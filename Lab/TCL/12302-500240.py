import os
import openpyxl
import secrets
import shutil
from datetime import datetime
import win32com.client
import time


# ===== 配置区域 =====
class Config:
    """程序配置类，集中管理所有可配置参数"""
    # 文件路径配置
    SOURCE_DIR = r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\3-周期验证\2025年\TCL\12302-500240(310100062)\(310100108)模板"
    OUTPUT_DIR = r"E:\System\desktop\PY\实验室"
    # 销售明细Excel文件路径
    SALES_DETAIL_FILE = r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\3-周期验证\TCL销售明细.xlsx"

    # PDF输出配置
    PDF_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "PDF输出")

    # 随机数生成范围配置 (最小值, 最大值, 最小差值)
    # B_C范围: 确保B列值大于C列值，差值至少为min_diff
    RANGE_CONFIG = {
        'B_C': (140, 150, 2.1),  # B列和C列范围
        'D_E': (5.8, 6.125, 0.12),  # D列和E列范围
        'F_G': (72.3, 74.7, 0.12),  # F列和G列范围
        'H_I': (3.8, 6.9, 0.81),  # H列和I列范围
    }

    # 数据填充区域（行范围）
    ROW_START = 12
    ROW_END = 16  # 包含此行

    # 文件匹配条件
    FILE_FILTERS = {
        'extensions': ['.xlsx', '.xls'],
        'keywords': ['310100108', '模板']
    }


# ===== 功能函数 =====
def generate_random_numbers(existing_values, value_range, ensure_first_larger=False):
    """
    生成两个不重复的随机数，可配置确保第一个数大于第二个数

    参数:
        existing_values (set): 已存在的值集合，用于避免重复
        value_range (tuple): 范围配置 (最小值, 最大值, 最小差值)
        ensure_first_larger (bool): 是否确保第一个数大于第二个数

    返回:
        tuple: 两个不重复的随机数
    """
    min_val, max_val, min_diff = value_range
    max_attempts = 100

    for _ in range(max_attempts):
        # 生成两个随机数
        value1 = round(secrets.SystemRandom().uniform(min_val, max_val), 3)
        value2 = round(secrets.SystemRandom().uniform(min_val, max_val), 3)

        # 确保两个数的差值符合要求
        if abs(value1 - value2) < min_diff:
            continue

        # 如果需要确保第一个数大于第二个数
        if ensure_first_larger and value1 <= value2:
            value1, value2 = value2, value1  # 交换值

        # 检查是否有重复
        if value1 not in existing_values and value2 not in existing_values:
            return value1, value2

    raise Exception("无法在100次尝试内生成不重复的随机数")


def process_excel_file(file_path, output_dir, order_date, order_number, config):
    """
    处理单个Excel文件：填充随机数并转换为PDF

    参数:
        file_path (str): 源Excel文件路径
        output_dir (str): 输出目录
        order_date (str): 订单日期
        order_number (str): 订单编号
        config (Config): 配置对象

    返回:
        bool: 处理是否成功
    """
    try:
        # 打开Excel工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet = workbook.active

        # 写入订单信息
        sheet['G2'] = order_date
        sheet['L2'] = order_number

        # 用于存储已生成的值，确保不重复
        existing_values = set()

        # 填充随机数到指定区域
        for row in range(config.ROW_START, config.ROW_END + 1):
            max_attempts = 100  # 最大尝试次数
            for attempt in range(max_attempts):
                # 临时集合，用于验证当前尝试的所有值
                temp_values = set(existing_values)

                # 生成所有列值
                try:
                    # B列和C列（C < B）
                    value_b, value_c = generate_random_numbers(temp_values, config.RANGE_CONFIG['B_C'],
                                                               ensure_first_larger=True)
                    temp_values.update([value_b, value_c])

                    # D列和E列（E > D）
                    value_d, value_e = generate_random_numbers(temp_values, config.RANGE_CONFIG['D_E'],
                                                               ensure_first_larger=False)
                    temp_values.update([value_d, value_e])

                    # F列和G列（G < F）
                    value_f, value_g = generate_random_numbers(temp_values, config.RANGE_CONFIG['F_G'],
                                                               ensure_first_larger=True)
                    temp_values.update([value_f, value_g])

                    # H列和I列（I > H）
                    value_h, value_i = generate_random_numbers(temp_values, config.RANGE_CONFIG['H_I'],
                                                               ensure_first_larger=False)
                    temp_values.update([value_h, value_i])

                    # 验证所有条件
                    if (value_c < value_b and
                            value_e > value_d and
                            value_g < value_f and
                            value_i > value_h):
                        # 条件全部满足，更新existing_values并跳出循环
                        existing_values.update(temp_values)
                        break

                except Exception as e:
                    # 生成失败，继续尝试
                    pass

                if attempt == max_attempts - 1:
                    raise Exception(f"行 {row}: 无法在{max_attempts}次尝试内生成满足所有条件的随机数")

            # 写入数据到对应单元格
            sheet[f'B{row}'] = value_b  # B列值（较大值）
            sheet[f'C{row}'] = value_c  # C列值（较小值）
            sheet[f'D{row}'] = value_d  # D列值（较小值）
            sheet[f'E{row}'] = value_e  # E列值（较大值）
            sheet[f'F{row}'] = value_f  # F列值（较大值）
            sheet[f'G{row}'] = value_g  # G列值（较小值）
            sheet[f'H{row}'] = value_h  # H列值（较小值）
            sheet[f'I{row}'] = value_i  # I列值（较大值）

            # 更新已存在的值集合
            existing_values.update([value_b, value_c, value_d, value_e, value_f, value_g, value_h, value_i])

        # 保存修改后的Excel文件
        os.makedirs(output_dir, exist_ok=True)
        file_name = os.path.basename(file_path)
        new_name = file_name.replace("模板", f"_{order_number}")
        output_file_path = os.path.join(output_dir, new_name)
        workbook.save(output_file_path)
        print(f"成功处理Excel: {file_name} -> {new_name}")

        # 转换为PDF
        pdf_output_dir = os.path.join(config.PDF_OUTPUT_DIR, os.path.relpath(output_dir, config.OUTPUT_DIR))
        pdf_path = os.path.join(pdf_output_dir, os.path.splitext(new_name)[0] + ".pdf")

        if excel_to_pdf(output_file_path, pdf_path):
            print(f"成功转换为PDF: {pdf_path}")
            return True
        else:
            print(f"PDF转换失败: {output_file_path}")
            return False

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}")
        return False


def excel_to_pdf(excel_path, pdf_path):
    """
    使用Excel COM接口将Excel文件转换为PDF

    参数:
        excel_path (str): Excel文件路径
        pdf_path (str): PDF输出路径

    返回:
        bool: 转换是否成功
    """
    try:
        # 创建Excel应用实例
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # 打开工作簿
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))

        # 创建PDF输出目录
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

        # 导出为PDF（所有工作表）
        workbook.ExportAsFixedFormat(0, pdf_path)

        # 关闭工作簿和Excel应用
        workbook.Close()
        excel.Quit()

        # 释放COM对象
        del workbook
        del excel

        return True
    except Exception as e:
        print(f"Excel转PDF失败: {excel_path} -> {pdf_path}, 错误: {e}")
        return False
    finally:
        # 确保资源被释放
        time.sleep(1)  # 等待Excel完全退出


def get_input_pairs(config):
    """
    从销售明细Excel文件中获取日期和订单编号对
    条件：物料编码对应实发数量 > 6000

    参数:
        config (Config): 配置对象

    返回:
        list: 包含元组 (日期, 订单编号) 的列表
    """
    pairs = []

    try:
        if not os.path.exists(config.SALES_DETAIL_FILE):
            print(f"错误: 销售明细文件不存在 - {config.SALES_DETAIL_FILE}")
            return pairs

        # 打开销售明细Excel文件
        workbook = openpyxl.load_workbook(config.SALES_DETAIL_FILE, data_only=True)
        sheet = workbook.active

        # 获取表头行，确定各列索引
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        date_col = None
        order_col = None
        material_col = None
        quantity_col = None

        # 查找各列对应的索引
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_value = str(cell_value).strip().lower()
            if '日期' in cell_value:
                date_col = idx
            elif '单据编号' in cell_value:
                order_col = idx
            elif '物料编码' in cell_value:
                material_col = idx
            elif '实发数量' in cell_value:
                quantity_col = idx

        # 检查是否找到了所有需要的列
        if any(col is None for col in [date_col, order_col, material_col, quantity_col]):
            missing = [col_name for col_name, col_idx in
                       [('日期', date_col), ('单据编号', order_col), ('物料编码', material_col),
                        ('实发数量', quantity_col)]
                       if col_idx is None]
            print(f"错误: 在销售明细文件中找不到以下列: {', '.join(missing)}")
            return pairs

        # 从第二行开始遍历数据行
        processed_orders = set()  # 用于记录已处理的订单编号，避免重复
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not any(row):
                continue

            # 获取各列值
            order_date = row[date_col]
            order_number = row[order_col]
            # 实发数量需要转换为数值类型
            try:
                quantity = float(row[quantity_col]) if row[quantity_col] is not None else 0
            except (ValueError, TypeError):
                quantity = 0

            # 检查条件：实发数量 > 6000 且订单编号未处理过
            if quantity > 6000 and order_number not in processed_orders:
                # 处理日期格式
                if isinstance(order_date, datetime):
                    formatted_date = order_date.strftime('%Y-%m-%d')
                else:
                    # 尝试解析字符串日期
                    try:
                        if isinstance(order_date, str):
                            # 处理常见日期格式
                            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y%m%d'):
                                try:
                                    date_obj = datetime.strptime(order_date, fmt)
                                    formatted_date = date_obj.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                            else:
                                print(f"警告: 无法解析日期格式 '{order_date}'，使用原始值")
                                formatted_date = str(order_date)
                        else:
                            formatted_date = str(order_date)
                    except Exception as e:
                        print(f"警告: 日期处理错误 '{order_date}': {e}，使用原始值")
                        formatted_date = str(order_date)

                pairs.append((formatted_date, order_number))
                processed_orders.add(order_number)
                print(f"已添加: {formatted_date} {order_number}")

        workbook.close()
        print(f"从销售明细文件中提取了 {len(pairs)} 个符合条件的订单")

        if not pairs:
            print("警告: 未找到实发数量大于6000的记录")

        return pairs

    except Exception as e:
        print(f"读取销售明细文件时出错: {e}")
        return pairs


def get_excel_files(config):
    """
    根据配置获取符合条件的Excel文件列表

    参数:
        config (Config): 配置对象

    返回:
        list: 符合条件的文件路径列表
    """
    excel_files = []
    if not os.path.exists(config.SOURCE_DIR):
        print(f"错误: 源目录不存在 - {config.SOURCE_DIR}")
        return excel_files

    for root, _, files in os.walk(config.SOURCE_DIR):
        for file in files:
            # 检查文件扩展名
            if not any(file.lower().endswith(ext) for ext in config.FILE_FILTERS['extensions']):
                continue
            # 检查关键词
            if not all(keyword in file for keyword in config.FILE_FILTERS['keywords']):
                continue
            excel_files.append(os.path.join(root, file))

    return excel_files


def main():
    """程序主入口"""
    # 创建配置实例
    config = Config()

    print(f"\n使用配置:")
    print(f"  源目录: {config.SOURCE_DIR}")
    print(f"  输出目录: {config.OUTPUT_DIR}")
    print(f"  PDF输出目录: {config.PDF_OUTPUT_DIR}")
    print(f"  销售明细文件: {config.SALES_DETAIL_FILE}")

    # 从销售明细Excel文件获取日期和订单编号对
    input_pairs = get_input_pairs(config)
    if not input_pairs:
        print("未找到符合条件的数据，程序退出")
        return

    # 获取符合条件的Excel文件
    excel_files = get_excel_files(config)
    if not excel_files:
        print("未找到符合条件的Excel文件")
        return

    print(f"找到 {len(excel_files)} 个符合条件的文件")

    # 批量处理文件
    for order_date, order_number in input_pairs:
        print(f"\n处理订单: {order_date} {order_number}")
        success_count = 0

        for file_path in excel_files:
            if process_excel_file(file_path, config.OUTPUT_DIR, order_date, order_number, config):
                success_count += 1

        print(f"订单 {order_number} 处理完成: 成功 {success_count} 个, 失败 {len(excel_files) - success_count} 个")


if __name__ == "__main__":
    main()