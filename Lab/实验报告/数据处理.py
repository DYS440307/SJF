import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import time
import numpy as np
import os

# 文件路径配置
CONFIG_DIR = r"E:\System\pic\A报告\模板\配置文件"
LAB_RECORD_FILE = r"E:\System\pic\A报告\老化实验记录.xlsx"


def read_config(file_path):
    """读取配置文件并返回配置字典"""
    config = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                # 跳过注释行和空行
                if line.startswith('#') or not line:
                    continue
                # 解析配置项
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip()
        return config
    except Exception as e:
        print(f"读取配置文件时发生错误: {e}")
        return None


def parse_config(config):
    """解析配置字典并返回结构化配置，不使用默认值"""
    if not config:
        raise ValueError("配置字典为空")

    # 定义配置参数元数据（不包含默认值）
    CONFIG_METADATA = {
        # 基本配置
        'TARGET_VALUE': {'type': int},
        'FILE_PATH': {'type': str},
        'A_RANGE_LOW': {'type': int},
        'A_RANGE_HIGH': {'type': int},
        'FIND_MAX': {'type': lambda x: x.lower() == 'true'},

        # SPL原档配置
        'SPL_MODE': {'type': str},
        'SPL_FIXED_TARGETS': {'type': lambda x: [int(v) for v in x.split(',')]},
        'SPL_RANGE_STEP': {'type': int},
        'SPL_CUSTOM_TARGETS': {'type': lambda x: [int(v) for v in x.split(',')]},
        'SPL_RANGE_LOW': {'type': int},
        'SPL_RANGE_HIGH': {'type': int},

        # THD原档配置
        'THD_A_RANGE_LOW': {'type': int},
        'THD_A_RANGE_HIGH': {'type': int}
    }

    # 自动解析配置，不使用默认值
    parsed_config = {}
    for key, meta in CONFIG_METADATA.items():
        if key not in config:
            raise ValueError(f"配置文件中缺少必需的参数: {key}")

        try:
            parsed_config[key] = meta['type'](config[key])
        except Exception as e:
            raise ValueError(f"解析配置项 '{key}' 时发生错误: {e}") from e

    return parsed_config


def validate_config(config):
    """验证配置是否完整且有效"""
    required_keys = [
        'TARGET_VALUE', 'FILE_PATH', 'A_RANGE_LOW', 'A_RANGE_HIGH', 'FIND_MAX',
        'SPL_MODE', 'SPL_FIXED_TARGETS', 'SPL_RANGE_STEP', 'SPL_CUSTOM_TARGETS',
        'SPL_RANGE_LOW', 'SPL_RANGE_HIGH', 'THD_A_RANGE_LOW', 'THD_A_RANGE_HIGH'
    ]

    # 检查是否缺少必需的键
    missing_keys = [key for key in required_keys if key not in config]
    if missing_keys:
        raise ValueError(f"配置文件缺少以下必需参数: {', '.join(missing_keys)}")

    # 验证数值范围
    if config['A_RANGE_LOW'] >= config['A_RANGE_HIGH']:
        raise ValueError("A_RANGE_LOW必须小于A_RANGE_HIGH")

    if config['SPL_RANGE_LOW'] >= config['SPL_RANGE_HIGH']:
        raise ValueError("SPL_RANGE_LOW必须小于SPL_RANGE_HIGH")

    if config['THD_A_RANGE_LOW'] >= config['THD_A_RANGE_HIGH']:
        raise ValueError("THD_A_RANGE_LOW必须小于THD_A_RANGE_HIGH")

    # 验证SPL_MODE值
    valid_spl_modes = ['FIXED', 'RANGE', 'CUSTOM', 'RANGE_ALL']
    if config['SPL_MODE'] not in valid_spl_modes:
        raise ValueError(
            f"无效的SPL_MODE值: {config['SPL_MODE']}，必须是{'FIXED', 'RANGE', 'CUSTOM', 'RANGE_ALL'}中的一个")

    return True


def find_nearest_value(df, target):
    # 确保数据是数值类型
    df = pd.to_numeric(df.iloc[:, 0], errors='coerce').dropna().reset_index(drop=True)

    if df.empty:
        return None

    left, right = 0, len(df) - 1
    nearest_idx = 0
    min_diff = float('inf')

    while left <= right:
        mid = (left + right) // 2
        current_diff = abs(df.iloc[mid] - target)

        if current_diff < min_diff:
            min_diff = current_diff
            nearest_idx = mid

        if df.iloc[mid] < target:
            left = mid + 1
        elif df.iloc[mid] > target:
            right = mid - 1
        else:
            break

    return df.iloc[nearest_idx]


def generate_spl_targets(spl_mode, a_range_low, a_range_high, step, fixed_targets, custom_targets, range_low,
                         range_high):
    """根据配置生成SPL目标值列表"""
    if spl_mode == "FIXED":
        return fixed_targets
    elif spl_mode == "RANGE":
        return list(range(a_range_low, a_range_high + 1, step))
    elif spl_mode == "CUSTOM":
        return custom_targets
    elif spl_mode == "RANGE_ALL":
        print(f"警告: RANGE_ALL模式将处理A列中所有在{range_low}~{range_high}范围内的值")
        return None  # 返回None表示处理范围内的所有值
    else:
        print(f"警告: 未知的SPL_MODE值 '{spl_mode}'，使用默认的固定目标值")
        return fixed_targets


def find_config_file(report_id):
    """在老化实验记录中查找报告编号并返回对应的配置文件路径"""
    try:
        # 读取老化实验记录文件
        print(f"正在读取老化实验记录文件: {LAB_RECORD_FILE}")
        df = pd.read_excel(LAB_RECORD_FILE)

        # 假设报告编号在第N列，这里遍历所有列查找
        found = False
        report_col = None

        # 从最后一行开始向上查找
        for col in df.columns:
            for row_idx in reversed(df.index):
                cell_value = df.at[row_idx, col]
                if pd.notna(cell_value) and str(report_id) in str(cell_value):
                    report_col = col
                    found = True
                    break
            if found:
                break

        if not found:
            raise ValueError(f"在老化实验记录中未找到报告编号: {report_id}")

        print(f"找到报告编号 '{report_id}' 在列: {report_col}")

        # 获取对应的第I列单元格内容
        i_col_idx = 8  # 假设I列是第9列(索引为8)，根据实际情况调整
        if i_col_idx >= len(df.columns):
            raise ValueError(f"未找到第I列数据")

        cell_value = df.at[row_idx, df.columns[i_col_idx]]
        if pd.isna(cell_value):
            raise ValueError(f"第I列对应单元格内容为空")

        # 分割单元格内容并移除TCL
        parts = [part.strip() for part in str(cell_value).split(';') if part.strip() != 'TCL']

        if not parts:
            raise ValueError(f"第I列单元格内容分割后没有有效关键字")

        print(f"从第I列获取的有效关键字: {parts}")

        # 在配置文件目录中查找匹配的配置文件
        config_files = os.listdir(CONFIG_DIR)
        matched_files = []

        # 为每个关键字查找匹配的配置文件
        for part in parts:
            part_matched = False
            # 尝试匹配包含所有关键字的单个配置文件
            combined_keywords = part
            for file in config_files:
                if all(keyword.strip() in file for keyword in combined_keywords.split('；')) and file.endswith('.txt'):
                    matched_files.append(os.path.join(CONFIG_DIR, file))
                    part_matched = True
                    print(f"  找到匹配的配置文件: {file}，匹配关键字: {combined_keywords}")
                    break  # 找到一个匹配项后就退出循环

            # 如果没有找到组合匹配，则尝试单独匹配每个关键字
            if not part_matched:
                for keyword in combined_keywords.split('；'):
                    keyword = keyword.strip()
                    if not keyword:
                        continue
                    for file in config_files:
                        if keyword in file and file.endswith('.txt'):
                            matched_files.append(os.path.join(CONFIG_DIR, file))
                            part_matched = True
                            print(f"  找到匹配的配置文件: {file}，匹配关键字: {keyword}")
                            break  # 找到一个匹配项后就退出循环
                    if part_matched:
                        break  # 找到一个关键字匹配后就不再检查其他关键字

            if not part_matched:
                print(f"  警告: 未找到与关键字 '{part}' 匹配的配置文件")

        # 移除重复的文件路径
        matched_files = list(dict.fromkeys(matched_files))

        if not matched_files:
            raise ValueError(f"未找到匹配的配置文件，关键字: {parts}")

        print(f"找到 {len(matched_files)} 个匹配的配置文件")
        for file in matched_files:
            print(f"  - {file}")

        # 返回第一个匹配的配置文件
        return matched_files[0]

    except Exception as e:
        print(f"查找配置文件时发生错误: {e}")
        return None


try:
    # 获取用户输入的报告编号
    report_id = input("请输入报告编号: ").strip()
    if not report_id:
        raise ValueError("报告编号不能为空")

    print(f"正在查找报告编号 '{report_id}' 对应的配置文件...")

    # 查找配置文件
    config_file_path = find_config_file(report_id)
    if not config_file_path:
        raise ValueError("无法找到匹配的配置文件")

    print(f"使用配置文件: {config_file_path}")

    # 读取配置文件
    config = read_config(config_file_path)
    if not config:
        raise ValueError("无法读取配置文件或配置文件为空")

    # 解析配置
    config = parse_config(config)

    # 验证配置
    validate_config(config)

    print(f"开始执行Excel数据处理脚本")
    print(f"配置参数:")
    for key, value in config.items():
        print(f"  {key}: {value}")

    # 生成SPL目标值列表
    spl_targets = generate_spl_targets(
        config['SPL_MODE'],
        config['A_RANGE_LOW'],
        config['A_RANGE_HIGH'],
        config['SPL_RANGE_STEP'],
        config['SPL_FIXED_TARGETS'],
        config['SPL_CUSTOM_TARGETS'],
        config['SPL_RANGE_LOW'],
        config['SPL_RANGE_HIGH']
    )
    print(f"  SPL查找模式: {config['SPL_MODE']}")
    if config['SPL_MODE'] == "RANGE_ALL":
        print(f"  SPL查找范围: {config['SPL_RANGE_LOW']}~{config['SPL_RANGE_HIGH']} (处理范围内的所有值)")
    else:
        print(f"  SPL查找目标: {spl_targets} (共{len(spl_targets)}个值)")

    print(f"  THD原档A列范围: {config['THD_A_RANGE_LOW']}~{config['THD_A_RANGE_HIGH']}")

    start_time = time.time()

    # 读取Excel文件
    print(f"正在读取Excel文件...")
    excel_file = pd.ExcelFile(config['FILE_PATH'])

    # 获取IMP原档中的数据
    print(f"正在解析'IMP原档'工作表...")
    df = excel_file.parse("IMP原档")

    # 检查IMP原档中有数值的列数是否为偶数
    non_empty_columns = df.dropna(axis=1, how='all').shape[1]

    if non_empty_columns % 2 != 0:
        raise ValueError(f"IMP原档中有数值的列数为{non_empty_columns}，必须为偶数")

    # 使用openpyxl将值写入ACR表
    print(f"准备写入'ACR'工作表...")
    wb = openpyxl.load_workbook(config['FILE_PATH'])
    ws = wb["ACR"]

    # 处理所有奇数列(1,3,5...)及其后续偶数列(2,4,6...)
    print(f"正在处理列对数据...")
    max_col = df.shape[1]
    processed_pairs = 0
    skipped_pairs = 0

    for col_pair in range(0, max_col, 2):
        # 检查是否有足够的列
        if col_pair + 1 >= max_col:
            break

        # 获取当前奇数列和偶数列
        odd_col = col_pair
        even_col = col_pair + 1

        # 获取列字母表示（用于日志输出）
        col_letter_odd = get_column_letter(odd_col + 1)
        col_letter_even = get_column_letter(even_col + 1)

        print(f"  处理列对: {col_letter_odd}&{col_letter_even}")

        # 提取当前奇数列和偶数列的数据
        current_df = df.iloc[:, [odd_col, even_col]].copy()

        # 转换为数值类型并删除非数值
        current_df.iloc[:, 0] = pd.to_numeric(current_df.iloc[:, 0], errors='coerce')
        current_df.iloc[:, 1] = pd.to_numeric(current_df.iloc[:, 1], errors='coerce')
        current_df = current_df.dropna()

        # 跳过空列对
        if current_df.empty:
            print(f"    跳过空列对: {col_letter_odd}&{col_letter_even}")
            skipped_pairs += 1
            continue

        # 找到奇数列最接近目标值的值对应的偶数列的值
        nearest_value = find_nearest_value(current_df, config['TARGET_VALUE'])

        # 如果找不到合适的值，则跳过
        if nearest_value is None:
            print(f"    在列 {col_letter_odd} 中未找到合适的值")
            skipped_pairs += 1
            continue

        # 获取对应偶数列的值
        nearest_row = current_df.iloc[(current_df.iloc[:, 0] - config['TARGET_VALUE']).abs().argsort()[:1]]
        value_to_write = nearest_row.iloc[0, 1]

        # 计算ACR表中的行号(从1开始，每对占一行)
        row_in_acr = (col_pair // 2) + 1

        # 将值写入ACR表的对应行的第一列
        ws.cell(row=row_in_acr, column=1).value = value_to_write
        processed_pairs += 1

    print(f"列对数据处理完成: 已处理 {processed_pairs} 对, 跳过 {skipped_pairs} 对")

    # 读取ACR表中的所有数据
    print(f"正在分析'ACR'工作表数据...")
    acr_data = []
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None:
            acr_data.append(cell_value)

    # 计算分割点
    split_index = len(acr_data) // 2

    # 将后一半数据移至B列的起始行，并清空A列对应位置
    print(f"正在重新排列'ACR'工作表数据...")
    moved_values = 0

    for i in range(split_index, len(acr_data)):
        source_row = i + 1
        target_row = (i - split_index) + 1
        ws.cell(row=target_row, column=2).value = acr_data[i]
        ws.cell(row=source_row, column=1).value = None  # 清空A列原数据
        moved_values += 1

    print(f"数据重新排列完成: 已移动 {moved_values} 个值到B列")

    # 比较AB两列相邻数据，确保B列数值大于A列
    print(f"正在验证'ACR'工作表AB列数据关系...")
    swap_count = 0
    max_compare_row = max(ws.max_row, len(acr_data) - split_index)

    for row in range(1, max_compare_row + 1):
        a_value = ws.cell(row=row, column=1).value
        b_value = ws.cell(row=row, column=2).value

        # 尝试转换为数值
        try:
            a_value = float(a_value) if a_value is not None else None
            b_value = float(b_value) if b_value is not None else None
        except (ValueError, TypeError):
            continue

        # 确保两个单元格都有数值
        if a_value is not None and b_value is not None:
            # 如果B列值小于等于A列值，则交换
            if b_value <= a_value:
                ws.cell(row=row, column=1).value = b_value
                ws.cell(row=row, column=2).value = a_value
                swap_count += 1

    print(f"数据验证完成: 执行了 {swap_count} 次交换操作")

    # 处理Fb工作表 - 对所有偶数列执行相同操作
    print(f"正在处理'Fb'工作表数据...")

    # 创建或获取Fb工作表
    if "Fb" in wb.sheetnames:
        fb_sheet = wb["Fb"]
    else:
        fb_sheet = wb.create_sheet("Fb")

    # 清空Fb工作表中已有的数据
    for row in range(1, fb_sheet.max_row + 1):
        for col in range(1, fb_sheet.max_column + 1):
            fb_sheet.cell(row=row, column=col).value = None

    # 处理所有偶数列（B,D,F...）
    result_values = []

    for col in range(1, max_col, 2):
        # 获取列字母表示
        col_letter = get_column_letter(col + 1)

        # 获取当前偶数列和前一列(奇数列)的数据
        odd_col = col - 1
        even_col = col

        # 检查列索引是否有效
        if odd_col < 0 or even_col >= max_col:
            print(f"    跳过无效列索引: {odd_col}和{even_col}")
            continue

        # 提取数据
        current_df = df.iloc[:, [odd_col, even_col]].copy()

        # 转换为数值类型并删除非数值
        current_df.iloc[:, 0] = pd.to_numeric(current_df.iloc[:, 0], errors='coerce')
        current_df.iloc[:, 1] = pd.to_numeric(current_df.iloc[:, 1], errors='coerce')
        current_df = current_df.dropna()

        # 跳过空列
        if current_df.empty:
            print(f"    跳过空列对: {get_column_letter(odd_col + 1)}&{col_letter}")
            continue

        # 筛选奇数列在指定范围内的数据
        filtered_df = current_df[
            (current_df.iloc[:, 0] >= config['A_RANGE_LOW']) & (current_df.iloc[:, 0] <= config['A_RANGE_HIGH'])]

        if not filtered_df.empty:
            # 重置索引以确保索引连续
            filtered_df = filtered_df.reset_index(drop=True)

            # 确保筛选后的数据有足够的行
            if len(filtered_df) == 0:
                print(f"    跳过: {col_letter}列在范围内没有有效值")
                continue

            # 根据配置查找偶数列的最大值或最小值
            if config['FIND_MAX']:
                extreme_value = filtered_df.iloc[:, 1].max()
                extreme_type = "最大值"
            else:
                extreme_value = filtered_df.iloc[:, 1].min()
                extreme_type = "最小值"

            # 找到对应的行
            extreme_rows = filtered_df[filtered_df.iloc[:, 1] == extreme_value]

            # 如果有多个匹配值，取第一个
            if not extreme_rows.empty:
                extreme_row = extreme_rows.iloc[0]

                # 获取奇数列和偶数列的值
                odd_col_value = extreme_row.iloc[0]
                even_col_value = extreme_row.iloc[1]

                print(
                    f"  在{get_column_letter(odd_col + 1)}列范围 {config['A_RANGE_LOW']}~{config['A_RANGE_HIGH']} 内找到{col_letter}列的{extreme_type}: {even_col_value}")
                print(f"    对应的{get_column_letter(odd_col + 1)}列值为: {odd_col_value}")

                # 保存结果
                result_values.append(odd_col_value)
            else:
                print(f"    警告: 未找到符合条件的行")
        else:
            print(
                f"  在{get_column_letter(odd_col + 1)}列中未找到范围在 {config['A_RANGE_LOW']}~{config['A_RANGE_HIGH']} 之间的值")

    # 将结果按顺序写入Fb工作表的第一列
    if result_values:
        for i, value in enumerate(result_values):
            fb_sheet.cell(row=i + 1, column=1).value = value

        print(f"已将所有结果按顺序写入'Fb'工作表的第一列")

        # 将Fb工作表后一半数据移至B列
        total_results = len(result_values)
        split_index_fb = (total_results + 1) // 2  # 向上取整

        print(f"正在重新排列'Fb'工作表数据...")
        moved_values_fb = 0

        for i in range(split_index_fb, total_results):
            source_row = i + 1
            target_row = (i - split_index_fb) + 1
            fb_sheet.cell(row=target_row, column=2).value = result_values[i]
            fb_sheet.cell(row=source_row, column=1).value = None  # 清空A列原数据
            moved_values_fb += 1

        print(f"Fb表数据重新排列完成: 已移动 {moved_values_fb} 个值到B列")

        # 比较Fb表AB两列相邻数据，确保B列数值小于A列
        print(f"正在验证'Fb'工作表AB列数据关系...")
        swap_count_fb = 0
        max_row_fb = max(fb_sheet.max_row, split_index_fb)

        for row in range(1, max_row_fb + 1):
            a_value = fb_sheet.cell(row=row, column=1).value
            b_value = fb_sheet.cell(row=row, column=2).value

            # 尝试转换为数值
            try:
                a_value = float(a_value) if a_value is not None else None
                b_value = float(b_value) if b_value is not None else None
            except (ValueError, TypeError):
                continue

            # 确保两个单元格都有数值
            if a_value is not None and b_value is not None:
                # 如果B列值大于等于A列值，则交换
                if b_value >= a_value:
                    fb_sheet.cell(row=row, column=1).value = b_value
                    fb_sheet.cell(row=row, column=2).value = a_value
                    swap_count_fb += 1

        print(f"Fb表数据验证完成: 执行了 {swap_count_fb} 次交换操作")
    else:
        print(f"没有找到符合条件的数据，'Fb'工作表保持为空")

    # 处理SPL原档工作表
    print(f"正在处理'SPL原档'工作表数据...")

    try:
        # 获取SPL原档中的数据
        spl_df = excel_file.parse("SPL原档")

        # 创建或获取SPL工作表
        if "SPL" in wb.sheetnames:
            spl_sheet = wb["SPL"]
        else:
            spl_sheet = wb.create_sheet("SPL")

        # 清空SPL工作表中已有的数据
        for row in range(1, spl_sheet.max_row + 1):
            for col in range(1, spl_sheet.max_column + 1):
                spl_sheet.cell(row=row, column=col).value = None

        # 获取SPL原档的总列数
        total_columns = spl_df.shape[1]

        if total_columns < 2:
            print(f"SPL原档中至少需要两列数据")
        else:
            # 提取SPL原档A列数据
            spl_col_a = pd.to_numeric(spl_df.iloc[:, 0], errors='coerce').dropna()

            # 处理所有偶数列（B,D,F...）
            for col_idx in range(1, total_columns, 2):
                # 获取当前偶数列的数据
                even_col = pd.to_numeric(spl_df.iloc[:, col_idx], errors='coerce').dropna()

                # 合并A列和当前偶数列数据
                merged_df = pd.concat([spl_col_a, even_col], axis=1).dropna()
                merged_df.columns = ['A', 'Current']

                if not merged_df.empty:
                    # 获取当前列的字母表示(用于日志)
                    col_letter = get_column_letter(col_idx + 1)
                    print(f"  正在处理偶数列 {col_letter}...")

                    # 查找A列中所有在范围内的行，并记录对应的偶数列值
                    target_values = []

                    if config['SPL_MODE'] == "RANGE_ALL":
                        # 处理范围内的所有值
                        filtered_df = merged_df[
                            (merged_df['A'] >= config['SPL_RANGE_LOW']) & (merged_df['A'] <= config['SPL_RANGE_HIGH'])]

                        if not filtered_df.empty:
                            target_values = filtered_df['Current'].tolist()
                            print(
                                f"    在A列中找到{len(target_values)}个值在{config['SPL_RANGE_LOW']}~{config['SPL_RANGE_HIGH']}范围内")
                        else:
                            print(f"    在A列中未找到值在{config['SPL_RANGE_LOW']}~{config['SPL_RANGE_HIGH']}范围内")
                    else:
                        # 处理指定目标值列表
                        missing_targets = []

                        for target in spl_targets:
                            # 找到A列中等于目标值的所有行
                            target_rows = merged_df[merged_df['A'] == target]

                            if not target_rows.empty:
                                # 获取对应的偶数列值
                                current_values = target_rows['Current'].tolist()

                                # 如果有多个匹配值，取第一个
                                if current_values:
                                    target_values.append(current_values[0])
                                    print(f"    在A列中找到值 {target}，对应的{col_letter}列值为 {current_values[0]}")
                                else:
                                    print(f"    警告: 在A列中找到值 {target}，但对应的{col_letter}列值为空")
                                    missing_targets.append(target)
                            else:
                                # 如果没有找到完全匹配的值，使用最接近的值
                                print(f"    警告: 在A列中未找到值 {target}，尝试查找最接近的值...")
                                nearest_row = merged_df.iloc[(merged_df['A'] - target).abs().argsort()[:1]]

                                if not nearest_row.empty:
                                    a_value = nearest_row.iloc[0]['A']
                                    current_value = nearest_row.iloc[0]['Current']

                                    # 存储对应的当前列值
                                    target_values.append(current_value)
                                    print(
                                        f"    在A列中找到最接近的值 {a_value}，对应的{col_letter}列值为 {current_value}")
                                else:
                                    print(f"    警告: 在A列中未找到接近的值")
                                    missing_targets.append(target)

                    # 计算当前偶数列的平均值并写入SPL工作表
                    if target_values:
                        # 计算平均值
                        average_value = sum(target_values) / len(target_values)

                        # 记录缺失的目标值（如果有）
                        if config['SPL_MODE'] != "RANGE_ALL" and missing_targets:
                            print(f"    注意: 在{col_letter}列中未找到以下目标值: {missing_targets}")

                        # 确定写入位置（B列对应A1，D列对应A2，依此类推）
                        row_in_spl = (col_idx + 1) // 2
                        spl_sheet.cell(row=row_in_spl, column=1).value = average_value
                        print(
                            f"    已将{col_letter}列对应值的平均值 {average_value:.4f} 写入'SPL'工作表的A列第{row_in_spl}行")
                    else:
                        print(f"    没有找到符合条件的数据，无法计算平均值")
                else:
                    print(f"  偶数列 {col_letter} 与A列合并后的数据为空")

            print(f"SPL原档所有偶数列处理完成")

            # 读取SPL表A列中的所有数据
            print(f"正在分析'SPL'工作表数据...")
            spl_data = []
            max_row_spl = spl_sheet.max_row
            for row in range(1, max_row_spl + 1):
                cell_value = spl_sheet.cell(row=row, column=1).value
                if cell_value is not None:
                    spl_data.append(cell_value)

            # 计算分割点
            split_index_spl = len(spl_data) // 2

            # 将后一半数据移至B列的起始行，并清空A列对应位置
            print(f"正在重新排列'SPL'工作表数据...")
            moved_values_spl = 0

            for i in range(split_index_spl, len(spl_data)):
                source_row = i + 1
                target_row = (i - split_index_spl) + 1
                spl_sheet.cell(row=target_row, column=2).value = spl_data[i]
                spl_sheet.cell(row=source_row, column=1).value = None  # 清空A列原数据
                moved_values_spl += 1

            print(f"SPL表数据重新排列完成: 已移动 {moved_values_spl} 个值到B列")

            # 比较SPL表AB两列相邻数据，确保A列数值大于B列
            print(f"正在验证'SPL'工作表AB列数据关系...")
            swap_count_spl = 0
            max_compare_row_spl = max(spl_sheet.max_row, split_index_spl)

            for row in range(1, max_compare_row_spl + 1):
                a_value = spl_sheet.cell(row=row, column=1).value
                b_value = spl_sheet.cell(row=row, column=2).value

                # 尝试转换为数值
                try:
                    a_value = float(a_value) if a_value is not None else None
                    b_value = float(b_value) if b_value is not None else None
                except (ValueError, TypeError):
                    continue

                # 确保两个单元格都有数值
                if a_value is not None and b_value is not None:
                    # 如果A列值小于等于B列值，则交换
                    if a_value <= b_value:
                        spl_sheet.cell(row=row, column=1).value = b_value
                        spl_sheet.cell(row=row, column=2).value = a_value
                        swap_count_spl += 1

            print(f"SPL表数据验证完成: 执行了 {swap_count_spl} 次交换操作")

    except Exception as spl_e:
        print(f"处理'SPL原档'工作表时发生错误: {spl_e}")

    # 处理THD原档工作表
    print(f"正在处理'THD原档'工作表数据...")

    try:
        # 获取THD原档中的数据
        thd_df = excel_file.parse("THD原档")

        # 创建或获取THD工作表
        if "THD" in wb.sheetnames:
            thd_sheet = wb["THD"]
        else:
            thd_sheet = wb.create_sheet("THD")

        # 清空THD工作表中已有的数据
        for row in range(1, thd_sheet.max_row + 1):
            for col in range(1, thd_sheet.max_column + 1):
                thd_sheet.cell(row=row, column=col).value = None

        # 获取THD原档的总列数
        total_columns = thd_df.shape[1]

        if total_columns < 2:
            print(f"THD原档中至少需要两列数据")
        else:
            # 提取THD原档A列数据
            thd_col_a = pd.to_numeric(thd_df.iloc[:, 0], errors='coerce').dropna()

            # 处理所有偶数列（B,D,F...）
            for col_idx in range(1, total_columns, 2):
                # 获取当前偶数列的数据
                even_col = pd.to_numeric(thd_df.iloc[:, col_idx], errors='coerce').dropna()

                # 合并A列和当前偶数列数据
                merged_df = pd.concat([thd_col_a, even_col], axis=1).dropna()
                merged_df.columns = ['A', 'Current']

                if not merged_df.empty:
                    # 获取当前列的字母表示(用于日志)
                    col_letter = get_column_letter(col_idx + 1)
                    print(f"  正在处理偶数列 {col_letter}...")

                    # 筛选A列在指定范围内的数据
                    filtered_df = merged_df[
                        (merged_df['A'] >= config['THD_A_RANGE_LOW']) & (merged_df['A'] <= config['THD_A_RANGE_HIGH'])]

                    if not filtered_df.empty:
                        # 找到偶数列的最大值
                        max_value = filtered_df['Current'].max()

                        # 找到对应的A列值
                        max_row = filtered_df[filtered_df['Current'] == max_value].iloc[0]
                        a_value = max_row['A']

                        print(
                            f"    在A列范围 {config['THD_A_RANGE_LOW']}~{config['THD_A_RANGE_HIGH']} 内找到{col_letter}列的最大值: {max_value}")
                        print(f"    对应的A列值为: {a_value}")

                        # 确定写入位置（B列对应A1，D列对应A2，依此类推）
                        row_in_thd = (col_idx + 1) // 2
                        thd_sheet.cell(row=row_in_thd, column=1).value = max_value
                    else:
                        print(
                            f"    在A列中未找到范围在 {config['THD_A_RANGE_LOW']}~{config['THD_A_RANGE_HIGH']} 之间的值")
                else:
                    print(f"  偶数列 {col_letter} 与A列合并后的数据为空")

            print(f"THD原档所有偶数列处理完成")

            # 读取THD表A列中的所有数据
            print(f"正在分析'THD'工作表数据...")
            thd_data = []
            max_row_thd = thd_sheet.max_row
            for row in range(1, max_row_thd + 1):
                cell_value = thd_sheet.cell(row=row, column=1).value
                if cell_value is not None:
                    thd_data.append(cell_value)

            # 计算分割点
            split_index_thd = len(thd_data) // 2

            # 将后一半数据移至B列的起始行，并清空A列对应位置
            print(f"正在重新排列'THD'工作表数据...")
            moved_values_thd = 0

            for i in range(split_index_thd, len(thd_data)):
                source_row = i + 1
                target_row = (i - split_index_thd) + 1
                thd_sheet.cell(row=target_row, column=2).value = thd_data[i]
                thd_sheet.cell(row=source_row, column=1).value = None  # 清空A列原数据
                moved_values_thd += 1

            print(f"THD表数据重新排列完成: 已移动 {moved_values_thd} 个值到B列")

            # 比较THD表AB两列相邻数据，确保B列数值大于A列
            print(f"正在验证'THD'工作表AB列数据关系...")
            swap_count_thd = 0
            max_compare_row_thd = max(thd_sheet.max_row, split_index_thd)

            for row in range(1, max_compare_row_thd + 1):
                a_value = thd_sheet.cell(row=row, column=1).value
                b_value = thd_sheet.cell(row=row, column=2).value

                # 尝试转换为数值
                try:
                    a_value = float(a_value) if a_value is not None else None
                    b_value = float(b_value) if b_value is not None else None
                except (ValueError, TypeError):
                    continue

                # 确保两个单元格都有数值
                if a_value is not None and b_value is not None:
                    # 如果B列值小于等于A列值，则交换
                    if b_value <= a_value:
                        thd_sheet.cell(row=row, column=1).value = b_value
                        thd_sheet.cell(row=row, column=2).value = a_value
                        swap_count_thd += 1

            print(f"THD表数据验证完成: 执行了 {swap_count_thd} 次交换操作")

    except Exception as thd_e:
        print(f"处理'THD原档'工作表时发生错误: {thd_e}")

    # 保存修改后的Excel文件
    print(f"正在保存修改后的Excel文件...")
    try:
        wb.save(config['FILE_PATH'])
        print(f"文件已成功保存到: {config['FILE_PATH']}")
    except Exception as save_e:
        print(f"保存文件时发生错误: {save_e}")

    end_time = time.time()
    print(f"脚本执行完成，耗时: {end_time - start_time:.2f}秒")

except Exception as e:
    print(f"执行脚本时发生错误: {e}")