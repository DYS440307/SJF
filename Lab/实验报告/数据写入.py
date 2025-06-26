import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 配置参数 - 方便修改
EXCEL_DIR = r"E:\System\pic\A报告"  # Excel文件所在目录
SEARCH_TEXT = "SYS"  # 文件名中要包含的文本
SOURCE_FILE = r"E:\System\pic\A报告\IMP数据.xlsx"  # 源数据文件路径
CONFIG_DIR = r"E:\System\pic\A报告\模板\配置文件"  # 配置文件目录
LAB_RECORD_FILE = r"Z:\3-品质部\实验室\邓洋枢\2-实验记录汇总表\2025年\老化实验记录.xlsx"  # 老化实验记录文件

# 数据映射：源工作表 -> (起始单元格)
SHEET_MAPPING = {
    "Fb": "B12",  # 源文件的Fb工作表数据复制到目标文件第一个工作表的B12起始
    "ACR": "D12",  # 源文件的ACR工作表数据复制到目标文件第一个工作表的D12起始
    "SPL": "F12",  # 源文件的SPL工作表数据复制到目标文件第一个工作表的F12起始
    "THD": "H12"  # 源文件的THD工作表数据复制到目标文件第一个工作表的H12起始
}

# 定义各列范围检查的基准单元格
COLUMN_RANGE_MAPPING = {
    'B': 'B10',  # B列和C列的数据检查B10的范围
    'D': 'D10',  # D列和E列的数据检查D10的范围
    'F': 'F10',  # F列和G列的数据检查F10的范围
    'H': 'H10'  # H列和I列的数据检查H10的范围
}

# 定义红色加粗字体样式
ERROR_FONT = Font(color="FF0000", bold=True)


def read_config(file_path):
    """读取配置文件并返回配置字典"""
    config = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                # 跳过注释行和空行
                if line.startswith('#') or not line:
                    continue
                # 解析配置项
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip()
        return config
    except Exception as e:
        print(f"读取配置文件时发生错误: {e}")
        return None


def parse_cell_data(config):
    """从配置中解析CELL_DATA"""
    cell_data = {}
    for key, value in config.items():
        if key.startswith('CELL_'):
            # 提取单元格地址，例如将CELL_B9转换为B9
            cell_address = key[5:]
            cell_data[cell_address] = value
    return cell_data


def parse_value_range(value_str):
    """解析值范围字符串，返回最小值和最大值"""
    try:
        if "±" in value_str:
            base, tolerance = value_str.split("±")
            base = float(base)
            if tolerance.endswith('%'):
                tolerance_value = base * float(tolerance.strip('%')) / 100
            else:
                tolerance_value = float(tolerance)
            return base - tolerance_value, base + tolerance_value
        elif "≤" in value_str:
            max_val = float(value_str.replace("≤", ""))
            return 0, max_val
        elif "≥" in value_str:
            min_val = float(value_str.replace("≥", ""))
            return min_val, float('inf')
        else:
            # 默认为精确值
            val = float(value_str)
            return val, val
    except Exception as e:
        print(f"解析值范围失败: {value_str}, 错误: {e}")
        return None, None


def validate_cell_value(cell_value, range_cell, cell_data):
    """验证单元格值是否在指定范围单元格定义的范围内"""
    # 如果单元格值不是数值类型，直接返回False
    if not isinstance(cell_value, (int, float)):
        return False

    # 获取范围单元格的值
    range_value_str = cell_data.get(range_cell)
    if not range_value_str:
        return False

    # 解析范围
    min_val, max_val = parse_value_range(range_value_str)
    if min_val is None or max_val is None:
        return False

    # 检查值是否在范围内
    return min_val <= cell_value <= max_val


def find_excel_files(directory, search_text):
    """查找目录中名称包含指定文本的Excel文件"""
    excel_files = []
    for filename in os.listdir(directory):
        if search_text in filename and filename.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(directory, filename))
    return excel_files


def find_config_file(report_id):
    """在老化实验记录中查找报告编号并返回对应的配置文件路径"""
    try:
        # 读取老化实验记录文件
        print(f"正在读取老化实验记录文件: {LAB_RECORD_FILE}")
        df = pd.read_excel(LAB_RECORD_FILE)

        # 假设报告编号在第N列，这里遍历所有列查找
        found = False
        report_col = None

        # 从最后一行开始向上查找
        for col in df.columns:
            for row_idx in reversed(df.index):
                cell_value = df.at[row_idx, col]
                if pd.notna(cell_value) and str(report_id) in str(cell_value):
                    report_col = col
                    found = True
                    break
            if found:
                break

        if not found:
            raise ValueError(f"在老化实验记录中未找到报告编号: {report_id}")

        print(f"找到报告编号 '{report_id}' 在列: {report_col}")

        # 获取对应的第I列单元格内容
        i_col_idx = 8  # 假设I列是第9列(索引为8)，根据实际情况调整
        if i_col_idx >= len(df.columns):
            raise ValueError(f"未找到第I列数据")

        cell_value = df.at[row_idx, df.columns[i_col_idx]]
        if pd.isna(cell_value):
            raise ValueError(f"第I列对应单元格内容为空")

        # 分割单元格内容并移除TCL
        parts = [part.strip() for part in str(cell_value).split(';') if part.strip() != 'TCL']

        if not parts:
            raise ValueError(f"第I列单元格内容分割后没有有效关键字")

        print(f"从第I列获取的有效关键字: {parts}")

        # 在配置文件目录中查找匹配的配置文件
        config_files = os.listdir(CONFIG_DIR)
        matched_files = []

        # 为每个关键字查找匹配的配置文件
        for part in parts:
            part_matched = False
            # 尝试匹配包含所有关键字的单个配置文件
            combined_keywords = part
            for file in config_files:
                if all(keyword.strip() in file for keyword in combined_keywords.split('；')) and file.endswith('.txt'):
                    matched_files.append(os.path.join(CONFIG_DIR, file))
                    part_matched = True
                    print(f"  找到匹配的配置文件: {file}，匹配关键字: {combined_keywords}")
                    break  # 找到一个匹配项后就退出循环

            # 如果没有找到组合匹配，则尝试单独匹配每个关键字
            if not part_matched:
                for keyword in combined_keywords.split('；'):
                    keyword = keyword.strip()
                    if not keyword:
                        continue
                    for file in config_files:
                        if keyword in file and file.endswith('.txt'):
                            matched_files.append(os.path.join(CONFIG_DIR, file))
                            part_matched = True
                            print(f"  找到匹配的配置文件: {file}，匹配关键字: {keyword}")
                            break  # 找到一个匹配项后就退出循环
                    if part_matched:
                        break  # 找到一个关键字匹配后就不再检查其他关键字

            if not part_matched:
                print(f"  警告: 未找到与关键字 '{part}' 匹配的配置文件")

        # 移除重复的文件路径
        matched_files = list(dict.fromkeys(matched_files))

        if not matched_files:
            raise ValueError(f"未找到匹配的配置文件，关键字: {parts}")

        print(f"找到 {len(matched_files)} 个匹配的配置文件")
        for file in matched_files:
            print(f"  - {file}")

        # 返回第一个匹配的配置文件
        return matched_files[0]

    except Exception as e:
        print(f"查找配置文件时发生错误: {e}")
        return None


def ensure_enough_rows(target_ws, source_ws, start_row):
    """确保目标工作表有足够的行来容纳源工作表的数据"""
    source_max_row = source_ws.max_row
    target_max_row = target_ws.max_row

    # 计算需要的行数
    required_rows = start_row + source_max_row - 1

    # 如果目标工作表的行数不足，则插入新行
    if target_max_row < required_rows:
        rows_to_insert = required_rows - target_max_row
        print(f"目标工作表需要增加 {rows_to_insert} 行以容纳源数据")
        target_ws.insert_rows(target_max_row + 1, amount=rows_to_insert)
        print(f"已成功增加 {rows_to_insert} 行")
    else:
        print(f"目标工作表已有足够的行 ({target_max_row} >= {required_rows})")


def copy_sheet_data(source_wb, source_sheet_name, target_wb, start_cell):
    """将源工作表的全部数据复制到目标工作簿的第一个工作表的指定位置"""
    # 检查源工作表是否存在
    if source_sheet_name not in source_wb:
        print(f"源文件中不存在工作表: {source_sheet_name}")
        return

    # 获取目标工作簿的第一个工作表
    target_ws = target_wb.active
    source_ws = source_wb[source_sheet_name]

    # 解析起始单元格（例如：B12 -> 列索引=2, 行索引=12）
    start_col_letter = start_cell[0]
    start_row = int(start_cell[1:])

    # 确保目标工作表有足够的行
    ensure_enough_rows(target_ws, source_ws, start_row)

    # 将列字母转换为数字索引（A=1, B=2, ...）
    start_col = ord(start_col_letter) - 64  # ASCII值减去64

    # 获取源工作表的最大行和列
    max_row = source_ws.max_row
    max_col = source_ws.max_column

    print(f"从源文件复制工作表 '{source_sheet_name}' 数据到目标文件的第一个工作表，起始位置: {start_cell}")
    print(f"源数据范围: 1-{max_row}行, A-{get_column_letter(max_col)}列")

    # 复制数据
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            # 计算目标单元格位置
            target_row = start_row + row_idx - 1
            target_col = start_col + col_idx - 1

            # 获取源单元格的值
            source_cell = source_ws.cell(row=row_idx, column=col_idx)
            value = source_cell.value

            # 对数值类型的数据保留三位小数
            if isinstance(value, (int, float)):
                # 使用Python的格式化字符串保留三位小数
                value = round(value, 3)
                # 设置Excel单元格的数字格式为三位小数
                target_ws.cell(row=target_row, column=target_col).number_format = '0.000'

            # 获取目标单元格并设置值
            target_ws.cell(row=target_row, column=target_col).value = value

    print(f"成功复制 {max_row} 行 {max_col} 列数据到起始位置 {start_cell}")


def write_to_excel(file_path, cell_data):
    """向Excel文件的指定单元格写入数据，直接覆盖原文件"""
    try:
        # 加载源工作簿
        source_wb = load_workbook(SOURCE_FILE, data_only=True)

        # 加载目标工作簿
        target_wb = load_workbook(file_path)

        # 写入固定数据
        target_ws = target_wb.active
        for cell, value in cell_data.items():
            target_ws[cell] = value

            # 跳过B9单元格的解析
            if cell == "B9":
                continue

            # 计算并写入范围
            row = cell[1:]
            col_letter = cell[0]
            next_col_letter = get_column_letter(ord(col_letter) + 1)

            min_val, max_val = parse_value_range(value)
            if min_val is not None and max_val is not None:
                if min_val == max_val:
                    range_str = f"{min_val}"
                else:
                    range_str = f"{min_val}~{max_val}"
                target_ws[f"{next_col_letter}{row}"] = range_str

                # 在控制台打印解析的范围
                print(f"单元格 {cell}: {value} -> 范围: {range_str}")
            else:
                # 处理无法解析的数值类型
                target_ws[f"{next_col_letter}{row}"] = "N/A"
                print(f"单元格 {cell}: {value} -> 范围: N/A (无法解析)")

        # 复制工作表数据
        for source_sheet_name, start_cell in SHEET_MAPPING.items():
            print(f"\n开始复制工作表 '{source_sheet_name}' 到 {start_cell}")
            copy_sheet_data(source_wb, source_sheet_name, target_wb, start_cell)

        # 验证并格式化数据
        print("\n开始验证数据范围并设置格式...")
        for col_base_letter in COLUMN_RANGE_MAPPING.keys():
            range_cell = COLUMN_RANGE_MAPPING[col_base_letter]
            print(
                f"验证 {col_base_letter} 和 {get_column_letter(ord(col_base_letter) + 1)} 列的数据，参考范围: {range_cell}")

            # 获取当前列的范围值
            range_value = cell_data.get(range_cell)
            if not range_value:
                print(f"警告: 未找到参考范围单元格 {range_cell} 的值")
                continue

            # 获取起始行和结束行（假设数据从第12行开始）
            start_row = 12
            end_row = target_ws.max_row

            # 验证当前列和下一列的数据
            for row in range(start_row, end_row + 1):
                # 验证当前列
                current_col_letter = col_base_letter
                cell = f"{current_col_letter}{row}"
                cell_value = target_ws[cell].value

                if isinstance(cell_value, (int, float)):
                    if not validate_cell_value(cell_value, range_cell, cell_data):
                        target_ws[cell].font = ERROR_FONT
                        print(f"警告: 单元格 {cell} 的值 {cell_value} 不在 {range_cell} 的范围内: {range_value}")

                # 验证下一列
                next_col_letter = get_column_letter(ord(current_col_letter) + 1)
                cell = f"{next_col_letter}{row}"
                cell_value = target_ws[cell].value

                if isinstance(cell_value, (int, float)):
                    if not validate_cell_value(cell_value, range_cell, cell_data):
                        target_ws[cell].font = ERROR_FONT
                        print(f"警告: 单元格 {cell} 的值 {cell_value} 不在 {range_cell} 的范围内: {range_value}")

        # 直接保存并覆盖原文件
        target_wb.save(file_path)
        print(f"成功写入数据到 {file_path}")
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}")


def main():
    # 获取用户输入的报告编号
    report_id = input("请输入报告编号: ").strip()
    if not report_id:
        print("报告编号不能为空")
        return

    print(f"正在查找报告编号 '{report_id}' 对应的配置文件...")

    # 查找配置文件
    config_file_path = find_config_file(report_id)
    if not config_file_path:
        print("无法找到匹配的配置文件")
        return

    print(f"使用配置文件: {config_file_path}")

    # 读取配置文件
    config = read_config(config_file_path)
    if not config:
        print("无法读取配置文件或配置文件为空")
        return

    # 解析CELL_DATA
    cell_data = parse_cell_data(config)
    if not cell_data:
        print("配置文件中未找到CELL_DATA配置")
        return

    print(f"从配置文件加载了 {len(cell_data)} 个单元格配置")
    for cell, value in cell_data.items():
        print(f"  {cell}: {value}")

    # 查找符合条件的Excel文件
    excel_files = find_excel_files(EXCEL_DIR, SEARCH_TEXT)

    if not excel_files:
        print(f"未找到名称包含 '{SEARCH_TEXT}' 的Excel文件")
        return

    print(f"找到 {len(excel_files)} 个Excel文件")

    # 处理每个Excel文件
    for file in excel_files:
        print(f"\n处理文件: {file}")
        write_to_excel(file, cell_data)


if __name__ == "__main__":
    main()