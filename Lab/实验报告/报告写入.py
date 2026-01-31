import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta


def find_report_data(report_id, lab_record_file):
    """在老化实验记录中查找报告ID对应的数据行"""
    try:
        print(f"正在读取老化实验记录: {lab_record_file}")
        df = pd.read_excel(lab_record_file)

        # 查找N列中匹配报告ID的行
        report_rows = df[df.iloc[:, 13].astype(str).str.strip() == report_id]

        if report_rows.empty:
            print(f"未找到报告编号为 '{report_id}' 的记录")
            return None

        if len(report_rows) > 1:
            print(f"警告: 找到多条报告编号为 '{report_id}' 的记录，仅使用第一条")

        return report_rows.iloc[0]

    except Exception as e:
        print(f"读取老化实验记录时出错: {e}")
        return None


def parse_i_column(i_value):
    """解析I列数据（格式：TCL；G0202-000313；310100108）"""
    if pd.isna(i_value):
        return [None, None, None]

    parts = str(i_value).split('；')
    parts += [None] * (3 - len(parts))  # 确保返回3个元素
    return parts[:3]


def parse_m_column(m_value):
    """解析M列数据（格式：8.75W；7.24V；直通；振幅冲击1;50°C；90%R.H）"""
    if pd.isna(m_value):
        return [None, None, None, None, None, None]

    # 处理中文和西文分号
    parts = str(m_value).replace(';', '；').split('；')
    parts += [None] * (6 - len(parts))  # 确保返回6个元素
    return parts[:6]


def parse_q_column(q_value):
    """解析Q列数据（格式：扬声器寿命测试系统-精深PS5018S（A组）；恒温恒湿箱-）"""
    if pd.isna(q_value):
        return [None, None]

    parts = str(q_value).split('；')
    parts += [None] * (2 - len(parts))  # 确保返回2个元素
    return parts[:2]


def calculate_time_difference(ws):
    """计算F4单元格的值（D4时间减去B4日期并转换为小时，只保留整数部分）"""
    try:
        # 获取B4单元格的日期值
        b4_value = ws['B4'].value
        # 获取D4单元格的时间值
        d4_value = ws['D4'].value

        # 验证数据类型
        if not isinstance(b4_value, (datetime, pd.Timestamp)):
            print(f"错误: B4单元格不是有效的日期格式 - {b4_value}")
            return

        if not isinstance(d4_value, (datetime, pd.Timestamp)):
            print(f"错误: D4单元格不是有效的时间格式 - {d4_value}")
            return

        # 确保日期时间包含日期部分
        if isinstance(d4_value, datetime) and d4_value.date() == datetime(1900, 1, 1).date():
            # 如果时间值没有日期信息，使用B4的日期
            d4_datetime = datetime.combine(b4_value.date(), d4_value.time())
        else:
            d4_datetime = d4_value

        # 计算时间差（小时）
        time_delta = d4_datetime - b4_value
        hours = int(time_delta.total_seconds() // 3600)  # 只保留整数部分

        # 将结果写入F4单元格，格式化为"XH"
        ws['F4'] = f"{hours}H"
        print(f"已计算时间差并写入F4单元格: {hours}H")

    except Exception as e:
        print(f"计算时间差时出错: {e}")


def write_to_report_template(report_data, template_file):
    """将提取的数据写入到试验报告模板中"""
    try:
        os.makedirs(os.path.dirname(template_file), exist_ok=True)
        print(f"正在打开试验报告模板: {template_file}")
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active  # 确保ws变量在使用前被正确定义

        # 基础数据映射
        cell_mapping = {
            'G2': 0,  # A列数据写入G2
            'B4': 1,  # B列数据写入B4
            'D4': 2,  # C列数据写入D4
            'H3': 7,  # H列数据写入H3
            'H4': 9,  # J列数据写入H4
            'J3': 10,  # K列数据写入J3
            'L3': 11,  # L列数据写入L3
            'L2': 13,  # N列数据写入L2
            'B2': 5  # F列数据写入B2
        }

        # 填充基础数据
        for cell, col_idx in cell_mapping.items():
            if pd.notna(report_data.iloc[col_idx]):
                # 特殊处理日期时间格式
                value = report_data.iloc[col_idx]
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()

                ws[cell] = value
                print(f"已将数据从列 {get_column_letter(col_idx + 1)} 写入到单元格 {cell}")
            else:
                print(f"列 {get_column_letter(col_idx + 1)} 数据为空，跳过单元格 {cell}")

        # 解析并写入I列数据
        i_parts = parse_i_column(report_data.iloc[8])  # I列索引为8
        i_mapping = {
            'F3': i_parts[0],  # TCL
            'B3': i_parts[1],  # G0202-000313
            'D3': i_parts[2]  # 310100108
        }

        for cell, value in i_mapping.items():
            if value is not None:
                ws[cell] = value
                print(f"已将I列数据 '{value}' 写入到单元格 {cell}")

        # 解析并写入M列数据
        m_parts = parse_m_column(report_data.iloc[12])  # M列索引为12
        m_mapping = {
            'B7': m_parts[0],  # 8.75W
            'C7': m_parts[1],  # 7.24V
            'D7': m_parts[2],  # 直通
            'E7': m_parts[3],  # 振幅冲击1
            'J4': m_parts[4],  # 50°C
            'L4': m_parts[5]  # 90%R.H
        }

        for cell, value in m_mapping.items():
            if value is not None:
                ws[cell] = value
                print(f"已将M列数据 '{value}' 写入到单元格 {cell}")

        # 新增：解析并写入Q列数据
        q_parts = parse_q_column(report_data.iloc[16])  # Q列索引为16
        q_mapping = {
            'B5': q_parts[0],  # 扬声器寿命测试系统-精深PS5018S（A组）
            'F5': q_parts[1]   # 恒温恒湿箱-
        }

        for cell, value in q_mapping.items():
            if value is not None:
                ws[cell] = value
                print(f"已将Q列数据 '{value}' 写入到单元格 {cell}")

        # 计算并写入时间差
        calculate_time_difference(ws)

        # 将B3和J3内容拼接后写入D1
        b3_value = ws['B3'].value or ''
        j3_value = ws['J3'].value or ''
        d1_value = f"{b3_value}{j3_value}试验报告"
        ws['D1'] = d1_value
        print(f"已将B3和J3内容拼接后写入D1单元格: {d1_value}")

        # 构建新的文件名：B3+L3+J3+L2，各部分之间用连字符连接
        b3_value = ws['B3'].value or ''
        l3_value = ws['L3'].value or ''
        j3_value = ws['J3'].value or ''
        l2_value = ws['L2'].value or ''

        # 拼接文件名，使用连字符分隔各部分
        file_name_parts = [b3_value, l3_value, j3_value, l2_value]
        file_name = "-".join(part for part in file_name_parts if part)  # 过滤空部分

        # 直接删除换行符
        file_name = file_name.replace('\n', '')

        # 增强非法字符过滤，处理其他隐藏字符
        invalid_chars = r'\/:*?"<>|\r\t'
        valid_filename = "".join(c for c in file_name if c not in invalid_chars)

        # 确保文件名不为空
        if not valid_filename:
            valid_filename = "未命名报告"

        # 修改工作表名称
        ws.title = valid_filename[:31]  # Excel工作表名称最长31个字符

        # 修改工作簿属性中的标题
        wb.properties.title = valid_filename

        # 修改文件保存名称
        output_dir = os.path.dirname(template_file)
        output_file = os.path.join(output_dir, f"{valid_filename}.xlsx")

        # 保存修改后的模板
        wb.save(output_file)
        print(f"试验报告已保存至: {output_file}")

    except Exception as e:
        print(f"写入试验报告时出错: {e}")


def main():
    # 文件路径配置
    LAB_RECORD_FILE = r"E:\System\pic\A报告\老化实验记录.xlsx"
    REPORT_TEMPLATE = r"E:\System\pic\A报告\试验报告.xlsx"

    # 获取用户输入的报告编号
    report_id = input("请输入报告编号: ").strip()
    if not report_id:
        print("报告编号不能为空")
        return

    # 查找报告数据
    report_data = find_report_data(report_id, LAB_RECORD_FILE)
    if report_data is None:
        return

    # 写入报告模板
    write_to_report_template(report_data, REPORT_TEMPLATE)

    print("报告生成完成!")


if __name__ == "__main__":
    main()