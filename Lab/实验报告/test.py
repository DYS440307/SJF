import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import time

# 配置参数 - 可直接修改以下值
TARGET_VALUE = 600  # ACR表查找目标值
FILE_PATH = r"E:\System\pic\A报告\IMP数据.xlsx"  # Excel文件路径
A_RANGE_LOW = 200  # A列范围下限
A_RANGE_HIGH = 400  # A列范围上限
FIND_MAX = True  # True: 查找B列最大值, False: 查找B列最小值
SPL_TARGETS = [200, 400, 500, 800]  # SPL原档查找目标值


def find_nearest_value(df, target):
    # 确保数据是数值类型
    df = pd.to_numeric(df.iloc[:, 0], errors='coerce').dropna().reset_index(drop=True)

    if df.empty:
        return None

    left, right = 0, len(df) - 1
    nearest_idx = 0
    min_diff = float('inf')

    while left <= right:
        mid = (left + right) // 2
        current_diff = abs(df.iloc[mid] - target)

        if current_diff < min_diff:
            min_diff = current_diff
            nearest_idx = mid

        if df.iloc[mid] < target:
            left = mid + 1
        elif df.iloc[mid] > target:
            right = mid - 1
        else:
            break

    return df.iloc[nearest_idx]


try:
    print(f"开始执行Excel数据处理脚本")
    print(f"配置参数:")
    print(f"  目标值: {TARGET_VALUE}")
    print(f"  文件路径: {FILE_PATH}")
    print(f"  A列范围: {A_RANGE_LOW}~{A_RANGE_HIGH}")
    print(f"  查找: {'最大值' if FIND_MAX else '最小值'}")
    print(f"  SPL查找目标: {SPL_TARGETS}")
    start_time = time.time()

    # 读取Excel文件
    print(f"正在读取Excel文件...")
    excel_file = pd.ExcelFile(FILE_PATH)

    # 获取IMP原档中的数据
    print(f"正在解析'IMP原档'工作表...")
    df = excel_file.parse("IMP原档")

    # 检查IMP原档中有数值的列数是否为偶数
    non_empty_columns = df.dropna(axis=1, how='all').shape[1]

    if non_empty_columns % 2 != 0:
        raise ValueError(f"IMP原档中有数值的列数为{non_empty_columns}，必须为偶数")

    # 使用openpyxl将值写入ACR表
    print(f"准备写入'ACR'工作表...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb["ACR"]

    # 处理所有奇数列(1,3,5...)及其后续偶数列(2,4,6...)
    print(f"正在处理列对数据...")
    max_col = df.shape[1]
    processed_pairs = 0
    skipped_pairs = 0

    for col_pair in range(0, max_col, 2):
        # 检查是否有足够的列
        if col_pair + 1 >= max_col:
            break

        # 获取当前奇数列和偶数列
        odd_col = col_pair
        even_col = col_pair + 1

        # 获取列字母表示（用于日志输出）
        col_letter_odd = get_column_letter(odd_col + 1)
        col_letter_even = get_column_letter(even_col + 1)

        print(f"  处理列对: {col_letter_odd}&{col_letter_even}")

        # 提取当前奇数列和偶数列的数据
        current_df = df.iloc[:, [odd_col, even_col]].copy()

        # 转换为数值类型并删除非数值
        current_df.iloc[:, 0] = pd.to_numeric(current_df.iloc[:, 0], errors='coerce')
        current_df.iloc[:, 1] = pd.to_numeric(current_df.iloc[:, 1], errors='coerce')
        current_df = current_df.dropna()

        # 跳过空列对
        if current_df.empty:
            print(f"    跳过空列对: {col_letter_odd}&{col_letter_even}")
            skipped_pairs += 1
            continue

        # 找到奇数列最接近目标值的值对应的偶数列的值
        nearest_value = find_nearest_value(current_df, TARGET_VALUE)

        # 如果找不到合适的值，则跳过
        if nearest_value is None:
            print(f"    在列 {col_letter_odd} 中未找到合适的值")
            skipped_pairs += 1
            continue

        # 获取对应偶数列的值
        nearest_row = current_df.iloc[(current_df.iloc[:, 0] - TARGET_VALUE).abs().argsort()[:1]]
        value_to_write = nearest_row.iloc[0, 1]

        # 计算ACR表中的行号(从1开始，每对占一行)
        row_in_acr = (col_pair // 2) + 1

        # 将值写入ACR表的对应行的第一列
        ws.cell(row=row_in_acr, column=1).value = value_to_write
        processed_pairs += 1

    print(f"列对数据处理完成: 已处理 {processed_pairs} 对, 跳过 {skipped_pairs} 对")

    # 读取ACR表中的所有数据
    print(f"正在分析'ACR'工作表数据...")
    acr_data = []
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None:
            acr_data.append(cell_value)

    # 计算分割点
    split_index = len(acr_data) // 2

    # 将后一半数据移至B列的起始行，并清空A列对应位置
    print(f"正在重新排列'ACR'工作表数据...")
    moved_values = 0

    for i in range(split_index, len(acr_data)):
        source_row = i + 1
        target_row = (i - split_index) + 1
        ws.cell(row=target_row, column=2).value = acr_data[i]
        ws.cell(row=source_row, column=1).value = None  # 清空A列原数据
        moved_values += 1

    print(f"数据重新排列完成: 已移动 {moved_values} 个值到B列")

    # 比较AB两列相邻数据，确保B列数值大于A列
    print(f"正在验证'ACR'工作表AB列数据关系...")
    swap_count = 0
    max_compare_row = max(ws.max_row, len(acr_data) - split_index)

    for row in range(1, max_compare_row + 1):
        a_value = ws.cell(row=row, column=1).value
        b_value = ws.cell(row=row, column=2).value

        # 尝试转换为数值
        try:
            a_value = float(a_value) if a_value is not None else None
            b_value = float(b_value) if b_value is not None else None
        except (ValueError, TypeError):
            continue

        # 确保两个单元格都有数值
        if a_value is not None and b_value is not None:
            # 如果B列值小于等于A列值，则交换
            if b_value <= a_value:
                ws.cell(row=row, column=1).value = b_value
                ws.cell(row=row, column=2).value = a_value
                swap_count += 1

    print(f"数据验证完成: 执行了 {swap_count} 次交换操作")

    # 处理Fb工作表 - 对所有偶数列执行相同操作
    print(f"正在处理'Fb'工作表数据...")

    # 创建或获取Fb工作表
    if "Fb" in wb.sheetnames:
        fb_sheet = wb["Fb"]
    else:
        fb_sheet = wb.create_sheet("Fb")

    # 清空Fb工作表中已有的数据
    for row in range(1, fb_sheet.max_row + 1):
        for col in range(1, fb_sheet.max_column + 1):
            fb_sheet.cell(row=row, column=col).value = None

    # 处理所有偶数列（B,D,F...）
    result_values = []

    for col in range(1, max_col, 2):
        # 获取列字母表示
        col_letter = get_column_letter(col + 1)

        # 获取当前偶数列和前一列(奇数列)的数据
        odd_col = col - 1
        even_col = col

        # 检查列索引是否有效
        if odd_col < 0 or even_col >= max_col:
            print(f"    跳过无效列索引: {odd_col}和{even_col}")
            continue

        # 提取数据
        current_df = df.iloc[:, [odd_col, even_col]].copy()

        # 转换为数值类型并删除非数值
        current_df.iloc[:, 0] = pd.to_numeric(current_df.iloc[:, 0], errors='coerce')
        current_df.iloc[:, 1] = pd.to_numeric(current_df.iloc[:, 1], errors='coerce')
        current_df = current_df.dropna()

        # 跳过空列
        if current_df.empty:
            print(f"    跳过空列对: {get_column_letter(odd_col + 1)}&{col_letter}")
            continue

        # 筛选奇数列在指定范围内的数据
        filtered_df = current_df[(current_df.iloc[:, 0] >= A_RANGE_LOW) & (current_df.iloc[:, 0] <= A_RANGE_HIGH)]

        if not filtered_df.empty:
            # 重置索引以确保索引连续
            filtered_df = filtered_df.reset_index(drop=True)

            # 确保筛选后的数据有足够的行
            if len(filtered_df) == 0:
                print(f"    跳过: {col_letter}列在范围内没有有效值")
                continue

            # 根据配置查找偶数列的最大值或最小值
            if FIND_MAX:
                extreme_value = filtered_df.iloc[:, 1].max()
                extreme_type = "最大值"
            else:
                extreme_value = filtered_df.iloc[:, 1].min()
                extreme_type = "最小值"

            # 找到对应的行
            extreme_rows = filtered_df[filtered_df.iloc[:, 1] == extreme_value]

            # 如果有多个匹配值，取第一个
            if not extreme_rows.empty:
                extreme_row = extreme_rows.iloc[0]

                # 获取奇数列和偶数列的值
                odd_col_value = extreme_row.iloc[0]
                even_col_value = extreme_row.iloc[1]

                print(
                    f"  在{get_column_letter(odd_col + 1)}列范围 {A_RANGE_LOW}~{A_RANGE_HIGH} 内找到{col_letter}列的{extreme_type}: {even_col_value}")
                print(f"    对应的{get_column_letter(odd_col + 1)}列值为: {odd_col_value}")

                # 保存结果
                result_values.append(odd_col_value)
            else:
                print(f"    警告: 未找到符合条件的行")
        else:
            print(f"  在{get_column_letter(odd_col + 1)}列中未找到范围在 {A_RANGE_LOW}~{A_RANGE_HIGH} 之间的值")

    # 将结果按顺序写入Fb工作表的第一列
    if result_values:
        for i, value in enumerate(result_values):
            fb_sheet.cell(row=i + 1, column=1).value = value

        print(f"已将所有结果按顺序写入'Fb'工作表的第一列")

        # 将Fb工作表后一半数据移至B列
        total_results = len(result_values)
        split_index_fb = (total_results + 1) // 2  # 向上取整

        print(f"正在重新排列'Fb'工作表数据...")
        moved_values_fb = 0

        for i in range(split_index_fb, total_results):
            source_row = i + 1
            target_row = (i - split_index_fb) + 1
            value = fb_sheet.cell(row=source_row, column=1).value
            fb_sheet.cell(row=target_row, column=2).value = value
            fb_sheet.cell(row=source_row, column=1).value = None  # 清空A列原数据
            moved_values_fb += 1

        print(f"Fb表数据重新排列完成: 已移动 {moved_values_fb} 个值到B列")

        # 比较Fb表AB两列相邻数据，确保B列数值小于A列
        print(f"正在验证'Fb'工作表AB列数据关系...")
        swap_count_fb = 0
        max_row_fb = max(fb_sheet.max_row, split_index_fb)

        for row in range(1, max_row_fb + 1):
            a_value = fb_sheet.cell(row=row, column=1).value
            b_value = fb_sheet.cell(row=row, column=2).value

            # 尝试转换为数值
            try:
                a_value = float(a_value) if a_value is not None else None
                b_value = float(b_value) if b_value is not None else None
            except (ValueError, TypeError):
                continue

            # 确保两个单元格都有数值
            if a_value is not None and b_value is not None:
                # 如果B列值大于等于A列值，则交换
                if b_value >= a_value:
                    fb_sheet.cell(row=row, column=1).value = b_value
                    fb_sheet.cell(row=row, column=2).value = a_value
                    swap_count_fb += 1

        print(f"Fb表数据验证完成: 执行了 {swap_count_fb} 次交换操作")
    else:
        print(f"没有找到符合条件的数据，'Fb'工作表保持为空")

    # 处理SPL原档工作表
    print(f"正在处理'SPL原档'工作表数据...")

    try:
        # 获取SPL原档中的数据
        spl_df = excel_file.parse("SPL原档")

        # 提取A列和B列的数据
        spl_col_a = pd.to_numeric(spl_df.iloc[:, 0], errors='coerce').dropna()
        spl_col_b = pd.to_numeric(spl_df.iloc[:, 1], errors='coerce').dropna()

        # 合并A列和B列数据
        spl_merged_df = pd.concat([spl_col_a, spl_col_b], axis=1).dropna()
        spl_merged_df.columns = ['A', 'B']

        if not spl_merged_df.empty:
            # 创建或获取SPL工作表
            if "SPL" in wb.sheetnames:
                spl_sheet = wb["SPL"]
            else:
                spl_sheet = wb.create_sheet("SPL")

            # 清空SPL工作表中已有的数据
            for row in range(1, spl_sheet.max_row + 1):
                for col in range(1, spl_sheet.max_column + 1):
                    spl_sheet.cell(row=row, column=col).value = None

            # 查找每个目标值最接近的A列值及其对应的B列值
            nearest_values = []

            for target in SPL_TARGETS:
                if not spl_merged_df.empty:
                    # 找到最接近目标值的A列值
                    nearest_value = spl_merged_df.iloc[(spl_merged_df['A'] - target).abs().argsort()[:1]]

                    if not nearest_value.empty:
                        a_value = nearest_value.iloc[0]['A']
                        b_value = nearest_value.iloc[0]['B']

                        print(f"  在A列中找到最接近{target}的值: {a_value}")
                        print(f"    对应的B列值为: {b_value}")

                        nearest_values.append(b_value)
                    else:
                        print(f"  警告: 在A列中未找到接近{target}的值")
                else:
                    print(f"  警告: SPL原档中没有有效的数据")

            # 计算B列值的平均值
            if nearest_values:
                average_value = sum(nearest_values) / len(nearest_values)

                # 将平均值写入SPL工作表的第一行第一列
                spl_sheet.cell(row=1, column=1).value = average_value

                print(f"已将B列对应值的平均值 {average_value:.4f} 写入'SPL'工作表的第一行第一列")
            else:
                print(f"没有找到符合条件的数据，无法计算平均值")
        else:
            print(f"SPL原档中没有有效的数据")

    except Exception as spl_e:
        print(f"处理'SPL原档'工作表时发生错误: {spl_e}")

    # 保存修改
    print(f"正在保存修改后的Excel文件...")
    wb.save(FILE_PATH)

    end_time = time.time()
    execution_time = end_time - start_time

    print(f"所有操作已成功完成!")
    print(f"程序运行时间: {execution_time:.4f} 秒")

except Exception as e:
    print(f"执行过程中发生错误: {e}")