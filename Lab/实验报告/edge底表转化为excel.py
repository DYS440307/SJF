import pandas as pd
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')  # 忽略Excel相关的无关警告

# 定义文件路径
file_path = r"E:\System\pic\A报告\老化实验记录.xlsx"

# ---------------------- 第一步：清空实验记录（保留表头） ----------------------
try:
    # 加载Excel工作簿
    wb = load_workbook(file_path)

    # 检查是否存在“实验记录”sheet
    if "实验记录" in wb.sheetnames:
        ws = wb["实验记录"]
        # 清空除第一行（表头）外的所有行（从第2行开始到最后一行）
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)  # 从第2行开始，删除所有数据行
        wb.save(file_path)  # 保存清空后的工作簿
    wb.close()  # 关闭工作簿
except FileNotFoundError:
    print("⚠️  文件未找到，请检查路径是否正确！")
    exit()
except Exception as e:
    print(f"⚠️  清空数据时出错：{str(e)}")
    exit()

# ---------------------- 第二步：读取实验底表数据 ----------------------
df_source = pd.read_excel(
    file_path,
    sheet_name="实验底表",
    header=0
    # 完全保留所有列的原始格式，不做强制类型转换
)

# 转换日期列（底表D/E/F列：委托时间、开始时间、结束时间）为“年月日”格式
date_columns = ["委托时间", "开始时间", "结束时间"]
for col in date_columns:
    df_source[col] = pd.to_datetime(df_source[col], errors='coerce').dt.strftime("%Y年%m月%d日")
    df_source[col] = df_source[col].fillna("")

# ---------------------- 第三步：精准映射（仅用“使用中设备名称”作为使用设备） ----------------------
df_target = pd.DataFrame()

# 严格按实验记录列顺序填充
df_target["委托时间"] = df_source["委托时间"]  # A列（年月日格式）
df_target["开始测试"] = df_source["开始时间"]  # B列（年月日格式）
df_target["结束测试"] = df_source["结束时间"]  # C列（年月日格式）
df_target["进度"] = df_source["实验进度"]  # D列
df_target["出具报告"] = "是"  # E列
df_target["送测部门"] = df_source["送测部门"]  # F列
df_target["送测人"] = df_source["送测人"]  # G列
df_target["生产批号"] = df_source["样品批号"]  # H列
df_target["型号"] = df_source["测试型号"]  # I列
df_target["数量"] = df_source["测试数量"]  # J列
df_target["试验项目"] = df_source["实验项目"]  # K列
df_target["试验目的"] = df_source["使用目的"]  # L列
df_target["条件"] = df_source["测试条件"]  # M列
df_target["报告编号"] = df_source.iloc[:, 0]  # N列
df_target["问题描述"] = ""  # O列
df_target["备注"] = ""  # P列
# 核心修改：仅使用“使用中设备名称”作为“使用设备”，完全不涉及设备通道
df_target["使用设备"] = df_source["使用中设备名称"]

# 替换全文的英文分号(;)为中文分号(；)
for col in df_target.columns:
    if df_target[col].dtype == "object":
        df_target[col] = df_target[col].astype(str).str.replace(";", "；", regex=False)

# ---------------------- 第四步：写入新数据 ----------------------
with pd.ExcelWriter(
        file_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay"
) as writer:
    df_target.to_excel(
        writer,
        sheet_name="实验记录",
        index=False,
        header=False,
        startrow=1
    )

print("✅ 操作完成！")
print("📌 核心变更：“使用设备”列仅包含“使用中设备名称”，无设备通道内容；其他格式要求均已满足。")