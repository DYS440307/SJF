import os
import openpyxl
import secrets
import shutil
from datetime import datetime
import win32com.client
import time
from pathlib import Path


class Config:
    """程序配置类，集中管理所有可配置参数"""
    # 基础路径配置
    BASE_DIR = r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\3-周期验证\2025年"
    OUTPUT_DIR = r"E:\System\desktop\PY\实验室"
    SALES_DETAIL_FILE = r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\3-周期验证\销售明细.xlsx"

    # PDF输出配置
    PDF_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "PDF输出")

    # 处理模式配置
    PROCESS_MODE = {
        'large_quantity': True,  # 处理实发数量>QUANTITY_THRESHOLD的单据
        'closest_small_quantity': True  # 处理最近且实发数量<QUANTITY_THRESHOLD的单据
    }

    # 实发数量阈值配置
    QUANTITY_THRESHOLD = 200

    # 动态查找的路径
    SOURCE_DIR = None
    CONFIG_FILE = None
    MATERIAL_CODE = None

    def __init__(self, material_code=None):
        """初始化配置，从文件加载动态配置"""
        # 如果未指定物料编码，则从销售明细中提取第一个物料编码
        if material_code:
            self.MATERIAL_CODE = material_code
        else:
            self.MATERIAL_CODE = self._get_material_code_from_sales_detail()

        self._find_config_and_source_dir(self.MATERIAL_CODE)

        if not self.SOURCE_DIR or not self.CONFIG_FILE:
            raise FileNotFoundError(f"无法找到物料编码为 {self.MATERIAL_CODE} 的配置文件和模板目录")

        self.load_config_from_file()

    def _get_material_code_from_sales_detail(self):
        """从销售明细文件中提取物料编码"""
        try:
            if not os.path.exists(self.SALES_DETAIL_FILE):
                raise FileNotFoundError(f"销售明细文件不存在 - {self.SALES_DETAIL_FILE}")

            # 打开销售明细Excel文件
            workbook = openpyxl.load_workbook(self.SALES_DETAIL_FILE, data_only=True)
            sheet = workbook.active

            # 获取表头行，确定物料编码列索引
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            material_col = None

            # 查找物料编码列对应的索引
            for idx, cell_value in enumerate(header_row):
                if cell_value is None:
                    continue
                cell_value = str(cell_value).strip().lower()
                if '物料编码' in cell_value:
                    material_col = idx
                    break

            # 检查是否找到了物料编码列
            if material_col is None:
                raise ValueError("在销售明细文件中找不到物料编码列")

            # 从第二行开始遍历数据行，获取第一个非空的物料编码
            for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
                material_code = row[material_col]
                if material_code is not None:
                    return str(material_code)

            raise ValueError("销售明细文件中没有找到有效的物料编码")

        except Exception as e:
            raise Exception(f"从销售明细文件获取物料编码时出错: {e}")
        finally:
            if 'workbook' in locals():
                workbook.close()

    def _find_config_and_source_dir(self, material_code):
        """在BASE_DIR下查找匹配的配置文件和模板目录"""
        config_file_name = f"{material_code}配置文件.txt"
        template_dir_name = f"({material_code})模板"

        print(f"正在查找物料编码为 {material_code} 的配置文件和模板目录...")

        for root, dirs, files in os.walk(self.BASE_DIR):
            # 检查是否包含配置文件
            if config_file_name in files:
                config_file_path = os.path.join(root, config_file_name)
                # 检查同级或子目录中是否有匹配的模板目录
                for d in [root] + [os.path.join(root, d) for d in dirs]:
                    if os.path.basename(d) == template_dir_name:
                        self.SOURCE_DIR = d
                        self.CONFIG_FILE = config_file_path
                        print(f"找到配置文件: {self.CONFIG_FILE}")
                        print(f"找到模板目录: {self.SOURCE_DIR}")
                        return

        print(f"警告: 未找到物料编码为 {material_code} 的配置文件和模板目录")
        # 如果找不到特定编码的模板，尝试使用通用模板
        for root, dirs, files in os.walk(self.BASE_DIR):
            if "通用模板" in dirs:
                self.SOURCE_DIR = os.path.join(root, "通用模板")
                print(f"使用通用模板目录: {self.SOURCE_DIR}")
                # 尝试查找配置文件
                for file in files:
                    if "配置文件.txt" in file:
                        self.CONFIG_FILE = os.path.join(root, file)
                        print(f"使用通用配置文件: {self.CONFIG_FILE}")
                        return
        raise FileNotFoundError(f"无法找到物料编码为 {material_code} 的配置文件和模板目录，且无通用模板可用")

    def load_config_from_file(self):
        """从配置文件加载动态配置"""
        if not os.path.exists(self.CONFIG_FILE):
            raise FileNotFoundError(f"配置文件不存在 - {self.CONFIG_FILE}")

        try:
            with open(self.CONFIG_FILE, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            config_section = None
            self.RANGE_CONFIG = {}
            self.FILE_FILTERS = {'extensions': [], 'keywords': []}

            for line in lines:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                # 检测配置区域
                if line.startswith('[') and line.endswith(']'):
                    config_section = line[1:-1].strip()
                    continue

                # 解析配置项
                if config_section == 'RANGE_CONFIG':
                    parts = line.split('=')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        try:
                            # 解析元组值
                            values = parts[1].strip()
                            # 移除括号并分割
                            values = values.strip('()').split(',')
                            values = [float(v.strip()) for v in values if v.strip()]
                            if len(values) == 3:
                                self.RANGE_CONFIG[key] = tuple(values)
                        except Exception as e:
                            raise ValueError(f"解析RANGE_CONFIG配置项 '{line}' 失败: {e}")

                elif config_section == 'DATA_RANGE':
                    parts = line.split('=')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        try:
                            value = int(parts[1].strip())
                            if key == 'ROW_START':
                                self.ROW_START = value
                            elif key == 'ROW_END':
                                self.ROW_END = value
                        except Exception as e:
                            raise ValueError(f"解析DATA_RANGE配置项 '{line}' 失败: {e}")

                elif config_section == 'FILE_FILTERS':
                    parts = line.split('=')
                    if len(parts) == 2:
                        key = parts[0].strip()
                        value = parts[1].strip()
                        if key == 'extensions':
                            # 解析扩展名列表
                            exts = [ext.strip() for ext in value.split(',') if ext.strip()]
                            self.FILE_FILTERS['extensions'] = exts
                        elif key == 'keywords':
                            # 解析关键词列表
                            keywords = [kw.strip() for kw in value.split(',') if kw.strip()]
                            self.FILE_FILTERS['keywords'] = keywords

            # 验证配置是否完整
            if not self.validate_config():
                raise ValueError("配置文件不完整或包含无效配置")

        except Exception as e:
            raise Exception(f"读取配置文件时出错: {e}")

    def validate_config(self):
        """验证配置是否完整有效"""
        # 验证RANGE_CONFIG
        required_ranges = {'B_C', 'D_E', 'F_G', 'H_I'}
        if not all(key in self.RANGE_CONFIG for key in required_ranges):
            return False
        if not all(len(v) == 3 and all(isinstance(x, (int, float)) for x in v) for v in self.RANGE_CONFIG.values()):
            return False

        # 验证ROW_START和ROW_END
        if not (isinstance(self.ROW_START, int) and isinstance(self.ROW_END, int) and self.ROW_START <= self.ROW_END):
            return False

        # 验证FILE_FILTERS
        if not ('extensions' in self.FILE_FILTERS and 'keywords' in self.FILE_FILTERS):
            return False
        if not (isinstance(self.FILE_FILTERS['extensions'], list) and isinstance(self.FILE_FILTERS['keywords'], list)):
            return False
        if not all(isinstance(ext, str) for ext in self.FILE_FILTERS['extensions']):
            return False
        if not all(isinstance(kw, str) for kw in self.FILE_FILTERS['keywords']):
            return False

        return True


# ===== 功能函数 =====
def generate_random_numbers(existing_values, value_range, ensure_first_larger=False):
    """
    生成两个不重复的随机数，可配置确保第一个数大于第二个数

    参数:
        existing_values (set): 已存在的值集合，用于避免重复
        value_range (tuple): 范围配置 (最小值, 最大值, 最小差值)
        ensure_first_larger (bool): 是否确保第一个数大于第二个数

    返回:
        tuple: 两个不重复的随机数
    """
    min_val, max_val, min_diff = value_range
    max_attempts = 100

    for _ in range(max_attempts):
        # 生成两个随机数
        value1 = round(secrets.SystemRandom().uniform(min_val, max_val), 3)
        value2 = round(secrets.SystemRandom().uniform(min_val, max_val), 3)

        # 确保两个数的差值符合要求
        if abs(value1 - value2) < min_diff:
            continue

        # 如果需要确保第一个数大于第二个数
        if ensure_first_larger and value1 <= value2:
            value1, value2 = value2, value1  # 交换值

        # 检查是否有重复
        if value1 not in existing_values and value2 not in existing_values:
            return value1, value2

    raise Exception("无法在100次尝试内生成不重复的随机数")


def process_excel_file(file_path, output_dir, order_date, order_number, material_code, config):
    """
    处理单个Excel文件：填充随机数并转换为PDF

    参数:
        file_path (str): 源Excel文件路径
        output_dir (str): 输出目录
        order_date (str): 订单日期 (格式: YYYY/M/D)
        order_number (str): 订单编号
        material_code (str): 物料编码
        config (Config): 配置对象

    返回:
        bool: 处理是否成功
    """
    try:
        # 从订单日期字符串提取年份和月份
        year, month, _ = order_date.split('/')
        month_folder = f"{year}年{month}月"

        # 基于物料编码和月份创建子目录
        material_dir = os.path.join(output_dir, str(material_code))
        month_dir = os.path.join(material_dir, month_folder)
        os.makedirs(month_dir, exist_ok=True)

        # 打开Excel工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet = workbook.active

        # 写入订单信息
        sheet['G2'] = order_date
        sheet['L2'] = order_number

        # 用于存储已生成的值，确保不重复
        existing_values = set()

        # 填充随机数到指定区域
        for row in range(config.ROW_START, config.ROW_END + 1):
            # 初始化变量，避免NameError
            value_b, value_c, value_d, value_e, value_f, value_g, value_h, value_i = None, None, None, None, None, None, None, None

            max_attempts = 100  # 最大尝试次数
            for attempt in range(max_attempts):
                # 临时集合，用于验证当前尝试的所有值
                temp_values = set(existing_values)

                # 生成所有列值
                try:
                    # B列和C列（C < B）
                    value_b, value_c = generate_random_numbers(temp_values, config.RANGE_CONFIG['B_C'],
                                                               ensure_first_larger=True)
                    temp_values.update([value_b, value_c])

                    # D列和E列（E > D）
                    value_d, value_e = generate_random_numbers(temp_values, config.RANGE_CONFIG['D_E'],
                                                               ensure_first_larger=False)
                    temp_values.update([value_d, value_e])

                    # F列和G列（G < F）
                    value_f, value_g = generate_random_numbers(temp_values, config.RANGE_CONFIG['F_G'],
                                                               ensure_first_larger=True)
                    temp_values.update([value_f, value_g])

                    # H列和I列（I > H）
                    value_h, value_i = generate_random_numbers(temp_values, config.RANGE_CONFIG['H_I'],
                                                               ensure_first_larger=False)
                    temp_values.update([value_h, value_i])

                    # 验证所有条件
                    if (value_c < value_b and
                            value_e > value_d and
                            value_g < value_f and
                            value_i > value_h):
                        # 条件全部满足，更新existing_values并跳出循环
                        existing_values.update(temp_values)
                        break

                except Exception as e:
                    # 生成失败，继续尝试
                    pass

                if attempt == max_attempts - 1:
                    raise Exception(f"行 {row}: 无法在{max_attempts}次尝试内生成满足所有条件的随机数")

            # 写入数据到对应单元格
            sheet[f'B{row}'] = value_b  # B列值（较大值）
            sheet[f'C{row}'] = value_c  # C列值（较小值）
            sheet[f'D{row}'] = value_d  # D列值（较小值）
            sheet[f'E{row}'] = value_e  # E列值（较大值）
            sheet[f'F{row}'] = value_f  # F列值（较大值）
            sheet[f'G{row}'] = value_g  # G列值（较小值）
            sheet[f'H{row}'] = value_h  # H列值（较小值）
            sheet[f'I{row}'] = value_i  # I列值（较大值）

            # 更新已存在的值集合
            existing_values.update([value_b, value_c, value_d, value_e, value_f, value_g, value_h, value_i])

        # 保存修改后的Excel文件
        file_name = os.path.basename(file_path)
        new_name = file_name.replace("模板", f"_{order_number}")
        output_file_path = os.path.join(month_dir, new_name)
        workbook.save(output_file_path)
        print(f"成功处理Excel: {file_name} -> {new_name} (物料编码: {material_code}, 月份: {month_folder})")

        # 转换为PDF
        pdf_material_dir = os.path.join(config.PDF_OUTPUT_DIR, str(material_code))
        pdf_month_dir = os.path.join(pdf_material_dir, month_folder)
        os.makedirs(pdf_month_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_month_dir, os.path.splitext(new_name)[0] + ".pdf")

        if excel_to_pdf(output_file_path, pdf_path):
            print(f"成功转换为PDF: {pdf_path}")
            return True
        else:
            print(f"PDF转换失败: {output_file_path}")
            return False

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {e}")
        return False


def excel_to_pdf(excel_path, pdf_path):
    """
    使用Excel COM接口将Excel文件转换为PDF

    参数:
        excel_path (str): Excel文件路径
        pdf_path (str): PDF输出路径

    返回:
        bool: 转换是否成功
    """
    try:
        # 创建Excel应用实例
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # 打开工作簿
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))

        # 创建PDF输出目录
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

        # 导出为PDF（所有工作表）
        workbook.ExportAsFixedFormat(0, pdf_path)

        # 关闭工作簿和Excel应用
        workbook.Close()
        excel.Quit()

        # 释放COM对象
        del workbook
        del excel

        return True
    except Exception as e:
        print(f"Excel转PDF失败: {excel_path} -> {pdf_path}, 错误: {e}")
        return False
    finally:
        # 确保资源被释放
        time.sleep(1)  # 等待Excel完全退出


def get_input_pairs(config, target_material_code):
    """
    从销售明细Excel文件中获取指定物料编码的单据对
    根据配置决定获取实发数量>6000的所有单据还是最近且实发数量<6000的单据

    参数:
        config (Config): 配置对象
        target_material_code (str): 目标物料编码

    返回:
        list: 包含元组 (日期, 订单编号, 物料编码) 的列表
    """
    pairs = []

    if config.PROCESS_MODE['large_quantity']:
        # 获取所有实发数量>阈值的单据
        pairs.extend(get_large_quantity_pairs(config, target_material_code))

    if config.PROCESS_MODE['closest_small_quantity']:
        # 获取最近且实发数量<阈值的单据
        closest_pair = get_closest_small_quantity_pair(config, target_material_code)
        if closest_pair:
            pairs.append(closest_pair)

    return pairs


def get_large_quantity_pairs(config, target_material_code):
    """
    获取指定物料编码且实发数量>阈值的所有单据

    参数:
        config (Config): 配置对象
        target_material_code (str): 目标物料编码

    返回:
        list: 包含元组 (日期, 订单编号, 物料编码) 的列表
    """
    pairs = []

    try:
        if not os.path.exists(config.SALES_DETAIL_FILE):
            raise FileNotFoundError(f"销售明细文件不存在 - {config.SALES_DETAIL_FILE}")

        # 打开销售明细Excel文件
        workbook = openpyxl.load_workbook(config.SALES_DETAIL_FILE, data_only=True)
        sheet = workbook.active

        # 获取表头行，确定各列索引
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        date_col = None
        order_col = None
        material_col = None
        quantity_col = None

        # 查找各列对应的索引
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_value = str(cell_value).strip().lower()
            if '日期' in cell_value:
                date_col = idx
            elif '单据编号' in cell_value:
                order_col = idx
            elif '物料编码' in cell_value:
                material_col = idx
            elif '实发数量' in cell_value:
                quantity_col = idx

        # 检查是否找到了所有需要的列
        if any(col is None for col in [date_col, order_col, material_col, quantity_col]):
            missing = [col_name for col_name, col_idx in
                       [('日期', date_col), ('单据编号', order_col), ('物料编码', material_col),
                        ('实发数量', quantity_col)]
                       if col_idx is None]
            raise ValueError(f"在销售明细文件中找不到以下列: {', '.join(missing)}")

        # 从第二行开始遍历数据行
        processed_orders = set()  # 用于记录已处理的订单编号，避免重复
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not any(row):
                continue

            # 获取各列值
            order_date = row[date_col]
            order_number = row[order_col]
            material_code = row[material_col]
            # 实发数量需要转换为数值类型
            try:
                quantity = float(row[quantity_col]) if row[quantity_col] is not None else 0
            except (ValueError, TypeError):
                continue

            # 检查条件：物料编码匹配、实发数量 > 阈值 且订单编号未处理过
            if (str(material_code) == target_material_code and
                    quantity > config.QUANTITY_THRESHOLD and
                    order_number not in processed_orders):
                # 处理日期格式
                formatted_date = format_date(order_date)
                if formatted_date:
                    pairs.append((formatted_date, order_number, material_code))
                    processed_orders.add(order_number)
                    print(f"已添加: {formatted_date} {order_number} (物料编码: {material_code}, 实发数量: {quantity})")

        workbook.close()
        print(
            f"从销售明细文件中提取了 {len(pairs)} 个物料编码为 {target_material_code} 且实发数量大于{config.QUANTITY_THRESHOLD}的订单")

        return pairs

    except Exception as e:
        print(f"获取大数量单据时出错: {e}")
        return pairs


def get_closest_small_quantity_pair(config, target_material_code):
    """
    获取距离今日最近且物料编码匹配、实发数量<阈值的单据

    参数:
        config (Config): 配置对象
        target_material_code (str): 目标物料编码

    返回:
        tuple: (日期, 订单编号, 物料编码) 或 None
    """
    today = datetime.now().date()
    closest_record = None
    min_days_diff = float('inf')

    try:
        if not os.path.exists(config.SALES_DETAIL_FILE):
            raise FileNotFoundError(f"销售明细文件不存在 - {config.SALES_DETAIL_FILE}")

        # 打开销售明细Excel文件
        workbook = openpyxl.load_workbook(config.SALES_DETAIL_FILE, data_only=True)
        sheet = workbook.active

        # 获取表头行，确定各列索引
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        date_col = None
        order_col = None
        material_col = None
        quantity_col = None

        # 查找各列对应的索引
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_value = str(cell_value).strip().lower()
            if '日期' in cell_value:
                date_col = idx
            elif '单据编号' in cell_value:
                order_col = idx
            elif '物料编码' in cell_value:
                material_col = idx
            elif '实发数量' in cell_value:
                quantity_col = idx

        # 检查是否找到了所有需要的列
        if any(col is None for col in [date_col, order_col, material_col, quantity_col]):
            missing = [col_name for col_name, col_idx in
                       [('日期', date_col), ('单据编号', order_col), ('物料编码', material_col),
                        ('实发数量', quantity_col)]
                       if col_idx is None]
            raise ValueError(f"在销售明细文件中找不到以下列: {', '.join(missing)}")

        # 从第二行开始遍历数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not any(row):
                continue

            # 获取各列值
            order_date = row[date_col]
            order_number = row[order_col]
            material_code = row[material_col]
            # 实发数量需要转换为数值类型
            try:
                quantity = float(row[quantity_col]) if row[quantity_col] is not None else 0
            except (ValueError, TypeError):
                continue

            # 检查条件：物料编码匹配、实发数量 < 阈值
            if str(material_code) == target_material_code and quantity < config.QUANTITY_THRESHOLD:
                # 处理日期格式
                date_obj = parse_date(order_date)
                if date_obj:
                    # 计算与今日的天数差
                    days_diff = abs((date_obj - today).days)

                    # 如果是更接近的日期，更新最近记录
                    if days_diff < min_days_diff:
                        min_days_diff = days_diff
                        closest_record = (date_obj, order_number, material_code, quantity)

        workbook.close()

        if closest_record:
            # 格式化日期为 YYYY/M/D
            formatted_date = f"{closest_record[0].year}/{closest_record[0].month}/{closest_record[0].day}"
            print(
                f"找到最近的物料编码为 {target_material_code} 且实发数量小于{config.QUANTITY_THRESHOLD}的单据: {formatted_date} {closest_record[1]}，距离今日 {min_days_diff} 天，实发数量: {closest_record[3]}")
            return (formatted_date, closest_record[1], closest_record[2])
        else:
            print(f"未找到物料编码为 {target_material_code} 且实发数量小于{config.QUANTITY_THRESHOLD}的单据")
            return None

    except Exception as e:
        print(f"获取小数量单据时出错: {e}")
        return None


def format_date(date_value):
    """处理日期格式，返回 YYYY/M/D 格式的字符串"""
    if isinstance(date_value, datetime):
        return f"{date_value.year}/{date_value.month}/{date_value.day}"

    if isinstance(date_value, str):
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y%m%d'):
            try:
                date_obj = datetime.strptime(date_value, fmt)
                return f"{date_obj.year}/{date_obj.month}/{date_obj.day}"
            except ValueError:
                continue

    return None


def parse_date(date_value):
    """将日期值转换为datetime.date对象"""
    if isinstance(date_value, datetime):
        return date_value.date()

    if isinstance(date_value, str):
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y%m%d'):
            try:
                return datetime.strptime(date_value, fmt).date()
            except ValueError:
                continue

    return None


def get_excel_files(config):
    """
    根据配置获取符合条件的Excel文件列表

    参数:
        config (Config): 配置对象

    返回:
        list: 符合条件的文件路径列表
    """
    excel_files = []
    if not os.path.exists(config.SOURCE_DIR):
        print(f"错误: 源目录不存在 - {config.SOURCE_DIR}")
        return excel_files

    for root, _, files in os.walk(config.SOURCE_DIR):
        for file in files:
            # 检查文件扩展名
            if not any(file.lower().endswith(ext) for ext in config.FILE_FILTERS['extensions']):
                continue
            # 检查关键词
            if not all(keyword in file for keyword in config.FILE_FILTERS['keywords']):
                continue
            excel_files.append(os.path.join(root, file))

    return excel_files


def main():
    """程序主入口"""
    try:
        # 获取销售明细中的所有不同物料编码
        material_codes = get_unique_material_codes()

        if not material_codes:
            print("未在销售明细中找到有效的物料编码，程序退出")
            return

        print(f"找到 {len(material_codes)} 种不同的物料编码: {', '.join(material_codes)}")

        # 为每种物料编码处理相应的订单
        for material_code in material_codes:
            print(f"\n\n===== 处理物料编码 {material_code} 的订单 =====")

            # 创建对应物料编码的配置实例
            config = Config(material_code)

            print(f"\n使用配置:")
            print(f"  源目录: {config.SOURCE_DIR}")
            print(f"  输出目录: {config.OUTPUT_DIR}")
            print(f"  PDF输出目录: {config.PDF_OUTPUT_DIR}")
            print(f"  销售明细文件: {config.SALES_DETAIL_FILE}")
            print(f"  配置文件: {config.CONFIG_FILE}")
            print(f"  实发数量阈值: {config.QUANTITY_THRESHOLD}")
            print(
                f"  处理模式: 实发数量>{config.QUANTITY_THRESHOLD}的单据{'✓' if config.PROCESS_MODE['large_quantity'] else '✗'}, 最近且实发数量<{config.QUANTITY_THRESHOLD}的单据{'✓' if config.PROCESS_MODE['closest_small_quantity'] else '✗'}")

            # 从销售明细Excel文件获取该物料编码的单据对
            input_pairs = get_input_pairs(config, material_code)
            if not input_pairs:
                print(f"未找到物料编码为 {material_code} 的符合条件的数据，跳过处理")
                continue

            # 获取符合条件的Excel文件
            excel_files = get_excel_files(config)
            if not excel_files:
                print(f"未找到物料编码为 {material_code} 的符合条件的Excel文件，跳过处理")
                continue

            print(f"找到 {len(excel_files)} 个符合条件的文件")

            # 处理所有单据
            for order_date, order_number, material_code in input_pairs:
                print(f"\n处理订单: {order_date} {order_number} (物料编码: {material_code})")
                success_count = 0

                for file_path in excel_files:
                    if process_excel_file(file_path, config.OUTPUT_DIR, order_date, order_number, material_code,
                                          config):
                        success_count += 1

                print(
                    f"订单 {order_number} 处理完成: 成功 {success_count} 个, 失败 {len(excel_files) - success_count} 个")

    except Exception as e:
        print(f"程序执行失败: {e}")
        return 1

    return 0


def get_unique_material_codes():
    """从销售明细中提取所有不同的物料编码"""
    try:
        if not os.path.exists(Config.SALES_DETAIL_FILE):
            raise FileNotFoundError(f"销售明细文件不存在 - {Config.SALES_DETAIL_FILE}")

        # 打开销售明细Excel文件
        workbook = openpyxl.load_workbook(Config.SALES_DETAIL_FILE, data_only=True)
        sheet = workbook.active

        # 获取表头行，确定物料编码列索引
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        material_col = None

        # 查找物料编码列对应的索引
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_value = str(cell_value).strip().lower()
            if '物料编码' in cell_value:
                material_col = idx
                break

        # 检查是否找到了物料编码列
        if material_col is None:
            raise ValueError("在销售明细文件中找不到物料编码列")

        # 从第二行开始遍历数据行，收集所有不同的物料编码
        material_codes = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not any(row):
                continue

            material_code = row[material_col]
            if material_code is not None:
                material_codes.add(str(material_code))

        workbook.close()

        if not material_codes:
            print("警告: 销售明细中未找到任何物料编码")

        return list(material_codes)

    except Exception as e:
        print(f"从销售明细提取物料编码时出错: {e}")
        return []


if __name__ == "__main__":
    main()    