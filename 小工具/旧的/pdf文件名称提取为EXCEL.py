import os
import openpyxl
from openpyxl.styles import Alignment

# 配置路径
pdf_dir = r"E:\System\download\飞书下载"
excel_path = r"E:\System\download\飞书下载\名字.xlsx"

# 创建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PDF文件名提取"

# 设置表头
ws["A1"] = "第一列（编号）"
ws["B1"] = "第二列（名称）"

# 设置表头样式（可选，美化表格）
for col in ["A", "B"]:
    cell = ws[f"{col}1"]
    cell.font = openpyxl.styles.Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 调整列宽
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 25

# 遍历目录下的PDF文件
row = 2  # 从第二行开始写入数据
for filename in os.listdir(pdf_dir):
    # 只处理PDF文件
    if filename.lower().endswith(".pdf"):
        # 去除文件扩展名
        file_name_without_ext = os.path.splitext(filename)[0]

        # 按空格分隔，取前两部分
        parts = file_name_without_ext.split(" ", 2)  # split(" ", 2) 最多分割成3部分，取前两部分

        # 第一列数据
        col1 = parts[0] if len(parts) > 0 else ""
        # 第二列数据（如果有第二部分则取，没有则为空）
        col2 = parts[1] if len(parts) > 1 else ""

        # 写入Excel
        ws[f"A{row}"] = col1
        ws[f"B{row}"] = col2

        # 设置数据对齐方式
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].alignment = Alignment(horizontal="left")

        row += 1

# 保存Excel文件
try:
    wb.save(excel_path)
    print(f"成功提取 {row - 2} 个PDF文件名称")
    print(f"文件已保存到：{excel_path}")
except Exception as e:
    print(f"保存文件时出错：{str(e)}")
    print("请检查：1.路径是否存在 2.文件是否被占用 3.是否有写入权限")