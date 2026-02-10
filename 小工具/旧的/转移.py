import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# ===================== 核心配置区（已修正单元格映射）=====================
# 1. 文件路径配置
source_file_path = r"E:\System\download\供应商索赔清单 .xlsx"
template_sheet_name = "扣款通知单1"  # 模板工作表名
list_sheet_name = "索赔清单"  # 索赔清单工作表名
output_file_path = r"E:\System\download\供应商索赔清单_pandas生成.xlsx"  # 输出文件路径

# 2. 图片配置（路径+插入位置+尺寸（厘米））
IMAGE_CONFIGS = [
    {
        "path": r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\无归类文件\个人签名\声乐logo.png",
        "anchor": "A1",  # 嵌入位置
        "width_cm": 1.5,  # 宽度（厘米）
        "height_cm": 1.5  # 高度（厘米）
    },
    {
        "path": r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\无归类文件\个人签名\邓洋枢.png",
        "anchor": "B7",
        "width_cm": 2.5,
        "height_cm": 1.5
    },
    {
        "path": r"Z:\3-品质部\实验室\邓洋枢\1-实验室相关文件\无归类文件\个人签名\潘勇着.png",
        "anchor": "E7",
        "width_cm": 2.5,
        "height_cm": 1.5
    }
]

# 3. 数据填充单元格映射（已修正为匹配模板的B4/E4/B5）
CELL_MAP = {
    "supplier": "B4",  # 供方名称（原B3→修正为B4）
    "amount": "E4",    # 质量保证金（原E3→修正为E4）
    "abnormal": "B5"   # 异常描述（原B6→修正为B5）
}


# ===================== 工具函数（厘米转像素，固定逻辑无需修改）=====================
def cm_to_px(cm):
    """将厘米转换为像素（Excel默认96 DPI，1cm≈37.795像素）"""
    dpi = 96
    return int(cm * dpi / 2.54)


# ===================== 主生成函数 =====================
def batch_generate_with_pandas():
    # ========== 前置校验 ==========
    # 校验源文件是否存在
    if not os.path.exists(source_file_path):
        print(f"❌ 源文件 {source_file_path} 不存在，请检查路径！")
        return

    # 校验所有图片文件是否存在
    missing_imgs = []
    for cfg in IMAGE_CONFIGS:
        if not os.path.exists(cfg["path"]):
            missing_imgs.append(cfg["path"])
    if missing_imgs:
        print(f"❌ 以下图片文件不存在：")
        for path in missing_imgs:
            print(f"   - {path}")
        return

    # ========== 步骤1：读取并清洗索赔清单数据 ==========
    try:
        df_claim = pd.read_excel(
            source_file_path,
            sheet_name=list_sheet_name,
            header=0,
            usecols=["供方名称", "异常描述", "质量保证金(¥)"]
        )
        # 清理空行和空格
        df_claim = df_claim.dropna()
        df_claim["供方名称"] = df_claim["供方名称"].astype(str).str.strip()
        df_claim["异常描述"] = df_claim["异常描述"].astype(str).str.strip()
        df_claim["质量保证金(¥)"] = df_claim["质量保证金(¥)"].astype(str).str.strip()

        print(f"📊 成功读取 {len(df_claim)} 条有效索赔数据")
        print("📌 数据预览：")
        print(df_claim.head())

        if len(df_claim) == 0:
            print("❌ 索赔清单无有效数据，终止生成！")
            return
    except Exception as e:
        print(f"❌ 读取索赔清单失败：{str(e)}")
        return

    # ========== 步骤2：加载模板并批量生成通知单 ==========
    try:
        wb = load_workbook(source_file_path, read_only=False, data_only=False)

        # 校验模板工作表是否存在
        if template_sheet_name not in wb.sheetnames:
            print(f"❌ 未找到模板工作表「{template_sheet_name}」")
            wb.close()
            return

        template_ws = wb[template_sheet_name]

        # 遍历每条数据生成通知单
        for idx, row in df_claim.iterrows():
            # 1. 复制模板并重命名
            new_sheet_name = f"扣款通知单{idx + 1}"
            if new_sheet_name in wb.sheetnames:
                del wb[new_sheet_name]
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = new_sheet_name

            # 2. 填充核心数据（使用修正后的单元格映射）
            new_ws[CELL_MAP["supplier"]] = row["供方名称"]
            new_ws[CELL_MAP["amount"]] = row["质量保证金(¥)"]
            new_ws[CELL_MAP["abnormal"]] = row["异常描述"]

            # 3. 嵌入并调整每张图片
            img_result = {"成功": [], "失败": []}
            for cfg in IMAGE_CONFIGS:
                try:
                    # 加载图片
                    img = Image(cfg["path"])
                    # 转换尺寸（厘米→像素）并设置
                    img.width = cm_to_px(cfg["width_cm"])
                    img.height = cm_to_px(cfg["height_cm"])
                    # 插入到指定位置
                    new_ws.add_image(img, cfg["anchor"])
                    # 记录成功
                    img_name = cfg["path"].split("\\")[-1]
                    img_result["成功"].append(img_name)
                except Exception as e:
                    img_name = cfg["path"].split("\\")[-1]
                    img_result["失败"].append(f"{img_name}（{str(e)}）")

            # 打印当前通知单处理结果
            supplier = row["供方名称"]
            if img_result["失败"]:
                print(
                    f"⚠️ 通知单「{new_sheet_name}」（供方：{supplier}）：成功嵌入{img_result['成功']}，失败{img_result['失败']}")
            else:
                print(f"✅ 通知单「{new_sheet_name}」（供方：{supplier}）：所有图片嵌入成功！")

        # ========== 步骤3：保存文件 ==========
        wb.save(output_file_path)
        wb.close()
        print(f"\n🎉 全部生成完成！文件已保存至：{output_file_path}")
        print("✅ 图片已按指定尺寸（1.5cm*1.5cm/1.5cm*2.5cm）嵌入对应位置")

    except Exception as e:
        print(f"❌ 生成通知单失败：{str(e)}")
        if 'wb' in locals():
            wb.close()


# ===================== 执行生成 =====================
if __name__ == "__main__":
    batch_generate_with_pandas()