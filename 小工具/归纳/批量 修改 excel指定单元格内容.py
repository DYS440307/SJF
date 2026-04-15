import os
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import csv


# 统一格式化，确保 PyCharm 绝对对齐
def format_cell(value, width=10):
    s = str(value).strip() if value is not None else ""
    return s.ljust(width)[:width]


# ============================
# 【核查对比：递归遍历子文件夹版】
# ============================
def batch_check_fast_compare():
    print("=" * 80)
    print("📊 Excel 批量对比工具（递归遍历子文件夹）")
    print("=" * 80)

    folder_path = input("请输入文件夹路径：").strip()
    if not os.path.isdir(folder_path):
        print("❌ 路径错误")
        return

    target_area = input("请输入区域（如 A2:L4）：").strip()

    try:
        min_col, min_row, max_col, max_row = range_boundaries(target_area)
    except:
        print("❌ 格式错误")
        return

    # 读取所有文件（递归遍历子文件夹）
    files_data = []
    filenames = []

    # 递归遍历所有子文件夹
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if f.lower().endswith(".xlsx"):
                file_path = os.path.join(root, f)
                try:
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(min_row, max_row, min_col, max_col):
                        rows.append([cell.value for cell in row])
                    wb.close()
                    files_data.append(rows)
                    # 存储相对路径，方便区分子文件夹文件
                    rel_path = os.path.relpath(file_path, folder_path)
                    filenames.append(rel_path)
                    print(f"✅ {rel_path}")
                except Exception as e:
                    rel_path = os.path.relpath(file_path, folder_path)
                    print(f"❌ {rel_path}")

    if not files_data:
        print("\n未找到Excel")
        return

    # ============================
    # 核心：PyCharm 绝对对齐输出
    # ============================
    print("\n" + "=" * 120)
    print(f"🔍 对比区域：{target_area}")
    print("=" * 120)

    for row_idx in range(len(files_data[0])):
        current_row_num = min_row + row_idx
        print(f"\n📌 第 {current_row_num} 行")
        print("-" * 120)

        for file_idx, data in enumerate(files_data):
            # 文件名固定宽度（支持子文件夹路径显示）
            name = f"[{filenames[file_idx]:<35}]"

            # 每一列强制等宽
            cells = [format_cell(v) for v in data[row_idx]]
            line = " │ ".join(cells)

            print(f"{name} │ {line}")

    print("\n" + "=" * 120)
    print("✅ 对比完成！遍历所有子文件夹Excel")


# ============================
# 【优化：连续批量修改单元格 + 递归遍历】
# ============================
def batch_modify_cell():
    print("\n=== 批量连续修改单元格（递归遍历子文件夹）===")
    # 1. 只输入一次路径，全程复用
    folder = input("请输入文件夹路径：").strip()
    if not os.path.isdir(folder):
        print("❌ 路径错误")
        return

    # 2. 收集需要修改的单元格和内容（连续输入）
    modify_list = []
    print("\n📝 连续修改模式：输入单元格(如A1)和内容，输入 q 结束")
    while True:
        cell_addr = input("\n请输入单元格（输入 q 退出）：").strip()
        if cell_addr.lower() == "q":
            break
        if not cell_addr:
            print("❌ 单元格不能为空！")
            continue

        content = input(f"请输入 {cell_addr} 的新内容：").strip()
        modify_list.append((cell_addr, content))
        print(f"✅ 已添加修改：{cell_addr} = {content}")

    # 无修改任务直接退出
    if not modify_list:
        print("\n❌ 未添加任何修改任务")
        return

    # 3. 递归遍历所有子文件夹，批量执行修改
    count = 0
    for root, dirs, files in os.walk(folder):
        for filename in files:
            if filename.lower().endswith(".xlsx"):
                file_path = os.path.join(root, filename)
                rel_path = os.path.relpath(file_path, folder)
                try:
                    # 打开文件
                    wb = load_workbook(file_path)
                    ws = wb.active
                    # 批量修改所有单元格
                    for cell_addr, content in modify_list:
                        ws[cell_addr] = content
                    # 保存并关闭
                    wb.save(file_path)
                    wb.close()
                    count += 1
                    print(f"✅ 修改成功：{rel_path}")
                except Exception as e:
                    print(f"❌ 修改失败：{rel_path}，原因：{str(e)}")

    print(f"\n🎉 全部修改完成！成功修改 {count} 个文件，共 {len(modify_list)} 个单元格")


# ============================
# 【批量重命名 + 递归遍历子文件夹】
# ============================
def batch_rename_files():
    print("\n" + "=" * 50)
    print("📝 批量替换文件名称（递归遍历子文件夹）")
    print("=" * 50)

    folder = input("请输入文件夹路径：").strip()
    if not os.path.isdir(folder):
        print("❌ 路径错误")
        return

    old_str = input("请输入要替换的文字：").strip()
    new_str = input("请输入替换后的新文字：").strip()
    confirm = input(f"\n确认替换：「{old_str}」→「{new_str}」？(y/n)：").strip().lower()

    if confirm != "y":
        print("❌ 已取消重命名")
        return

    count = 0
    # 递归遍历所有子文件夹
    for root, dirs, files in os.walk(folder):
        for filename in files:
            # 只处理Excel文件
            if not filename.lower().endswith((".xlsx", ".xls")):
                continue
            # 包含要替换的文字才修改
            if old_str in filename:
                old_path = os.path.join(root, filename)
                new_filename = filename.replace(old_str, new_str)
                new_path = os.path.join(root, new_filename)
                rel_old = os.path.relpath(old_path, folder)
                rel_new = os.path.relpath(new_path, folder)

                try:
                    os.rename(old_path, new_path)
                    print(f"✅ {rel_old} → {rel_new}")
                    count += 1
                except Exception as e:
                    print(f"❌ 重命名失败：{rel_old}，原因：{str(e)}")

    print(f"\n🎉 重命名完成！成功修改 {count} 个Excel文件")


# ============================
# 主程序
# ============================
if __name__ == "__main__":
    print("===== Excel 批量工具合集（全递归版）=====")
    print("1 → 核查对比（遍历所有子文件夹）")
    print("2 → 批量连续修改单元格（遍历所有子文件夹）")
    print("3 → 批量替换文件名称（遍历所有子文件夹）")
    print("========================================")
    c = input("请选择功能序号：").strip()
    if c == "1":
        batch_check_fast_compare()
    elif c == "2":
        batch_modify_cell()
    elif c == "3":
        batch_rename_files()
    else:
        print("❌ 输入错误，请输入1/2/3")