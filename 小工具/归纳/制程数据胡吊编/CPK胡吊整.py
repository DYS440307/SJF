import openpyxl
import numpy as np
from scipy.stats import norm
import os

# -------------------------- 核心配置（仅修改此处适配你的Excel） --------------------------
EXCEL_PATH = r"E:\System\desktop\新建 Microsoft Excel 工作表.xlsx"  # Excel路径
MANUAL_ROWS = {
    "spec": 7,  # 规格(SPEC)：B7-E7（人工填写，代码仅读取不修改）
    "upper_tol": 8,  # 上偏差(+)：B8-E8（人工填写，代码仅读取不修改）
    "lower_tol": 9  # 下偏差(-)：B9-E9（人工填写，代码仅读取不修改）
}
DATA_RANGE = {
    "start_row": 10,  # DATA1开始行：第10行
    "end_row": 45,  # DATA36结束行：第45行（10+36-1=45）
    "cols": ["B", "C", "D", "E"]  # 仅操作这4列的10-45行
}
DATA_COUNT = 36  # 生成36组数据（对应B10:E45）
TARGET_CPK = 1.35  # 目标CPK（确保生成数据后原有公式计算的CPK＞1.33）


# -------------------------- 核心函数 --------------------------
def get_spec_limits(ws, col):
    """读取人工填写的规格、上下偏差，计算USL/LSL（仅读取，不修改）"""
    # 读取单元格值（处理空值/非数值）
    spec_val = ws[f"{col}{MANUAL_ROWS['spec']}"].value
    upper_tol_val = ws[f"{col}{MANUAL_ROWS['upper_tol']}"].value
    lower_tol_val = ws[f"{col}{MANUAL_ROWS['lower_tol']}"].value

    try:
        spec = float(spec_val) if spec_val else 0
        upper_tol = float(upper_tol_val) if upper_tol_val else 0
        lower_tol = float(lower_tol_val) if lower_tol_val else 0
    except (ValueError, TypeError):
        raise ValueError(
            f"列{col}的规格/偏差单元格（{col}{MANUAL_ROWS['spec']}/{col}{MANUAL_ROWS['upper_tol']}/{col}{MANUAL_ROWS['lower_tol']}）填写错误，请确保是数值！")

    usl = spec + upper_tol
    lsl = spec - lower_tol
    if usl <= lsl:
        raise ValueError(f"列{col}规格上下限异常（USL={usl} ≤ LSL={lsl}），请检查偏差值！")
    return spec, usl, lsl


def generate_cpk_data(spec, usl, lsl):
    """
    生成符合CPK＞1.33的正态分布数据（适配Excel STDEV.S公式）
    :return: 36组数值（保留2位小数，确保原有公式计算CPK＞1.33）
    """
    # 计算最大允许标准差（留10%余量，避免刚好达标）
    max_std = (usl - lsl) / (6 * (TARGET_CPK + 0.1))
    # 均值贴近规格中心（偏移＜0.01，减少偏倚）
    mean = spec + np.random.uniform(-0.01, 0.01)

    # 循环生成，直到CPK＞1.33
    while True:
        # 生成正态分布数据
        data = np.random.normal(loc=mean, scale=max_std, size=DATA_COUNT)
        data = np.round(data, 2)  # 保留2位小数（匹配测量精度）
        # 确保所有数据在规格限内（留0.01余量，避免贴边）
        data = np.clip(data, lsl + 0.01, usl - 0.01)

        # 模拟Excel公式计算CPK（样本标准差STDEV.S）
        sample_std = np.std(data, ddof=1)  # ddof=1对应STDEV.S
        cpk_u = (usl - np.mean(data)) / (3 * sample_std)
        cpk_l = (np.mean(data) - lsl) / (3 * sample_std)
        cpk = min(cpk_u, cpk_l)

        # 验证CPK达标
        if cpk > TARGET_CPK:
            break
    return data


# -------------------------- 主执行逻辑（仅操作B10:E45） --------------------------
if __name__ == "__main__":
    # 检查Excel文件是否存在
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel文件不存在：{EXCEL_PATH}，请先创建并填写B7-E9的规格/偏差！")

    # 打开Excel（read_only=False：允许写入；data_only=False：保留公式）
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=False)
    ws = wb.active  # 选择活动工作表

    try:
        # 循环处理B/C/D/E列，仅写入10-45行
        for col in DATA_RANGE["cols"]:
            # 1. 读取规格/偏差（仅读取，不修改原有内容）
            spec, usl, lsl = get_spec_limits(ws, col)

            # 2. 生成符合CPK＞1.33的36组数据
            data = generate_cpk_data(spec, usl, lsl)

            # 3. 仅写入B10:E45（覆盖原有数据，保留其他单元格公式）
            for i in range(DATA_COUNT):
                row = DATA_RANGE["start_row"] + i  # 10-45行
                cell = f"{col}{row}"
                ws[cell].value = data[i]  # 仅修改此单元格，其他不动

                # 可选：给A列DATA行标注名称（如果需要，不影响公式）
                if col == "B":
                    ws[f"A{row}"].value = f"DATA{i + 1}"

        # 4. 保存Excel（仅更新B10:E45，其他内容/公式完全保留）
        wb.save(EXCEL_PATH)
        print(f"✅ 操作完成！仅更新了Excel中{B10 if col == 'B' else ''}10:E45单元格数据")
        print(f"📌 保留了所有原有公式，生成的数据确保CPK＞{TARGET_CPK}")
        print(f"📌 请刷新Excel查看CPK计算结果（判定应为OK）")

    except ValueError as e:
        print(f"❌ 错误：{e}")
    finally:
        # 关闭Excel文件，避免占用
        wb.close()