import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

# ==============================
# 配置区域
# ==============================
excel_path = r"E:\System\download\声乐设备清单_2025年.xlsx"
output_folder = r"E:\System\download\图片文件夹"

# 确保输出文件夹存在
os.makedirs(output_folder, exist_ok=True)

# ==============================
# 清理文件名非法字符的函数
# ==============================
def clean_filename(filename):
    if not filename:
        return ""
    # 替换换行符、制表符等空白字符为下划线
    filename = re.sub(r'[\n\r\t]', '_', filename)
    # 移除Windows非法字符（\/:*?"<>|）
    filename = re.sub(r'[\\/:*?"<>|]', '', filename)
    # 替换多个下划线为单个，去除首尾下划线
    filename = re.sub(r'_+', '_', filename).strip('_')
    # 限制文件名长度（避免超长）
    return filename[:50]  # 可根据需要调整长度

# ==============================
# 打开Excel文件
# ==============================
wb = load_workbook(excel_path)
ws = wb.active

# 统计提取数量
extracted_count = 0

# 获取所有图片
for image in ws._images:
    try:
        # 图片锚点信息，确定所在单元格
        anchor = image.anchor._from
        row = anchor.row + 1  # openpyxl内部是从0开始的
        col = anchor.col + 1

        # 仅处理第2列（B列）
        if col != 2:
            continue

        # 获取序号和设备名称
        seq = ws.cell(row=row, column=1).value  # A列 序号
        name = ws.cell(row=row, column=3).value  # C列 设备名称

        # 校验基础数据
        if not seq or not name:
            print(f"⚠️ 第{row}行：序号/设备名称为空，跳过")
            continue

        # 清理序号和名称中的非法字符
        clean_seq = clean_filename(str(seq))
        clean_name = clean_filename(str(name))

        # 构建最终文件名（避免空文件名）
        if not clean_seq or not clean_name:
            print(f"⚠️ 第{row}行：清理后序号/名称为空，跳过")
            continue
        filename = f"{clean_seq}_{clean_name}.jpg"
        filepath = os.path.join(output_folder, filename)

        # 保存图片（处理PIL可能的格式问题）
        img_data = image._data()
        with PILImage.open(BytesIO(img_data)) as img:
            # 处理透明图片（PNG转JPG需填充白色背景）
            if img.mode in ('RGBA', 'P'):
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                background.save(filepath, "JPEG", quality=95)
            else:
                img.convert("RGB").save(filepath, "JPEG", quality=95)

        extracted_count += 1
        print(f"✅ 已提取图片: {filepath}")

    except Exception as e:
        print(f"❌ 第{row}行处理失败：{str(e)}")
        continue

# 最终统计
print(f"\n🎯 提取完成！共成功提取 {extracted_count} 张图片")