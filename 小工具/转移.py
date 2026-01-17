import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# 1. 配置参数（核对准确！）
source_file_path = r"E:\System\download\供应商索赔清单 .xlsx"
template_sheet_name = "扣款通知单1"  # 模板工作表
list_sheet_name = "索赔清单"  # 索赔清单工作表
output_file_path = r"E:\System\download\供应商索赔清单_pandas生成.xlsx"  # 输出文件

# 2. 模板单元格映射（列字母+行号，核对后修改）
CELL_MAP = {
    "supplier": "B3",  # 供方名称
    "amount": "E3",  # 质量保证金
    "abnormal": "B6"  # 异常描述
}


def batch_generate_with_pandas():
    # 校验源文件是否存在
    if not os.path.exists(source_file_path):
        print(f"❌ 源文件 {source_file_path} 不存在")
        return

    # ========== 步骤1：用pandas读取索赔清单数据 ==========
    try:
        # 读取索赔清单（跳过表头，header=0表示第1行是表头）
        df_claim = pd.read_excel(
            source_file_path,
            sheet_name=list_sheet_name,
            header=0,
            usecols=["供方名称", "异常描述", "质量保证金(¥)"]  # 只读取需要的列
        )
        # 清理空值和特殊字符
        df_claim = df_claim.dropna()  # 删除空行
        df_claim["供方名称"] = df_claim["供方名称"].astype(str).str.strip()
        df_claim["异常描述"] = df_claim["异常描述"].astype(str).str.strip()
        df_claim["质量保证金(¥)"] = df_claim["质量保证金(¥)"].astype(str).str.strip()
        print(f"📊 pandas读取到 {len(df_claim)} 条有效数据")
        print("📌 前5条数据预览：")
        print(df_claim.head())

        if len(df_claim) == 0:
            print("❌ 索赔清单无有效数据")
            return
    except Exception as e:
        print(f"❌ 读取索赔清单失败：{e}")
        return

    # ========== 步骤2：加载模板工作簿，批量生成通知单 ==========
    try:
        # 加载源工作簿（保留格式，data_only=False）
        wb = load_workbook(source_file_path, read_only=False, data_only=False)

        # 校验模板工作表是否存在
        if template_sheet_name not in wb.sheetnames:
            print(f"❌ 未找到模板工作表「{template_sheet_name}」")
            wb.close()
            return

        # 获取模板工作表
        template_ws = wb[template_sheet_name]

        # ========== 步骤3：遍历数据，复制模板并填充数据 ==========
        for idx, row in df_claim.iterrows():
            # 复制模板工作表（注意：pandas+openpyxl仍无法保留图片）
            new_sheet_name = f"扣款通知单{idx + 1}"
            # 若已存在同名工作表，删除后重建
            if new_sheet_name in wb.sheetnames:
                del wb[new_sheet_name]
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = new_sheet_name

            # 填充数据到指定单元格
            new_ws[CELL_MAP["supplier"]] = row["供方名称"]
            new_ws[CELL_MAP["amount"]] = row["质量保证金(¥)"]
            new_ws[CELL_MAP["abnormal"]] = row["异常描述"]

            print(f"✅ 已生成：{new_sheet_name}（供方：{row['供方名称']}）")

        # ========== 步骤4：保存新文件 ==========
        wb.save(output_file_path)
        wb.close()
        print(f"\n🎉 生成完成！文件保存至：{output_file_path}")
        print("⚠️ 注意：pandas+openpyxl无法保留Excel中的图片/logo/手写签名，仅保留单元格数据和基础格式")

    except Exception as e:
        print(f"❌ 生成通知单失败：{e}")
        wb.close() if 'wb' in locals() else None


if __name__ == "__main__":
    batch_generate_with_pandas()