import os
from openpyxl import load_workbook

# 定义源目录和目标目录
source_dir = r"F:\system\Desktop\奥克斯语音模组"
template_dir = "F:\\system\\Desktop\\PY\\语音模组\\模板"
destination_dir = "F:\\system\\Desktop\\PY\\语音模组\\Test"

# 遍历源目录中的所有文件
for file_name in os.listdir(source_dir):
    file_path = os.path.join(source_dir, file_name)

    # 加载工作簿并获取'K8'和'E6'单元格的值
    wb = load_workbook(file_path)
    sheet = wb.active
    k8_value = sheet['K8'].value
    e6_value = sheet['E6'].value

    # 从文件名中提取Temp1，不包括.xlsx扩展名
    temp1 = os.path.splitext(file_name.split('_')[-1])[0]

    # 遍历模板目录中的所有文件
    for template_file in os.listdir(template_dir):
        template_path = os.path.join(template_dir, template_file)

        # 加载模板工作簿
        template_wb = load_workbook(template_path)
        template_sheet = template_wb.active

        # 将值复制到'G3'和'B4'单元格
        template_sheet['G3'] = k8_value
        template_sheet['B4'] = e6_value

        # 将值复制到'G3'和'B4'单元格
        template_sheet['G3'] = k8_value
        template_sheet['B4'] = e6_value

        # 根据B4单元格的值设置E4单元格的值
        if e6_value == '3613-P':
            template_sheet['E4'] = 113291230000008
        elif e6_value == 'ESP-32-DU2306':
            template_sheet['E4'] = 113291230000009

        # 将Temp1复制到'L4'单元格
        template_sheet['L4'] = temp1

        # 创建目标目录的新文件夹
        temp1_folder = os.path.join(destination_dir, temp1)
        if not os.path.exists(temp1_folder):
            os.makedirs(temp1_folder)

        # 将修改后的文件保存到目标目录的新文件夹中
        new_file_path = os.path.join(temp1_folder, template_file)
        template_wb.save(new_file_path)

print("所有文件已处理并保存到目标目录。")
