import openpyxl
from datetime import datetime

# 忽略"-"的匹配函数
def normalize_string(s):
    if isinstance(s, str):  # 只有是字符串时才去掉 '-'
        return s.replace('-', '')
    elif isinstance(s, datetime):  # 如果是 datetime 对象，返回字符串形式
        return s.strftime('%Y-%m-%d %H:%M:%S')  # 根据需要调整日期格式
    return ''  # 如果不是字符串或日期，返回空字符串

# 加载工作簿和工作表
file1 = r'F:\system\Pictures\转中\文件1.xlsx'
file2 = r'F:\system\Pictures\转中\文件2.xlsx'

wb1 = openpyxl.load_workbook(file1)
wb2 = openpyxl.load_workbook(file2)

ws1 = wb1.active
ws2 = wb2.active

# 遍历文件2的C列
for row2 in range(2, ws2.max_row + 1):  # 从第二行开始遍历文件2的C列
    value2 = ws2[f'C{row2}'].value
    normalized_value2 = normalize_string(value2)

    # 遍历文件1的C列，进行部分匹配
    for row1 in range(2, ws1.max_row + 1):  # 从第二行开始遍历文件1的C列
        value1 = ws1[f'C{row1}'].value
        normalized_value1 = normalize_string(value1)

        # 部分匹配
        if normalized_value1 and normalized_value2 and normalized_value1 in normalized_value2:
            # 将文件1的D列的值写入文件2对应的D列
            ws2[f'D{row2}'] = ws1[f'D{row1}'].value
            print(f'匹配成功: 文件2 C{row2} 包含文件1 C{row1}，将文件1 D{row1} 的值写入文件2 D{row2}')
            break  # 找到匹配项后跳出循环

# 保存文件2
wb2.save(file2)
print("操作完成，文件已保存。")
