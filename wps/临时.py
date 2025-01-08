import random
from openpyxl import load_workbook

# Excel 文件路径
file_path = r"F:\system\Pictures\转中\工作簿1.xlsx"

# 加载工作簿和工作表
workbook = load_workbook(file_path)
sheet = workbook.active  # 如果是特定工作表，替换为 workbook["工作表名称"]

# 遍历偶数列，第15行到第31行
for row in range(15, 32):  # 包括第31行
    for col in range(2, sheet.max_column + 1, 2):  # 偶数列，从第2列开始，步长为2
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)) and cell.value > 9:
            cell.value = random.uniform(0, 8)  # 生成0到8之间的小数

# 保存工作簿
workbook.save(file_path)
print("数据修改完成！")
