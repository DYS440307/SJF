import os
from openpyxl import load_workbook

# 源文件夹路径
source_folder = "F:\\system\\Desktop\\Temp\\2024"
# 目标模板文件路径
template_file = "F:\\system\\Desktop\\Temp\\奥克斯语音模组检验报告单模板.xlsx"
# 新文件保存路径
new_folder = "F:\\system\\Desktop\\Temp\\2024新"

# 确保新文件夹存在
if not os.path.exists(new_folder):
    os.makedirs(new_folder)

# 遍历源文件夹中的所有Excel文件
for file_name in os.listdir(source_folder):
    if file_name.endswith('.xlsx'):
        # 加载源Excel文件
        source_path = os.path.join(source_folder, file_name)
        source_wb = load_workbook(source_path)
        source_ws = source_wb.active

        # 加载目标模板文件
        target_wb = load_workbook(template_file)
        target_ws = target_wb.active

        # 复制单元格
        target_ws['J4'] = source_ws['J4'].value
        target_ws['E6'] = source_ws['E6'].value
        target_ws['E7'] = source_ws['E7'].value
        target_ws['K7'] = source_ws['K7'].value

        # 生成新文件名，去除日期中的冒号和时分秒部分
        date_str = str(source_ws['K8'].value)[:10]  # 截取日期部分，格式为 YYYY-MM-DD
        new_file_name = "奥克斯语音模组_" + date_str.replace('-', '') + ".xlsx"
        new_file_path = os.path.join(new_folder, new_file_name)

        # 保存新文件
        target_wb.save(new_file_path)

        print(f"文件 {file_name} 已处理并保存为 {new_file_name}")

print("所有文件处理完毕。")