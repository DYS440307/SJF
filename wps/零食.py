from openpyxl import load_workbook

# 文件路径
file_path = r"F:\system\Downloads\记录表.xlsx"

# 加载工作簿和第一个工作表
wb = load_workbook(file_path)
sheet1 = wb.worksheets[0]

# 如果第二个工作表不存在，则创建
if len(wb.sheetnames) < 2:
    wb.create_sheet(title="结果")
sheet2 = wb.worksheets[1]

# 清空第二个工作表
for row in sheet2.iter_rows():
    for cell in row:
        cell.value = None

# 获取最大行和列数
max_row = sheet1.max_row
max_col = sheet1.max_column

# 遍历所有偶数列与前一奇数列配对
result_row = 1
for col in range(2, max_col + 1, 2):  # 遍历偶数列
    max_value = None
    corresponding_value = None

    # 找到偶数列中的最大值及其对应的奇数列值
    for row in range(1, max_row + 1):
        value = sheet1.cell(row=row, column=col).value
        if value is not None and (max_value is None or value > max_value):
            max_value = value
            corresponding_value = sheet1.cell(row=row, column=col - 1).value

    # 将对应的奇数列值写入第二个工作表
    if corresponding_value is not None:
        sheet2.cell(row=result_row, column=1).value = corresponding_value
        result_row += 1

# 保存工作簿
wb.save(file_path)
print("操作完成，结果已写入到第二个工作表中！")
