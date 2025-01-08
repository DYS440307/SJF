import openpyxl
from openpyxl.utils import get_column_letter


def merge_sheets(file_path):
    # 打开Excel文件
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames

    # 获取第一个工作表
    main_sheet = wb[sheets[0]]

    # 找到第一个工作表的最后一列
    main_last_col = main_sheet.max_column

    # 遍历其他工作表，将内容复制到第一个工作表的最右边一列
    for sheet_name in sheets[1:]:
        sheet = wb[sheet_name]

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_index, value in enumerate(row, start=1):
                main_sheet.cell(row=row_index, column=main_last_col + col_index, value=value)

        # 更新最后一列的位置
        main_last_col += sheet.max_column

    # 删除其余工作表
    for sheet_name in sheets[1:]:
        del wb[sheet_name]

    # 保存为新文件
    new_file_path = r"F:\\system\\Downloads\\工作簿2_合并后.xlsx"
    wb.save(new_file_path)
    print(f"所有内容已成功拼接到第一个工作表，并保存到 {new_file_path}")


# 使用示例
file_path = r"F:\\system\\Downloads\\工作簿2.xlsx"
merge_sheets(file_path)