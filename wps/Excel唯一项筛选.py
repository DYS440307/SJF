import openpyxl

# 提示用户输入Excel文件的路径
file_path = input("请输入Excel文件的路径：")

try:
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 获取工作表
    if '7月份' not in workbook.sheetnames:
        raise ValueError("工作表 '7月份' 不存在")
    sheet = workbook['7月份']

    # 存储已经出现过的组
    seen = set()
    rows_to_keep = []

    # 遍历工作表的每一行，跳过表头
    for row in sheet.iter_rows(min_row=2, values_only=True):
        group = (row[2], row[10])  # 第三列和第十一列的组合
        if group not in seen:
            seen.add(group)
            rows_to_keep.append(row)

    # 创建一个新的工作簿并添加工作表
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = '7月份'

    # 复制表头
    for cell in sheet[1]:
        new_sheet.cell(row=1, column=cell.column, value=cell.value)

    # 将保留的行写入新的工作表
    for row_idx, row in enumerate(rows_to_keep, start=2):
        for col_idx, value in enumerate(row, start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=value)

    # 保存新的工作簿
    output_path = file_path.replace('.xlsx', '.xlsx')
    new_workbook.save(output_path)
    print(f"处理后的数据已保存到 {output_path}")

except Exception as e:
    print(f"发生错误：{e}")
