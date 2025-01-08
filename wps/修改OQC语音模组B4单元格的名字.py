import os
from openpyxl import load_workbook

def modify_excel_files(directory):
    # 遍历目录及其子文件夹
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):  # 只处理Excel文件
                filepath = os.path.join(root, file)
                try:
                    wb = load_workbook(filepath)
                    ws = wb.active
                    # 遍历每个表格中的数据
                    for row in ws.iter_rows(min_row=4, min_col=2, max_col=2):
                        for cell in row:
                            if cell.value == 'VOI-M7':
                                ws['E' + str(cell.row)] = '112220720005'
                            elif cell.value == '3613-P':
                                ws['E' + str(cell.row)] = '113291230000008'
                            elif cell.value == 'ESP-32-DU2306':
                                ws['E' + str(cell.row)] = '113291230000009'
                    wb.save(filepath)
                    print(f"文件 '{file}' 修改成功！")
                except Exception as e:
                    print(f"处理文件 '{file}' 时出错：{e}")

# 修改指定目录下的Excel文件
modify_excel_files(r"F:\system\Desktop\语音模组")
