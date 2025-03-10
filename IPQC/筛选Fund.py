import openpyxl

# 打开工作簿和工作表
filepath = r'F:\system\Downloads\1.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

col = 2  # 从偶数列开始（B列）
while col <= ws.max_column:
    delete_flag = False

    # 原逻辑：检查整列是否有 <35 或 >95 的值
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col).value
        try:
            if cell_value is not None:
                value = float(cell_value)
                if value < 35 or value > 95:
                    delete_flag = True
                    break
        except ValueError:
            continue

    # 新增逻辑：再检查第6~27行是否有值不在70~95之间
    if not delete_flag:
        for row in range(6, 28):  # 注意range(6, 28)表示6~27行
            cell_value = ws.cell(row=row, column=col).value
            try:
                if cell_value is not None:
                    value = float(cell_value)
                    if value < 70 or value > 95:
                        delete_flag = True
                        break
            except ValueError:
                continue

    # 满足任一条件就删除该奇偶列组
    if delete_flag:
        ws.delete_cols(col - 1, 2)  # 删除对应奇数列+偶数列
        col = col - 2  # 回退2列以防跳过
        if col < 2:
            col = 2
    else:
        col += 2  # 检查下一个偶数列

# 保存工作簿
wb.save(filepath)
print("处理完成，已删除不符合条件的列组。")
