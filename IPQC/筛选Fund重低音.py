from openpyxl import load_workbook

# 读取 Excel 文件
file_path = r"E:\System\pic\1.xlsx"
wb = load_workbook(file_path)
ws = wb.active  # 默认操作第一个工作表

# 获取当前工作表的最大列数
max_col = ws.max_column

# 为了避免删除列时导致后续列索引发生变化，从最后一列向前遍历
for col in range(max_col, 1, -1):  # 从第二列开始（即列索引>=2），故range最后值为1
    delete_flag = False
    # 遍历第2行到第82行（Excel行号从1开始，包含82）
    for row in range(2, 94):
        cell = ws.cell(row=row, column=col)
        try:
            # 尝试将单元格值转换为浮点数进行比较
            value = float(cell.value)
        except (TypeError, ValueError):
            value = None
        if value is not None and (value < 40 or value > 100):
            delete_flag = True
            break  # 若发现一个单元格满足条件，则该列标记为删除

    if delete_flag:
        ws.delete_cols(col)
        print(f"已删除第 {col} 列")

# 保存工作簿
wb.save(file_path)
print("处理完成，已删除不符合条件的列组。")
