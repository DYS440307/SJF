from openpyxl import load_workbook

# 读取 Excel 文件
file_path = r"E:\System\pic\1.xlsx"
wb = load_workbook(file_path)
ws = wb.active  # 默认操作第一个工作表

# 获取当前工作表的最大列数
max_col = ws.max_column

# 为了避免删除列时导致后续列索引发生变化，从最后一列向前遍历
for col in range(max_col, 1, -1):  # 从第二列开始（即列索引>=2）
    delete_flag = False

    # 判断第2行到第82行，要求数值在0~15之间
    for row in range(2, 83):  # 包含第82行
        cell = ws.cell(row=row, column=col)
        try:
            value = float(cell.value)
        except (TypeError, ValueError):
            value = None
        if value is not None and (value < 0 or value > 15):
            delete_flag = True
            break  # 若发现不符合条件，立即标记该列为删除

    # 若前面的行未标记删除，再判断第83行到第94行，要求数值在0~10之间
    if not delete_flag:
        for row in range(83, 95):  # 包含第94行
            cell = ws.cell(row=row, column=col)
            try:
                value = float(cell.value)
            except (TypeError, ValueError):
                value = None
            if value is not None and (value < 0 or value > 10):
                delete_flag = True
                break

    if delete_flag:
        ws.delete_cols(col)
        print(f"已删除第 {col} 列")

# 保存工作簿
wb.save(file_path)
print("处理完成，已删除不符合条件的列组。")
