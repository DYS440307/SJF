from openpyxl import load_workbook
from itertools import combinations
from math import comb

# --- 配置区 ---
file_path     = r"E:\System\desktop\工作簿1.xlsx"  # Excel 文件路径
sheet_name    = None       # None 表示第一个 sheet，否则写表名
select_k      = 20         # 要选出的列数
enum_cutoff   = 20         # 当总列数 <= 该值且 C(N, k) 合理时使用全枚举

# --- 加载工作簿 & 表格 ---
wb = load_workbook(file_path, data_only=True)
ws = wb[sheet_name] if sheet_name else wb.active

# --- 读取第一行标签 & 数据 ---
labels = [cell.value for cell in ws[1]]      # 0-based 列序号对应的“标签”
n_cols = ws.max_column
n_rows = ws.max_row - 1                      # 排除表头
# 将每列的数值读成列表；空或非数值按 0 处理
cols = []
for c in range(1, n_cols+1):
    col_data = []
    for r in range(2, ws.max_row+1):
        v = ws.cell(r, c).value
        col_data.append(v if isinstance(v, (int, float)) else 0)
    cols.append(col_data)

# --- 计算给定一组列索引 combo 的“总离散度” ---
def total_dispersion(combo):
    """对 combo（列索引元组），累加每行的 max-min"""
    s = 0
    for row_idx in range(n_rows):
        row_vals = [cols[c][row_idx] for c in combo]
        s += (max(row_vals) - min(row_vals))
    return s

# --- 主流程：全枚举 or 贪心启发 ---
best_set, best_score = None, float('inf')

# 决定是否做全枚举
if n_cols <= enum_cutoff and comb(n_cols, select_k) <= 5e6:
    # 全枚举
    for combo in combinations(range(n_cols), select_k):
        disp = total_dispersion(combo)
        if disp < best_score:
            best_score, best_set = disp, combo
else:
    # 贪心启发式
    remaining = set(range(n_cols))
    # 初始：选一对离散度最小的两列
    min_pair = min(
        ((i,j) for i in range(n_cols) for j in range(i+1, n_cols)),
        key=lambda p: total_dispersion(p)
    )
    best_set = set(min_pair)
    remaining -= best_set
    # 逐步加入：每次从剩余列中选一个，使加入后对当前集合的总离散度增量最小
    while len(best_set) < select_k:
        cand, cand_inc = None, float('inf')
        for j in remaining:
            inc = total_dispersion(tuple(sorted(best_set | {j})))
            if inc < cand_inc:
                cand_inc, cand = inc, j
        best_set.add(cand)
        remaining.remove(cand)
    best_set = tuple(sorted(best_set))
    best_score = total_dispersion(best_set)

# --- 输出 ---
print("选出的列索引 (0-based)：", best_set)
print("对应第一行标签：", [labels[i] for i in best_set])
print("总离散度（各行 max-min 之和）：", best_score)
