import openpyxl

# 加载工作簿和工作表
filepath = r'F:\system\Downloads\1.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb.active

col = 2  # 从偶数列开始
while col <= ws.max_column:
    delete_flag = False
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col).value
        try:
            if cell_value is not None:
                value = float(cell_value)
                if value < 0 or value > 10:
                    delete_flag = True
                    break
        except ValueError:
            continue  # 跳过无法转成数值的单元格

    if delete_flag:
        # 删除对应奇数列和偶数列
        ws.delete_cols(col - 1, 2)
        col = col - 2  # 回退2列，重新判断当前位置
        if col < 2:
            col = 2
    else:
        col += 2  # 若未删除则继续下一个偶数列

# 保存修改
wb.save(filepath)
print("处理完成，已删除不符合条件的列组。")
