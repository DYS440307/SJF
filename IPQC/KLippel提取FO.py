import os
import openpyxl

# --- 路径设置 与 验证 ---
source_path = r'E:\System\pic\1.xlsx'
if not os.path.exists(source_path):
    raise FileNotFoundError(
        f"未找到文件：{source_path}\n"
        f"该目录下的文件列表为：{os.listdir(os.path.dirname(source_path))}"
    )

# --- 打开工作簿与工作表 ---
wb = openpyxl.load_workbook(source_path)

# 显式指定原始数据表名（请根据实际修改）
src_sheet_name = '原始数据'
if src_sheet_name not in wb.sheetnames:
    raise ValueError(f"未找到名为 '{src_sheet_name}' 的工作表！")
ws_source = wb[src_sheet_name]

# 获取或新建 FO提取 表，并清空旧内容（可选）
fo_name = 'FO提取'
if fo_name in wb.sheetnames:
    ws_fo = wb[fo_name]
    for row in ws_fo.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_fo = wb.create_sheet(title=fo_name)

# 数据范围
max_col = ws_source.max_column
max_row = ws_source.max_row

write_row = 1  # 写入行，从第1行开始

# 从第2列开始处理
for col in range(2, max_col + 1):
    max_value = None
    candidate_rows = []

    # 找最大值及对应行
    for row in range(1, max_row + 1):
        val = ws_source.cell(row=row, column=col).value
        if isinstance(val, (int, float)):
            if max_value is None or val > max_value:
                max_value = val
                candidate_rows = [row]
            elif val == max_value:
                candidate_rows.append(row)

    if not candidate_rows:
        continue

    # 在这些行里，选第一列最小的那一行
    min_first = None
    sel_row = None
    for r in candidate_rows:
        first = ws_source.cell(row=r, column=1).value
        if isinstance(first, (int, float)):
            if min_first is None or first < min_first:
                min_first = first
                sel_row = r

    if sel_row is None:
        continue

    # 写入到 FO提取 的第1列
    ws_fo.cell(row=write_row, column=1, value=min_first)
    write_row += 1

# 保存
wb.save(source_path)
print("处理完成！已将提取结果写入 'FO提取' 表的第一列。")
