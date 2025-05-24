from openpyxl import load_workbook

# 加载工作簿
file_path = r"E:\System\pic\1.xlsx"
wb = load_workbook(file_path)

# 获取工作表
ws_thd = wb["THD"]
ws_fund = wb["Fund"]
ws_imp = wb["IMP"]

# 初始化列索引（从第2列开始）和删除计数
col = 2
deleted_count = 0

# 注意：删除列时，后续列的索引会向前移动，所以处理时不应该预先增加 col
while col <= ws_thd.max_column:
    delete_flag = False
    # 检查第22~100行该列是否有值大于35
    for row in range(55, 83):
        cell_value = ws_thd.cell(row=row, column=col).value
        if isinstance(cell_value, (int, float)) and cell_value > 10:
            delete_flag = True
            break

    if delete_flag:
        # 删除 THD, Fund, IMP 中的对应列
        ws_thd.delete_cols(col)
        ws_fund.delete_cols(col)
        ws_imp.delete_cols(col)
        deleted_count += 1
        # 删除后当前列变成新的列，不递增
    else:
        col += 1

# 保存工作簿
wb.save(file_path)

# 输出结果
print(f"THD 工作表中删除的列数：{deleted_count}")
