from openpyxl import load_workbook

# 加载工作簿
file_path = 'E:/System/pic/1.xlsx'
wb = load_workbook(file_path)

# 获取工作表
imp_ws = wb['IMP']
fund_ws = wb['Fund']
thd_ws = wb['THD']

# 存储需要删除的列（从第2列开始）
cols_to_delete = []

# 从后往前遍历列（避免索引偏移）
for col in range(imp_ws.max_column, 1, -1):
    delete_flag = False

    # 条件1：第2~14行中是否有值 >7
    for row in range(2, 43):
        val = imp_ws.cell(row=row, column=col).value
        if isinstance(val, (int, float)) and val > 7:
            delete_flag = True
            break
    # 条件1：第2~14行中是否有值 >7
    for row in range(62, 75):
        val = imp_ws.cell(row=row, column=col).value
        if isinstance(val, (int, float)) and val > 6.3:
            delete_flag = True
            break

        # 条件1：第2~14行中是否有值 >7
        for row in range(22, 38):
            val = imp_ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)) and val >6:
                delete_flag = True
                break

            # 条件1：第2~14行中是否有值 >7
            for row in range(22, 38):
                val = imp_ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)) and val < 5.55:
                    delete_flag = True
                    break


    # 条件2：第14~110行是否有值 >50
    if not delete_flag:
        for row in range(2, 94):
            val = imp_ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)) and val > 10:
                delete_flag = True
                break

    # 如果满足任何条件，就删除
    if delete_flag:
        cols_to_delete.append(col)
        imp_ws.delete_cols(col)

# 在 Fund 和 THD 中删除相应列（从后往前）
cols_to_delete.sort(reverse=True)
for col in cols_to_delete:
    fund_ws.delete_cols(col)
    thd_ws.delete_cols(col)

# 输出信息
print(f"总共删除了 {len(cols_to_delete)} 列。")

# 保存修改
wb.save(file_path)
