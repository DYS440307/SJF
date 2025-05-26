from openpyxl import load_workbook

# 加载工作簿
file_path = r"E:\System\pic\全\中音\原档筛选_重复项核查_300PCS.xlsx"
wb = load_workbook(file_path)

# 获取工作表列表
sheetnames = wb.sheetnames

# 记录所有需要删除的列位置（以列号为基准）
# 格式: {sheetname: set([col1, col2, ...])}
delete_map = {name: set() for name in sheetnames}

# 处理每个工作表
for current_sheet in sheetnames:
    ws = wb[current_sheet]
    max_col = ws.max_column
    max_row = ws.max_row

    # 提取每列数据
    columns = []
    for col in range(1, max_col + 1):
        col_data = [ws.cell(row=row, column=col).value for row in range(1, max_row + 1)]
        columns.append(col_data)

    # 查找重复列（保留最前面一个）
    for i in range(len(columns)):
        for j in range(i + 1, len(columns)):
            if columns[i] == columns[j]:
                # 当前表删除列 j+1
                delete_map[current_sheet].add(j + 1)
                # 其余表也删除对应列号
                for other_sheet in sheetnames:
                    if other_sheet != current_sheet:
                        delete_map[other_sheet].add(j + 1)

# 对每个表执行删除操作（按列号倒序）
for sheetname in sheetnames:
    ws = wb[sheetname]
    for col_idx in sorted(delete_map[sheetname], reverse=True):
        ws.delete_cols(col_idx)

# 保存
wb.save(file_path)
print("处理完成，重复列已识别并在三个表中统一删除。")
