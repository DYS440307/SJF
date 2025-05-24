from openpyxl import load_workbook

# 加载工作簿
file_path = 'E:/System/pic/1.xlsx'
wb = load_workbook(file_path)
imp_ws = wb['IMP']
fund_ws = wb['Fund']
thd_ws = wb['THD']

# 将 IMP 表中所有数据读入内存（二维数组）
imp_data = [[imp_ws.cell(row=r, column=c).value for c in range(1, imp_ws.max_column + 1)]
            for r in range(1, imp_ws.max_row + 1)]

cols_to_delete = []

# 遍历每列（从后往前）
for col in range(len(imp_data[0]) - 1, 0, -1):  # 索引从0开始，第2列是索引1
    delete_flag = False

    # 条件1：第2~14行中有值 >7
    for row in range(1, 14):  # Excel第2~14行，对应Python索引1~13
        val = imp_data[row][col]
        if isinstance(val, (int, float)) and val > 7:
            delete_flag = True
            break

    # 条件2a：第39~74行有值 >10
    if not delete_flag:
        for row in range(38, 74):  # 索引对应Excel第39~75行
            val = imp_data[row][col]
            if isinstance(val, (int, float)) and val > 10:
                delete_flag = True
                break

    # 条件2b：第15~27行中有值 <5
    if not delete_flag:
        for row in range(14, 27):  # 索引对应Excel第15~28行
            val = imp_data[row][col]
            if isinstance(val, (int, float)) and val < 5:
                delete_flag = True
                break

    if delete_flag:
        cols_to_delete.append(col + 1)  # openpyxl列索引从1开始

# 删除列（从后往前）
for col in sorted(cols_to_delete, reverse=True):
    imp_ws.delete_cols(col)
    fund_ws.delete_cols(col)
    thd_ws.delete_cols(col)

print(f"总共删除了 {len(cols_to_delete)} 列。")

# 保存文件
wb.save(file_path)
