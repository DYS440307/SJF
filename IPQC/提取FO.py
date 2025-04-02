import openpyxl

# 加载源Excel文件
file_path = r'C:\Users\SL\Downloads\1.xlsx'
wb = openpyxl.load_workbook(file_path)

# 获取第一个工作簿
ws = wb.active

# 获取FO提取工作簿
fo_ws = wb['FO提取']

# 设置开始写入的行
write_row = 1

# 遍历第二列到最后一列
for col in range(2, ws.max_column + 1):
    max_value = None
    max_value_row = None

    # 遍历14行到29行，找到该列中的最大值及对应的第一列数值
    for row in range(14, 30):
        current_value = ws.cell(row=row, column=col).value
        if max_value is None or current_value > max_value:
            max_value = current_value
            max_value_row = row

    # 获取对应的第一列数值
    first_column_value = ws.cell(row=max_value_row, column=1).value

    # 将对应的数值写入FO提取工作簿的第一列
    fo_ws.cell(row=write_row, column=1, value=first_column_value)

    # 写入下一行
    write_row += 1

# 保存修改后的源工作簿
wb.save(file_path)

print("操作完成，所有列的最大值对应的第一列数值已写入 'FO提取' 工作簿。")
