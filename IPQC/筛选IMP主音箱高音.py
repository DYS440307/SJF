from openpyxl import load_workbook

# 读取 Excel 文件
file_path = r"E:\System\pic\1.xlsx"
wb = load_workbook(file_path)

# 检查所需工作表是否都存在
required_sheets = ["IMP", "Fund", "THD"]
for sheet_name in required_sheets:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到名为 '{sheet_name}' 的工作表")

# 获取目标工作表
ws_imp = wb["IMP"]
ws_fund = wb["Fund"]
ws_thd = wb["THD"]

# 获取 IMP 工作表的最大列数
max_col = ws_imp.max_column

# 记录需要删除的列索引
cols_to_delete = []

# 从最后一列向前遍历，避免删除列时索引变化
for col in range(max_col, 1, -1):  # 从第2列开始
    delete_flag = False

    # 第2行到第93行要求值在0~40
    for row in range(2, 94):
        cell = ws_imp.cell(row=row, column=col)
        try:
            value = float(cell.value)
        except (TypeError, ValueError):
            value = None
        if value is not None and (value < 0 or value > 50):
            delete_flag = True
            break

    # 若前段通过，再检查第83行到94行要求值在0~20
    if not delete_flag:
        for row in range(83, 90):
            cell = ws_imp.cell(row=row, column=col)
            try:
                value = float(cell.value)
            except (TypeError, ValueError):
                value = None
            if value is not None and (value < 0 or value > 50):
                delete_flag = True
                break

    if delete_flag:
        cols_to_delete.append(col)

# 按记录顺序删除 IMP、Fund、THD 中的列（从后往前）
for col in cols_to_delete:
    ws_imp.delete_cols(col)
    ws_fund.delete_cols(col)
    ws_thd.delete_cols(col)
    print(f"已删除第 {col} 列")

# 保存文件
wb.save(file_path)
print("处理完成，已同步删除 IMP、Fund、THD 工作表中不符合条件的列。")
