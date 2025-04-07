import openpyxl

# 加载源Excel文件
file_path = r'E:\System\pic\1.xlsx'
wb = openpyxl.load_workbook(file_path)

# 获取第一个工作表
ws = wb.active

# 检查是否存在名为“FO提取”的工作表，如果不存在则创建
if 'FO提取' not in wb.sheetnames:
    fo_ws = wb.create_sheet('FO提取')
else:
    fo_ws = wb['FO提取']

# 设置开始写入的行
write_row = 1

# 遍历第二列到最后一列
for col in range(2, ws.max_column + 1):
    max_value = None
    max_value_row = None

    # 遍历第1行到第92行，找到该列中的最大值及对应的第一列数值
    for row in range(1, 93):
        current_value = ws.cell(row=row, column=col).value
        if current_value is not None:  # 确保 current_value 不是 None
            if max_value is None or current_value > max_value:
                max_value = current_value
                max_value_row = row

    if max_value_row is not None:  # 确保找到了有效的最大值
        # 获取对应的第一列数值
        first_column_value = ws.cell(row=max_value_row, column=1).value

        # 将对应的数值写入“FO提取”工作表的第一列
        fo_ws.cell(row=write_row, column=1, value=first_column_value)

        # 写入下一行
        write_row += 1



# 保存修改后的源工作簿
wb.save(file_path)

print("操作完成，所有列的最大值对应的第一列数值已写入 'FO提取' 工作表。")
