import openpyxl

# 文件路径
file_path = r'C:\Users\SL\Downloads\1.xlsx'

# 打开Excel文件
wb = openpyxl.load_workbook(file_path)

# 获取活动工作表
ws = wb.active

# 遍历第一列的偶数行
row_num = 2  # 从第二行开始（偶数行）
for i in range(2, ws.max_row + 1, 2):  # 从2开始步进2，选择偶数行
    value = ws.cell(row=i, column=1).value  # 获取第一列偶数行的值
    ws.cell(row=row_num, column=2, value=value)  # 将值写入第二列
    row_num += 1

# 删除第二列中的空单元格
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=2).value is None:
        ws.delete_rows(row)

# 保存修改后的文件
wb.save(file_path)

print("操作完成，第一列偶数行已复制到第二列，并删除了空单元格。")
