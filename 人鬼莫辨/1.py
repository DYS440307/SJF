import openpyxl
import random

# Load the Excel file
file_path = r"F:\system\Desktop\12302-500111_额定功率_可靠性实验报告_.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Extracting num from cell H16
num_str = sheet['H16'].value.split('≤')[1].strip()
num = float(num_str)

# Calculate lower and upper limits for range 1
lower_limit = num * 0.121
upper_limit = num * 0.55

# Generate random numbers for column H and I and write to cells H18:I22
for row in range(18, 23):
    random_num_H = round(random.uniform(lower_limit, upper_limit), 3)
    random_num_I = round(random.uniform(random_num_H, upper_limit), 3)  # Ensure I > H
    sheet.cell(row=row, column=8).value = random_num_H
    sheet.cell(row=row, column=9).value = random_num_I

    # Print the written result
    print(f"Row {row}: H{row} = {random_num_H}, I{row} = {random_num_I}")

# Save the modified Excel file
workbook.save(file_path)
