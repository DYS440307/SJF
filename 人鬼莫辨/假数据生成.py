import openpyxl
import random

def swap_values(ws, col1, col2, row_start, row_end):
    # 交换两列指定行范围内的数值
    for row in range(row_start, row_end + 1):
        temp_value = ws[f'{col1}{row}'].value
        ws[f'{col1}{row}'].value = ws[f'{col2}{row}'].value
        ws[f'{col2}{row}'].value = temp_value

def generate_random_numbers(ws, column_letter, limits):
    # 读取指定列的第16行单元格的值
    cell_value = ws[f'{column_letter}16'].value
    print(f"{column_letter}16单元格的值: {cell_value}")

    # 识别出数值
    num_str = cell_value.split('±')[0].strip()
    num = float(num_str)
    print(f"识别出的数值: {num}")

    # 获取当前列的上下限
    lower_limit, upper_limit = limits[column_letter]

    # 计算范围的下限和上限
    range_lower = round(num * lower_limit, 3)
    range_upper = round(num * upper_limit, 3)
    print(f"{column_letter}列的范围下限: {range_lower}, 上限: {range_upper}")

    # 确保范围的下限小于上限
    if range_lower >= range_upper:
        print("错误：下限大于等于上限，请调整范围。")
        return

    # 生成随机数并写入到指定列的后两列
    for row in range(18, 23):  # 行号从18到22
        next_column_value = round(random.uniform(range_lower, range_upper), 3)
        current_column_value = round(random.uniform(next_column_value + 0.001, range_upper), 3)
        ws[f'{column_letter}{row}'] = current_column_value
        ws[f'{chr(ord(column_letter)+1)}{row}'] = next_column_value
        print(f"{column_letter}{row}写入的随机数: {current_column_value}, {chr(ord(column_letter)+1)}{row}写入的随机数: {next_column_value}")

    # 在保存之前交换D和E列的数值
    if column_letter == 'D':
        swap_values(ws, 'D', 'E', 18, 22)

def process_file(file_path, limits, columns):
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for column in columns:
        generate_random_numbers(ws, column, limits)

    # Extracting num from cell H16
    num_str = ws['H16'].value.split('≤')[1].strip()
    num = float(num_str)

    # H列的
    lower_limit = num * 0.121
    upper_limit = num * 0.652

    # Generate random numbers for column H and I and write to cells H18:I22
    for row in range(18, 23):
        random_num_H = round(random.uniform(lower_limit, upper_limit), 3)
        random_num_I = round(random.uniform(random_num_H, upper_limit), 3)  # Ensure I > H
        ws[f'H{row}'] = random_num_H
        ws[f'I{row}'] = random_num_I
        print(f"Row {row}: H{row} = {random_num_H}, I{row} = {random_num_I}")

    # 保存工作簿
    wb.save(file_path)
    print("操作完成，文件已保存。")
    print()  # 打印一个空行

# 手动输入文件路径
file_path = input("请输入Excel文件路径：")

# 每列的调节上下限
limits = {
    'B': (0.96, 1.05),
    'D': (0.95, 1.08),
    'F': (0.96, 1.05)
}

# 需要处理的列
columns = ['B', 'D', 'F']

# 处理文件
process_file(file_path, limits, columns)
