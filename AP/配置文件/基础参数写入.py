import openpyxl
def process_data(user_id, experiment_record_path, experiment_report_path):
    """根据用户ID，处理实验记录和实验报告"""
    column_index = 14  # 假设ID在第14列
    m_column_index = 13  # 假设M列是第13列
    i_column_index = 9  # 假设I列是第9列

    experiment_record_wb = None
    experiment_report_wb = None

    try:
        # 打开实验记录的工作簿和工作表
        experiment_record_wb = openpyxl.load_workbook(experiment_record_path)
        experiment_record_ws = experiment_record_wb.active

        # 打开实验报告的工作簿和工作表
        experiment_report_wb = openpyxl.load_workbook(experiment_report_path)
        experiment_report_ws = experiment_report_wb.active

        # 获取最大行数
        max_row = experiment_record_ws.max_row

        # 从最后一行向上检索
        match_found = False
        for row in range(max_row, 0, -1):  # 从 max_row 到第1行倒序遍历
            cell_value = experiment_record_ws.cell(row=row, column=column_index).value
            if str(cell_value) == user_id:  # 转换为字符串后比较
                match_found = True

                # 如果匹配，获取该行的其他列的值
                experiment_report_ws["G2"] = experiment_record_ws.cell(row=row, column=1).value  # 第A列 -> G2
                experiment_report_ws["B4"] = experiment_record_ws.cell(row=row, column=2).value  # 第B列 -> B4
                experiment_report_ws["D4"] = experiment_record_ws.cell(row=row, column=3).value  # 第C列 -> D4
                experiment_report_ws["B2"] = experiment_record_ws.cell(row=row, column=6).value  # 第F列 -> B2
                experiment_report_ws["H3"] = experiment_record_ws.cell(row=row, column=8).value  # 第H列 -> H3
                experiment_report_ws["B3"] = experiment_record_ws.cell(row=row, column=9).value  # 第I列 -> B3
                experiment_report_ws["H4"] = experiment_record_ws.cell(row=row, column=10).value  # 第J列 -> H4
                experiment_report_ws["J3"] = experiment_record_ws.cell(row=row, column=11).value  # 第K列 -> J3
                experiment_report_ws["L3"] = experiment_record_ws.cell(row=row, column=12).value  # 第L列 -> L3
                experiment_report_ws["L2"] = experiment_record_ws.cell(row=row, column=14).value  # 实验编号 -> L2


                # 处理I列数据并切割，写入实验报告的B3和D3单元格
                i_column_value = str(experiment_record_ws.cell(row=row, column=i_column_index).value)
                i_column_values = i_column_value.split("；")  # 切割I列的内容

                # 如果I列有多个值，按顺序写入B3, D3
                if len(i_column_values) > 0:
                    experiment_report_ws["B3"] = i_column_values[0]  # 写入B3
                if len(i_column_values) > 1:
                    experiment_report_ws["D3"] = i_column_values[1]  # 写入D3

                # 处理M列数据并切割，写入实验报告的B7, C7, D7, E7单元格
                m_column_value = str(experiment_record_ws.cell(row=row, column=m_column_index).value)
                m_column_values = m_column_value.split("；")  # 切割M列的内容

                # 如果M列有多个值，按顺序写入B7, C7, D7, E7
                for i in range(min(4, len(m_column_values))):  # 最多切割4个部分
                    experiment_report_ws[chr(66 + i) + "7"] = m_column_values[i]  # B7, C7, D7, E7

                # 处理I列和K列的拼接：取I列切割后的第1个值与K列拼接并写入D1单元格
                i_column_value_first = i_column_values[0] if len(i_column_values) > 0 else ""  # 获取I列切割后的第1个值
                k_column_value = str(experiment_record_ws.cell(row=row, column=11).value)  # 第K列
                experiment_report_ws["D1"] = i_column_value_first + k_column_value + "实验"  # 拼接后写入D1

                # 直接退出循环，因为匹配已完成
                break

        if match_found:
            # 保存实验报告的修改
            experiment_report_wb.save(experiment_report_path)
            print(f"找到匹配的ID: {user_id}，相关数据已写入实验报告。")
        else:
            print(f"未找到匹配的ID: {user_id}。")

    except Exception as e:
        print(f"发生错误: {e}")

    finally:
        # 确保在完成后关闭工作簿
        if experiment_record_wb:
            experiment_record_wb.close()
        if experiment_report_wb:
            experiment_report_wb.close()
