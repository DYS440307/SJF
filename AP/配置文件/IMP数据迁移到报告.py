from openpyxl import load_workbook
from copy import copy
import os  # 用于文件重命名

from AP.配置文件.路径配置 import IMP_path, experiment_report_path


def write_data_to_report(IMP_path, experiment_report_path):
    # 加载 IMP 数据的工作簿
    imp_workbook = load_workbook(IMP_path)

    # 定义工作表名称及目标单元格
    sheet_to_cell_mapping = {
        "Fb归纳": "B12",
        "ACR归纳": "D12",
        "SPL归纳": "F12",
        "THD归纳": "H12"
    }

    # 加载实验报告的工作簿
    report_workbook = load_workbook(experiment_report_path)

    # 获取目标工作表（假设为活动工作表）
    report_sheet = report_workbook.active

    for sheet_name, target_cell in sheet_to_cell_mapping.items():
        if sheet_name in imp_workbook.sheetnames:
            # 获取对应的工作表
            sheet = imp_workbook[sheet_name]

            # 获取该工作表的数据
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)

            # 获取目标单元格的行号和列号
            target_row = int(target_cell[1:])
            target_col = ord(target_cell[0]) - ord('A') + 1  # 转换字母列为数字列

            # 将数据写入到目标单元格区域
            for i, row in enumerate(sheet_data):
                for j, value in enumerate(row):
                    report_sheet.cell(row=target_row + i, column=target_col + j, value=value)

    # 保存实验报告
    report_workbook.save(experiment_report_path)

    print(f"所有数据已成功写入到实验报告：Fb归纳到B12，ACR归纳到D12，SPL归纳到F12，THD归纳到H12")

    # 获取重命名所需的单元格内容
    B3_value = report_sheet["B3"].value or "B3未填写"
    J3_value = report_sheet["J3"].value or "J3未填写"
    L3_value = report_sheet["L3"].value or "L3未填写"
    L2_value = report_sheet["L2"].value or "L2未填写"

    # 生成新的文件名
    new_filename = f"{B3_value}-{J3_value}-{L3_value}-{L2_value}.xlsx"
    new_filepath = os.path.join(os.path.dirname(experiment_report_path), new_filename)

    # 重命名文件
    os.rename(experiment_report_path, new_filepath)
    print(f"实验报告已重命名为：{new_filename}")


# 调用函数
# write_data_to_report(IMP_path, experiment_report_path)
