from openpyxl.reader.excel import load_workbook
from AP.配置文件.路径配置 import IMP_path

def process_spl_data(target_values):
    print("开始运行")

    def binary_search_excel(sheet, column, target):
        """
        使用二分查找法在指定列中找到最接近目标值的单元格，返回行号。
        假设列数据已经按升序排列。
        :param sheet: 工作表对象
        :param column: 列号（如 'A', 'B'）
        :param target: 目标值
        :return: 最接近目标值的行号
        """
        low, high = 1, sheet.max_row  # 从第一行到最后一行
        closest_row = low  # 初始最接近行号

        while low <= high:
            mid = (low + high) // 2
            cell_value = sheet[f"{column}{mid}"].value

            # 跳过空值
            if cell_value is None:
                low = mid + 1
                continue

            # 更新最接近行号
            if abs(cell_value - target) < abs(sheet[f"{column}{closest_row}"].value - target):
                closest_row = mid

            if cell_value < target:
                low = mid + 1
            elif cell_value > target:
                high = mid - 1
            else:
                return mid  # 精确找到目标值

        return closest_row

    # 1. 加载工作簿
    workbook = load_workbook(IMP_path)

    # 2. 操作 SPL原档 表，删除前三行
    SPL原档_sheet = workbook["SPL原档"]
    SPL原档_sheet.delete_rows(1, 3)

    # 3. 创建/获取 SPL提取 工作表
    if "SPL提取" in workbook.sheetnames:
        SPL提取_sheet = workbook["SPL提取"]
    else:
        SPL提取_sheet = workbook.create_sheet("SPL提取")  # 如果不存在则新建

    # 清空 SPL提取 工作表内容
    for row in SPL提取_sheet.iter_rows():
        for cell in row:
            cell.value = None

    # 4. 对 SPL原档 中所有奇数列和偶数列进行操作
    for col_idx in range(1, SPL原档_sheet.max_column + 1, 2):
        odd_column_letter = SPL原档_sheet.cell(1, col_idx).column_letter
        even_column_letter = SPL原档_sheet.cell(1, col_idx + 1).column_letter  # 偶数列

        # 在每列中对每个目标值进行处理
        for target_idx, target in enumerate(target_values, start=1):
            closest_row = binary_search_excel(SPL原档_sheet, odd_column_letter, target)

            if closest_row:
                even_value = SPL原档_sheet[f"{even_column_letter}{closest_row}"].value
                # 将结果写入到 SPL提取 工作表的相应单元格
                SPL提取_sheet.cell(target_idx, (col_idx + 1) // 2, value=even_value)

    # 5. 保存文件
    workbook.save(IMP_path)

    # 6. 计算每一列的平均值，并将其写入到每列的第五行
    SPL提取_sheet = workbook["SPL提取"]
    for col_idx in range(1, SPL提取_sheet.max_column + 1):
        col_values = [SPL提取_sheet.cell(row=row_idx, column=col_idx).value for row_idx in range(1, 5)]

        # 排除 None 值的情况，计算非空单元格的平均值
        non_empty_values = [value for value in col_values if value is not None]
        if non_empty_values:
            avg_value = sum(non_empty_values) / len(non_empty_values)
        else:
            avg_value = None  # 如果该列没有有效数据，设置为 None

        # 将平均值写入到第五行
        SPL提取_sheet.cell(row=5, column=col_idx, value=avg_value)

    # 7. 获取或创建 SPL归纳 工作表
    if "SPL归纳" in workbook.sheetnames:
        SPL归纳_sheet = workbook["SPL归纳"]
    else:
        SPL归纳_sheet = workbook.create_sheet("SPL归纳")

    # 清空 SPL归纳 工作表内容
    for row in SPL归纳_sheet.iter_rows():
        for cell in row:
            cell.value = None

    # 8. 获取 SPL提取 中的平均值数据（剔除空值后处理第5行的数据）
    avg_values = [
        SPL提取_sheet.cell(row=5, column=col_idx).value
        for col_idx in range(1, SPL提取_sheet.max_column + 1)
        if SPL提取_sheet.cell(row=5, column=col_idx).value is not None
    ]

    # 调试输出 avg_values
    # print("SPL提取的第5行平均值数据:", avg_values)

    # 分成前一半和后一半
    half = len(avg_values) // 2
    front_half = avg_values[:half]  # 前一半
    back_half = avg_values[half:]  # 后一半

    # 写入前一半到 SPL归纳 工作表的第一列
    for row_idx, value in enumerate(front_half, start=1):
        SPL归纳_sheet.cell(row=row_idx, column=1, value=value)

    # 写入后一半到 SPL归纳 工作表的第二列
    for row_idx, value in enumerate(back_half, start=1):
        SPL归纳_sheet.cell(row=row_idx, column=2, value=value)

    # 9. 保存文件
    workbook.save(IMP_path)
    workbook.close()
    print("SPL实验数据处理完成")

# 调用封装后的函数，传递目标值作为参数
# process_spl_data([400, 500, 600, 800])
