from openpyxl.reader.excel import load_workbook
from AP.配置文件.路径配置 import IMP_path

def process_thd_data(range_min, range_max):
    print("开始处理 THD 原档...")

    # 1. 加载工作簿
    workbook = load_workbook(IMP_path)

    # 检查工作表是否存在
    if "THD原档" not in workbook.sheetnames:
        raise ValueError("THD原档 工作表不存在，请检查输入文件！")

    # 2. 删除 THD原档 工作表的前三行
    THD_sheet = workbook["THD原档"]
    THD_sheet.delete_rows(1, 3)

    # 3. 在奇数列的数值范围 range_min ~ range_max 中，寻找偶数列的最大值，并写入到 THD提取 的第一行
    if "THD提取" in workbook.sheetnames:
        THD提取_sheet = workbook["THD提取"]  # 使用已有工作表
    else:
        THD提取_sheet = workbook.create_sheet("THD提取")  # 如果不存在则新建

    # 清空 THD提取 工作表的内容
    for row in THD提取_sheet.iter_rows():
        for cell in row:
            cell.value = None

    THD_values = []  # 用于存储最终写入的数据

    # 遍历奇数列
    for col_idx in range(1, THD_sheet.max_column + 1, 2):
        odd_column_letter = THD_sheet.cell(1, col_idx).column_letter  # 奇数列字母
        even_column_letter = THD_sheet.cell(1, col_idx + 1).column_letter  # 偶数列字母

        max_value = float('-inf')  # 初始化为负无穷
        max_row = None

        # 遍历所有行，筛选奇数列值在范围 range_min ~ range_max 的行
        for row_idx in range(1, THD_sheet.max_row + 1):
            odd_value = THD_sheet[f"{odd_column_letter}{row_idx}"].value
            if odd_value is not None and range_min <= odd_value <= range_max:
                even_value = THD_sheet[f"{even_column_letter}{row_idx}"].value
                if even_value is not None and even_value > max_value:
                    max_value = even_value
                    max_row = row_idx

        if max_row:
            THD_values.append(max_value)

    # 将结果写入到 THD提取 工作表的第一行
    for col_idx, value in enumerate(THD_values, start=1):
        THD提取_sheet.cell(row=1, column=col_idx, value=value)

    # 4. 将 THD提取 工作表中第一行的数值均分为两分，前一半写入到 THD归纳 的第一列，后一半写入到第二列
    if "THD归纳" in workbook.sheetnames:
        THD归纳_sheet = workbook["THD归纳"]  # 使用已有工作表
    else:
        THD归纳_sheet = workbook.create_sheet("THD归纳")  # 如果不存在则新建

    # 清空 THD归纳 工作表的内容
    for row in THD归纳_sheet.iter_rows():
        for cell in row:
            cell.value = None

    # 获取 THD提取 第一行的数值
    THD_first_row_values = [cell.value for cell in THD提取_sheet[1] if cell.value is not None]
    half = len(THD_first_row_values) // 2

    # 写入前半部分到第一列
    for row_idx, value in enumerate(THD_first_row_values[:half], start=1):
        THD归纳_sheet.cell(row=row_idx, column=1, value=value)

    # 写入后半部分到第二列
    for row_idx, value in enumerate(THD_first_row_values[half:], start=1):
        THD归纳_sheet.cell(row=row_idx, column=2, value=value)

    # 保存文件
    workbook.save(IMP_path)
    workbook.close()
    print("THD原档处理完成")

# 调用封装后的函数，传递 RANGE_MIN 和 RANGE_MAX 参数
# process_thd_data(300, 1500)
