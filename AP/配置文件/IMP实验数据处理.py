from openpyxl.reader.excel import load_workbook
from AP.配置文件.路径配置 import IMP_path


def process_excel(target_value, range_min, range_max, mode):
    """
    处理Excel文件的主函数，将多个操作封装为函数。
    :param target_value: 搜索最接近的目标值
    :param range_min: 奇数列数值范围的最小值
    :param range_max: 奇数列数值范围的最大值
    :param mode: 操作模式 (1: 寻找最大值；2: 寻找最小值)
    """
    print("开始工作")

    def binary_search_excel(sheet, column, target):
        low, high = 1, sheet.max_row  # 从第一行到最后一行
        closest_row = low  # 初始最接近行号

        while low <= high:
            mid = (low + high) // 2
            cell_value = sheet[f"{column}{mid}"].value

            if cell_value is None:
                low = mid + 1
                continue

            if abs(cell_value - target) < abs(sheet[f"{column}{closest_row}"].value - target):
                closest_row = mid

            if cell_value < target:
                low = mid + 1
            elif cell_value > target:
                high = mid - 1
            else:
                return mid  # 精确找到目标值

        return closest_row

    # 1. 加载工作簿
    workbook = load_workbook(IMP_path)

    # 2. 操作 IMP原档 表，删除前三行
    IMP_sheet = workbook["IMP原档"]
    IMP_sheet.delete_rows(1, 3)

    # 3. 遍历奇数列，搜索最接近目标值的值并写入到 ACR 工作表
    ACR_sheet = workbook["ACR"]
    output_row = 1

    for col_idx in range(1, IMP_sheet.max_column + 1, 2):
        column_letter = IMP_sheet.cell(1, col_idx).column_letter
        closest_row = binary_search_excel(IMP_sheet, column_letter, target_value)

        if closest_row:
            even_col_letter = IMP_sheet.cell(1, col_idx + 1).column_letter
            closest_value = IMP_sheet[f"{even_col_letter}{closest_row}"].value
            ACR_sheet.cell(output_row, (col_idx + 1) // 2, value=closest_value)

    # 4. ACR归纳表操作
    if "ACR归纳" in workbook.sheetnames:
        ACR归纳_sheet = workbook["ACR归纳"]
    else:
        ACR归纳_sheet = workbook.create_sheet("ACR归纳")

    for row in ACR归纳_sheet.iter_rows():
        for cell in row:
            cell.value = None

    first_row_values = [cell.value for cell in ACR_sheet[1] if cell.value is not None]
    half = len(first_row_values) // 2

    for row_idx, value in enumerate(first_row_values[:half], start=1):
        ACR归纳_sheet.cell(row=row_idx, column=1, value=value)

    for row_idx, value in enumerate(first_row_values[half:], start=1):
        ACR归纳_sheet.cell(row=row_idx, column=2, value=value)

    # 5. Fb表操作
    if "Fb" in workbook.sheetnames:
        Fb_sheet = workbook["Fb"]
    else:
        Fb_sheet = workbook.create_sheet("Fb")

    Fb_values = []

    for col_idx in range(1, IMP_sheet.max_column + 1, 2):
        odd_column_letter = IMP_sheet.cell(1, col_idx).column_letter
        even_column_letter = IMP_sheet.cell(1, col_idx + 1).column_letter

        if mode == 1:
            target_value = float('-inf')
        elif mode == 2:
            target_value = float('inf')
        else:
            raise ValueError("MODE 值无效，仅支持 1 或 2。")

        target_row = None

        for row_idx in range(1, IMP_sheet.max_row + 1):
            odd_value = IMP_sheet[f"{odd_column_letter}{row_idx}"].value
            if odd_value is not None and range_min <= odd_value <= range_max:
                even_value = IMP_sheet[f"{even_column_letter}{row_idx}"].value
                if even_value is not None:
                    if (mode == 1 and even_value > target_value) or (mode == 2 and even_value < target_value):
                        target_value = even_value
                        target_row = row_idx

        if target_row:
            corresponding_value = IMP_sheet[f"{odd_column_letter}{target_row}"].value
            Fb_values.append(corresponding_value)

    for col_idx, value in enumerate(Fb_values, start=1):
        Fb_sheet.cell(row=1, column=col_idx, value=value)

    # 6. Fb归纳表操作
    if "Fb归纳" in workbook.sheetnames:
        Fb归纳_sheet = workbook["Fb归纳"]
    else:
        Fb归纳_sheet = workbook.create_sheet("Fb归纳")

    for row in Fb归纳_sheet.iter_rows():
        for cell in row:
            cell.value = None

    Fb_first_row_values = [cell.value for cell in Fb_sheet[1] if cell.value is not None]
    Fb_half = len(Fb_first_row_values) // 2

    for row_idx, value in enumerate(Fb_first_row_values[:Fb_half], start=1):
        Fb归纳_sheet.cell(row=row_idx, column=1, value=value)

    for row_idx, value in enumerate(Fb_first_row_values[Fb_half:], start=1):
        Fb归纳_sheet.cell(row=row_idx, column=2, value=value)

    workbook.save(IMP_path)
    print("处理IMP完成")
    workbook.close()


# 示例调用
# process_excel(target_value=1000, range_min=200, range_max=400, mode=1)

