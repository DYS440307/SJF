import openpyxl
import os
import re

# ================ 可配置参数 ================
process_all_months = False  # True=处理12个月份，False=处理单个月份
target_month_num = 11  # 单个月份模式时生效（1-12）
should_merge_cells = False  # 是否合并第二列相同单元格
delete_zero_rows = False  # 是否删除第四列值为0的行（注意：默认用 hide 模式）
delete_mode = 'hide'  # 'hide' 或 'openpyxl'
create_backup = False  # 是否创建备份文件

# 文件路径设置
file1_path = r"E:\System\desktop\PY\SQE\关系梳理\声乐（惠州）品控履历表_IQC检验记录（量产）.xlsx"
file2_path = r"E:\System\desktop\PY\SQE\声乐QCDS综合评分表_优化 - 副本.xlsx"
# ================================================================


def clean_text(text):
    """清洗文本：去除空格、特殊字符并统一为小写，增强匹配度"""
    if not text:
        return ""
    text_str = str(text).strip()
    text_str = re.sub(r'\s+', '', text_str)  # 去除所有空格
    text_str = re.sub(r'[^\w\u4e00-\u9fa5]', '', text_str)  # 保留字母、数字和中文
    return text_str.lower()  # 统一小写，忽略大小写差异


def merge_same_cells(worksheet, column):
    """合并指定列中连续相同的单元格（原样保留你的实现）"""
    if worksheet.max_row < 2:
        return
    start_row = 6
    current_value = worksheet.cell(row=start_row, column=column).value
    for row in range(start_row + 1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=column).value
        if cell_value != current_value:
            if row - 1 > start_row:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column
                )
            start_row = row
            current_value = cell_value
    if worksheet.max_row > start_row:
        worksheet.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=worksheet.max_row,
            end_column=column
        )


def delete_rows_with_zero(worksheet, mode='hide'):
    """
    删除或隐藏第四列（D列）中值为0的行（从第6行开始）
    mode:
        - 'hide'（默认）: 将匹配行设置为隐藏，不改变行号，能保证公式不被移动/错乱（推荐）
        - 'openpyxl'      : 使用 openpyxl 的 delete_rows 逐行删除（会改变行号，可能导致引用需额外处理）
    返回: (deleted_count, deleted_rows_list)
    """
    if worksheet.max_row < 6:
        return 0, []

    rows_to_process = []
    # 收集要处理的行（自上而下收集，后面根据 mode 决定处理顺序）
    for r in range(6, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=r, column=4).value
        try:
            if cell_value is not None and float(cell_value) == 0:
                rows_to_process.append(r)
        except (ValueError, TypeError):
            continue

    if not rows_to_process:
        return 0, []

    deleted_rows = []
    if mode == 'hide':
        # 隐藏行（安全，不改变其他单元格的引用）
        for r in rows_to_process:
            # 只隐藏行，不改动单元格内容与公式
            worksheet.row_dimensions[r].hidden = True
            deleted_rows.append(r)
        return len(deleted_rows), deleted_rows

    elif mode == 'openpyxl':
        # 物理删除行 —— 必须从下往上删除以避免索引错位
        for r in sorted(rows_to_process, reverse=True):
            worksheet.delete_rows(r, 1)
            deleted_rows.append(r)
        return len(deleted_rows), deleted_rows

    else:
        raise ValueError("不支持的 mode，选择 'hide' 或 'openpyxl'。")


def process_month(ws1, ws2, month_num, month_name, should_merge, delete_zero,
                  file1_suppliers, file2_suppliers):
    """处理单个月份的数据，收集供应商信息"""
    total_count = {}
    ng_count = {}

    # 处理file1数据，收集供应商
    for row in ws1.iter_rows(min_row=2, values_only=True):
        date, supplier, part, status = row[:4]
        if not supplier or not part:
            continue
        is_target_month = False
        if isinstance(date, str):
            if month_name in date:
                is_target_month = True
        elif hasattr(date, "month"):
            if date.month == month_num:
                is_target_month = True
        if not is_target_month:
            continue
        cleaned_supplier = clean_text(supplier)
        original_supplier = str(supplier).strip()
        file1_suppliers.add((cleaned_supplier, original_supplier))
        cleaned_part = clean_text(part)
        key = (cleaned_supplier, cleaned_part)
        total_count[key] = total_count.get(key, 0) + 1
        if status and str(status).strip().lower() == "ng":
            ng_count[key] = ng_count.get(key, 0) + 1

    # 处理file2数据，写入 total / ng
    for row_num in range(6, ws2.max_row + 1):
        part = ws2.cell(row=row_num, column=2).value
        supplier = ws2.cell(row=row_num, column=3).value
        if not supplier or not part:
            continue
        cleaned_supplier = clean_text(supplier)
        file2_suppliers.add(cleaned_supplier)
        cleaned_part = clean_text(part)
        key = (cleaned_supplier, cleaned_part)

        total_cell = ws2.cell(row=row_num, column=4)
        total_cell.value = total_count.get(key, 0)

        ng_cell = ws2.cell(row=row_num, column=5)
        ng_cell.value = ng_count.get(key, 0)

    # 执行删除/隐藏零值行操作（**不对名为 "汇总" 的 sheet 生效**）
    deleted_count = 0
    deleted_rows = []
    if delete_zero and ws2.title != "汇总":
        deleted_count, deleted_rows = delete_rows_with_zero(ws2, mode=delete_mode)

    # 执行合并单元格操作
    if should_merge:
        merge_same_cells(ws2, 2)

    return True


# ================== 主程序 ==================
if __name__ == "__main__":
    if not os.path.exists(file1_path):
        raise FileNotFoundError(f"文件1不存在：{file1_path}")
    if not os.path.exists(file2_path):
        raise FileNotFoundError(f"文件2不存在：{file2_path}")

    if create_backup:
        backup_path = os.path.splitext(file2_path)[0] + "_backup.xlsx"
        try:
            import shutil
            shutil.copy2(file2_path, backup_path)
        except Exception as e:
            print(f"创建备份文件失败：{str(e)}")

    file1_suppliers = set()
    file2_suppliers = set()

    try:
        wb1 = openpyxl.load_workbook(file1_path, read_only=True, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=False, keep_vba=True)
        ws1 = wb1.active

        all_months = [(i, f"{i}月") for i in range(1, 13)]
        target_month_name = f"{target_month_num}月"

        if process_all_months:
            for month_num, month_name in all_months:
                if month_name in wb2.sheetnames:
                    ws2 = wb2[month_name]
                    process_month(ws1, ws2, month_num, month_name,
                                  should_merge_cells, delete_zero_rows,
                                  file1_suppliers, file2_suppliers)
        else:
            if target_month_name in wb2.sheetnames:
                ws2 = wb2[target_month_name]
                process_month(ws1, ws2, target_month_num, target_month_name,
                              should_merge_cells, delete_zero_rows,
                              file1_suppliers, file2_suppliers)

        wb1.close()
        wb2.save(file2_path)
        wb2.close()

        missing_suppliers = set()
        for cleaned, original in file1_suppliers:
            if cleaned not in file2_suppliers:
                missing_suppliers.add(original)

        if missing_suppliers:
            print("file1中存在但file2中不存在的供应商：")
            for supplier in sorted(missing_suppliers):
                print(f"- {supplier}")
        else:
            print("file1中的所有供应商在file2中都存在")

    except PermissionError:
        print("文件可能被其他程序占用，请关闭后再试")
    except Exception as e:
        print(f"处理时发生错误：{str(e)}")
