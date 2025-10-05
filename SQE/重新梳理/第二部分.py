import openpyxl
import os
import unicodedata
import re

# --------------------------
# 配置参数
# --------------------------
# 要处理的Excel文件路径
EXCEL_PATH = r"E:\System\desktop\PY\SQE\关系梳理\2_惠州声乐品质履历_IQC检验记录汇总 - 副本.xlsx"

# C列关键字段替换规则（全字段匹配，仅当内容与关键词完全一致时替换）
REPLACEMENT_RULES = [
    ("上壳端子加工", "上壳"), ("L-箱壳组", "箱壳"), ("L下壳组", "箱壳"),
    ("R下壳组", "箱壳"), ("R-下壳组", "箱壳"), ("R-箱壳组", "箱壳"),
    ("上壳组件", "箱壳"), ("上壳组", "箱壳"), ("下壳组件", "下壳"),
    ("盆架组", "盆架组件"), ("箱壳组件", "箱壳"), ("上壳", "箱壳"),
    ("下壳", "箱壳"), ("减震绵", "减震棉"), ("吸音绵", "吸音棉"),
    ("尾数箱", "纸箱"), ("外箱", "纸箱"), ("鼓纸组件", "鼓纸"),
    ("海绵", "减震棉"), ("内盒", "纸箱"), ("PCB", "端子板"),
    ("盖板", "纸箱"), ("音膜组件", "音膜"), ("音膜支架组件", "音膜"),
    ("盆架组件", "盆架"), ("散件成品（橡胶圈）", "橡胶圈"), ("底板", "纸箱"),
    ("低音面罩", "面罩"), ("EVA", "减震棉"), ("面盖", "面罩"),
    ("平卡", "纸箱"), ("钕铁硼", "磁铁"), ("磁钢", "磁铁"),
    ("PCB板", "端子板"), ("音膜组", "音膜"), ("压线卡", "纸箱"),
    ("刀卡", "纸箱"), ("高音面板", "箱壳"), ("连接线", "电线"),
    ("面罩", "箱壳"), ("高音面罩", "箱壳"), ("珍珠棉刀卡", "纸箱"),
    ("防尘网", "防尘帽"), ("吸音棉", "减震棉"), ("海绵圈", "减震棉"),
    ("支架组件", "箱壳"), ("高音支架", "箱壳"), ("支架", "箱壳"),
    ("啤卡", "纸箱")
]

# C列模糊匹配替换规则（只要单元格包含关键词就替换）
FUZZY_REPLACEMENT_RULES = [
    ("刀卡", "纸箱"), ("CD纹", "防尘帽")  # 只要包含关键词就替换
]

# 要在C列中匹配并删除的关键词（包含匹配）
DELETE_C_KEYWORDS = [
    "鼓纸胶", "RA溶剂", "去渍水", "双组份中心胶", "双组份内磁磁路胶",
    "天那水", "干燥剂", "弹波胶", "模组", "八字胶", "出线孔胶",
    "双组份外磁磁路胶", "无源音箱", "无铅焊锡丝", "全音扬声器",
    "粘异物胶", "防尘帽胶", "磁液", "酒精", "锦丝线固定胶", "保鲜膜",
    "低音扬声器", "塑料袋", "调音纸", "贴纸", "纸垫圈", "保护膜",
    "套管", "PP垫圈", "高音扬声器", "810", "pp垫", "中心胶", "磁路胶",
    "胶水", "热缩管", "号角组件"
]

# C列自定义排序顺序（按此列表顺序排列）
C_COLUMN_SORT_ORDER = [
    "防尘帽", "音圈", "鼓纸", "弹波", "T铁", "U铁", "磁铁",
    "华司", "盆架", "端子板", "锦丝线", "电线", "减震棉",
    "箱壳", "螺丝", "橡胶圈", "纸箱"
]

# 其他配置
MAX_REPLACE_ROUNDS = 2  # 多次替换的最大轮数


def clean_text(text):
    """增强版文本清理函数，处理各种空格和格式问题"""
    if text is None:
        return ""

    # 转换为字符串并标准化 Unicode 字符（处理全角/半角等）
    str_text = str(text)
    normalized = unicodedata.normalize('NFKC', str_text)

    # 移除所有类型的空格（包括普通空格、不间断空格、全角空格等）
    cleaned = re.sub(r'\s+', '', normalized)

    # 移除特殊控制字符
    cleaned = re.sub(r'[\u0000-\u001F\u007F-\u009F\u2000-\u200F\u2028-\u202F\u205F-\u206F\uFEFF]', '', cleaned)

    # 转换为小写以提高匹配一致性
    return cleaned.lower()


def preprocess_rules(rules):
    """预处理规则列表，清理关键词格式，确保匹配准确性"""
    return [(clean_text(k), clean_text(v)) for k, v in rules if clean_text(k)]


def delete_c_rows_by_keyword(sheet):
    """删除C列中包含指定关键词的行（高效版）"""
    if not DELETE_C_KEYWORDS:
        return 0

    # 预处理删除关键词，统一格式
    cleaned_keywords = [clean_text(kw) for kw in DELETE_C_KEYWORDS]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # 收集需要保留的行数据
    rows_to_keep = []

    for row in range(1, max_row + 1):
        cell_value = sheet[f'C{row}'].value
        cell_clean = clean_text(cell_value)

        # 检查是否包含任何关键词（包含匹配）
        contains_keyword = any(kw in cell_clean for kw in cleaned_keywords)

        # 如果不包含关键词，则保留此行
        if not contains_keyword:
            row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
            rows_to_keep.append(row_data)

    # 计算删除的行数
    deleted_count = max_row - len(rows_to_keep)

    # 清空工作表
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).value = None

    # 写入保留的行数据
    for new_row, row_data in enumerate(rows_to_keep, start=1):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=new_row, column=col).value = value

        # 显示进度
        if new_row % 5000 == 0:
            print(f"已保留 {new_row}/{len(rows_to_keep)} 行数据")

    return deleted_count


def sort_c_column(sheet):
    """按自定义顺序对C列进行排序"""
    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_row <= 1:
        return 0  # 没有数据或只有表头，无需排序

    # 创建排序优先级字典，清理排序关键词格式
    sort_priority = {clean_text(item): idx for idx, item in enumerate(C_COLUMN_SORT_ORDER)}
    default_priority = len(C_COLUMN_SORT_ORDER)  # 未在排序列表中的项的优先级

    # 收集所有行数据及排序键
    rows_data = []
    for row in range(1, max_row + 1):
        c_value = sheet[f'C{row}'].value
        c_clean = clean_text(c_value)

        # 获取当前行的排序优先级
        priority = sort_priority.get(c_clean, default_priority)
        # 排序键：(优先级, 清理后的值)，确保相同优先级的按值排序
        sort_key = (priority, c_clean)

        # 保存整行数据
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        rows_data.append((sort_key, row_data))

    # 按自定义顺序排序
    rows_data.sort(key=lambda x: x[0])

    # 清空工作表
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).value = None

    # 写入排序后的数据
    for new_row, (_, row_data) in enumerate(rows_data, start=1):
        for col, value in enumerate(row_data, start=1):
            sheet.cell(row=new_row, column=col).value = value

        # 显示进度
        if new_row % 5000 == 0:
            print(f"已写入 {new_row} 行排序后的数据")

    return len(rows_data)


def process_c_column(file_path):
    """处理C列完整流程：清理→替换→删除→排序"""
    try:
        if not os.path.exists(file_path):
            print(f"错误: 文件 '{file_path}' 不存在")
            return

        print("开始加载Excel文件...")
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        sheet = workbook.active
        print("Excel文件加载完成")

        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"原始数据：共 {max_row} 行，{max_col} 列\n")

        # 预处理所有规则（清理格式，确保匹配准确性）
        cleaned_rules = preprocess_rules(REPLACEMENT_RULES)
        cleaned_fuzzy_rules = preprocess_rules(FUZZY_REPLACEMENT_RULES)

        # --------------------------
        # 步骤1：C列模糊匹配替换（包含关键词即替换）
        # --------------------------
        print("===== 步骤1：C列模糊匹配替换 =====")
        fuzzy_replaced_count = 0

        for row in range(1, max_row + 1):
            cell = sheet[f'C{row}']
            original_value = cell.value

            if original_value is not None:
                # 先清理单元格内容格式
                cleaned_value = clean_text(str(original_value))
                initial_value = cleaned_value

                # 应用模糊替换
                for key, value in cleaned_fuzzy_rules:
                    if key in cleaned_value:  # 模糊匹配：包含关键词即替换
                        cleaned_value = value
                        fuzzy_replaced_count += 1
                        break  # 每个单元格只应用一次模糊替换

                # 如果有变化，更新单元格值
                if cleaned_value != initial_value:
                    cell.value = cleaned_value

            # 显示进度
            if row % 5000 == 0:
                print(f"已处理 {row}/{max_row} 行")

        print(f"模糊替换完成：共替换 {fuzzy_replaced_count} 处\n")

        # --------------------------
        # 步骤2：C列全字段匹配替换
        # --------------------------
        print("===== 步骤2：C列全字段匹配替换 =====")
        exact_replaced_count = 0
        max_row_after_fuzzy = sheet.max_row  # 可能在之前步骤中发生了变化

        for row in range(1, max_row_after_fuzzy + 1):
            cell = sheet[f'C{row}']
            original_value = cell.value

            if original_value is not None:
                # 再次清理，确保格式统一
                cleaned_value = clean_text(str(original_value))
                initial_value = cleaned_value

                # 应用多次全字段替换
                for _ in range(MAX_REPLACE_ROUNDS):
                    changed = False
                    for key, value in cleaned_rules:
                        if cleaned_value == key:  # 全字段精确匹配
                            cleaned_value = value
                            exact_replaced_count += 1
                            changed = True
                            break
                    if not changed:
                        break

                # 如果有变化，更新单元格值
                if cleaned_value != initial_value:
                    cell.value = cleaned_value

            # 显示进度
            if row % 5000 == 0:
                print(f"已处理 {row}/{max_row_after_fuzzy} 行")

        print(f"全字段替换完成：共替换 {exact_replaced_count} 处\n")

        # --------------------------
        # 步骤3：删除C列包含指定关键词的行
        # --------------------------
        print(f"===== 步骤3：删除C列包含指定关键词的行 =====")
        deleted_c = delete_c_rows_by_keyword(sheet)
        print(f"C列删除完成：共删除 {deleted_c} 行\n")
        max_row_after_delete = sheet.max_row

        # --------------------------
        # 步骤4：按自定义顺序对C列进行排序
        # --------------------------
        print("===== 步骤4：按自定义顺序对C列进行排序 =====")
        if max_row_after_delete > 0:
            sorted_rows = sort_c_column(sheet)
            print(f"C列排序完成：共 {sorted_rows} 行数据按指定顺序排列")
            print(f"排序顺序：{', '.join(C_COLUMN_SORT_ORDER)}，其他项排在最后\n")
        else:
            print("无数据可排序，跳过排序步骤\n")

        # 保存文件
        print("\n正在保存文件...")
        workbook.save(file_path)
        print(f"所有处理完成，已覆盖原文件: {file_path}")
        print(f"最终统计：模糊替换 {fuzzy_replaced_count} 处，全字段替换 {exact_replaced_count} 处，"
              f"删除 {deleted_c} 行，排序后剩余 {sheet.max_row} 行")

    except Exception as e:
        print(f"\n处理错误: {str(e)}")


if __name__ == "__main__":
    process_c_column(EXCEL_PATH)
