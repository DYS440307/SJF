import openpyxl
import os
import unicodedata
import re
from difflib import SequenceMatcher

# ==============================================================================
# 配置区域
# ==============================================================================
ABBR_EXCEL_PATH = r"E:\System\desktop\PY\SQE\关系梳理\2_惠州声乐品质履历_IQC检验记录汇总 - 副本.xlsx"
FULLNAME_EXCEL_PATH = r"E:\System\desktop\PY\SQE\关系梳理\1_采购入库单 - 副本.xlsx"

ABBR_COLUMN = 2
FULLNAME_COLUMN = 1
FULLNAME_TARGET_COLUMN = None  # None 表示覆盖原列

MIN_MATCH_RATIO = 0.5  # 仅在 fallback 使用
MIN_CHINESE_MATCH_LEN = 2  # 至少匹配连续2个汉字

REPLACE_RULES = {"韵锦": "昀锦",
                 "锦韵": "昀锦",
                 "林飞宇": "林菲宇",
                 "超了": "超乐",
                 "新升": "兴升",
                 "丰晟": "佳晟",
                 "鑫和盛": "和鑫盛",
                 "大连": "大联",
                 "铂韵": "珀韵",
                 "宴钦": "彦钦",
                 "林非宇": "林菲宇",
                 "旭音": "旭声",
                 "陈意达": "诚意达",
                 "滨澄": "宾澄"



                 }
DELETE_CONDITIONS = {
    "exact_match": ["科技"],
    "contains": ["佳音达", "美佳", "强翔", "泓达","宏飛洋","宇创","三台","深辉","鸿宇","/","环盛"]
}

SHOW_MATCH_SUGGESTIONS = True


# ==============================================================================
# 工具函数
# ==============================================================================
def clean_text(text):
    """只保留中文字符"""
    if text is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(text))
    return "".join(re.findall(r"[\u4e00-\u9fff]", normalized))


# ==============================================================================
# 匹配逻辑（核心改进）
# ==============================================================================
def find_best_match(abbreviation, fullname_list, min_ratio):
    abbr_clean = clean_text(abbreviation)
    if not abbr_clean:
        return None, 0

    # 第一优先：如果全称中包含任意 ≥2 连续汉字的片段
    for length in range(len(abbr_clean), MIN_CHINESE_MATCH_LEN - 1, -1):
        for start in range(0, len(abbr_clean) - length + 1):
            sub = abbr_clean[start:start + length]
            for full_clean, full_original in fullname_list:
                if sub in full_clean:
                    return full_original, 1.0  # 满分匹配

    # 第二优先：用 difflib 做模糊匹配兜底
    best_match, best_ratio = None, 0
    for full_clean, full_original in fullname_list:
        ratio = SequenceMatcher(None, abbr_clean, full_clean).ratio()
        if ratio > best_ratio and ratio >= min_ratio:
            best_match, best_ratio = full_original, ratio

    return best_match, best_ratio


# ==============================================================================
# 文件预处理（保持不变）
# ==============================================================================
def preprocess_abbreviation_file(workbook):
    sheet = workbook.active
    max_row = sheet.max_row
    max_col = sheet.max_column

    print("\n===== 文本替换 =====")
    replace_count = 0
    for row in range(1, max_row + 1):
        cell = sheet.cell(row=row, column=ABBR_COLUMN)
        if cell.value:
            value = str(cell.value)
            for old, new in REPLACE_RULES.items():
                if old in value:
                    value = value.replace(old, new)
                    replace_count += 1
            cell.value = value
    print(f"替换完成: {replace_count}")

    print("\n===== 删除指定行 =====")
    rows_to_keep = []
    for row in range(1, max_row + 1):
        val = sheet.cell(row=row, column=ABBR_COLUMN).value or ""
        keep = True
        for kw in DELETE_CONDITIONS["exact_match"]:
            if val.strip() == kw:
                keep = False
                break
        for kw in DELETE_CONDITIONS["contains"]:
            if kw in val:
                keep = False
                break
        if keep:
            rows_to_keep.append([sheet.cell(row=row, column=c).value for c in range(1, max_col + 1)])

    # 重写
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            sheet.cell(r, column=c).value = None
    for i, row_data in enumerate(rows_to_keep, start=1):
        for c, v in enumerate(row_data, start=1):
            sheet.cell(i, column=c).value = v

    print(f"删除完成: {max_row - len(rows_to_keep)} 行")
    return len(rows_to_keep)


# ==============================================================================
# 获取Excel唯一值
# ==============================================================================
def get_unique_values(file_path, column_index):
    if not os.path.exists(file_path):
        print(f"错误: 文件 '{file_path}' 不存在")
        return []
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    sheet = workbook.active
    values = set()
    for row in sheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index, values_only=True):
        val = clean_text(row[0])
        if val:
            values.add((val, row[0]))
    workbook.close()
    return list(values)


# ==============================================================================
# 主函数
# ==============================================================================
def complete_abbreviations():
    workbook = openpyxl.load_workbook(ABBR_EXCEL_PATH)
    sheet = workbook.active
    print(f"打开文件: {ABBR_EXCEL_PATH}")
    max_row = preprocess_abbreviation_file(workbook)
    fullname_list = get_unique_values(FULLNAME_EXCEL_PATH, FULLNAME_COLUMN)
    print(f"读取全称 {len(fullname_list)} 条")

    completed = 0
    for i in range(1, max_row + 1):
        abbr = sheet.cell(row=i, column=ABBR_COLUMN).value
        if not abbr:
            continue
        match, ratio = find_best_match(abbr, fullname_list, MIN_MATCH_RATIO)
        if match:
            sheet.cell(row=i, column=FULLNAME_TARGET_COLUMN or ABBR_COLUMN).value = match
            completed += 1
        if SHOW_MATCH_SUGGESTIONS:
            print(f"行{i}: '{abbr}' → '{match or '未匹配'}' (相似度 {ratio:.2f})")

    workbook.save(ABBR_EXCEL_PATH)
    workbook.close()
    print(f"\n✅ 匹配完成！成功补全: {completed} / {max_row}")


if __name__ == "__main__":
    complete_abbreviations()
