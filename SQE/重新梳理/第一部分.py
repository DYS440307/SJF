import openpyxl
import os
import re
import unicodedata

# --------------------------
# 配置参数（在此处修改关键词和规则）
# --------------------------
# 1. 要在A列中匹配并删除的关键词（包含匹配）
DELETE_A_KEYWORDS = ["劳保"]  # 支持多个关键词，如 ["劳保", "报废"]

# 2. 要在B列中匹配并删除的关键词（包含匹配）
DELETE_B_KEYWORDS = [
    "鼓纸胶", "RA溶剂", "去渍水", "双组份中心胶", "双组份内磁磁路胶",
    "天那水", "干燥剂", "弹波胶", "模组", "八字胶", "出线孔胶",
    "双组份外磁磁路胶", "无源音箱", "无铅焊锡丝", "全音扬声器",
    "粘异物胶", "防尘帽胶", "磁液", "酒精", "锦丝线固定胶", "保鲜膜",
    "低音扬声器", "塑料袋", "调音纸", "贴纸", "纸垫圈", "保护膜",
    "套管", "PP垫圈", "高音扬声器"
]  # 可添加更多关键词

# 3. A列关键字段替换替换规则（全字段匹配，仅当内容与关键词完全一致时替换）
REPLACEMENT_RULES = [
    ("上壳端子加工", "上壳"), ("L-箱壳组", "箱壳"), ("L下壳组", "箱壳"),
    ("R下壳组", "箱壳"), ("R-下壳组", "箱壳"), ("R-箱壳组", "箱壳"),
    ("上壳组件", "箱壳"), ("上壳组", "箱壳"), ("下壳组件", "下壳"),
    ("盆架组", "盆架组件"), ("箱壳组件", "箱壳"), ("上壳", "箱壳"),
    ("下壳", "箱壳"), ("减震绵", "减震棉"), ("吸音绵", "吸音棉"),
    ("尾数箱", "纸箱"), ("外箱", "纸箱"), ("鼓纸组件", "鼓纸"),
    ("海绵", "减震棉"), ("内盒", "纸箱"), ("PCB", "端子板"),
    ("盖板", "纸箱"), ("音膜组件", "音膜"), ("音膜支架组件", "音膜"),
    ("盆架组件", "盆架"), ("散件成品（橡胶圈）", "橡胶圈"), ("底板", "纸箱"),
    ("刀卡", "纸箱"), ("低音面罩", "面罩"), ("EVA", "减震棉"),("面盖", "面罩")
]

# 4. 其他配置
MAX_REPLACE_ROUNDS = 2  # 多次替换的最大轮数
# 多个特殊公司配置：每个公司可设置不同的允许B列值
SPECIAL_COMPANIES = [
    {
        "company": "池州赛唯特电子科技有限公司",
        "allowed_b": "箱壳"  # 该公司仅保留B列为"箱壳"的行
    }
]

# B列自定义排序顺序（按此列表顺序排列）
B_COLUMN_SORT_ORDER = [
    "防尘帽", "音圈", "鼓纸", "弹波", "T铁", "U铁", "磁铁",
    "华司", "盆架", "端子板", "锦丝线", "电线", "减震棉",
    "箱壳", "螺丝", "橡胶圈", "纸箱"
]


def clean_text(text):
    """增强版文本清理函数，移除不可见字符并标准化"""
    if text is None:
        return ""

    str_text = str(text)
    normalized = unicodedata.normalize('NFKC', str_text)  # 处理全角/半角
    cleaned = re.sub(
        r'[\s\u0000-\u001F\u007F-\u009F\u2000-\u200F\u2028-\u202F\u205F-\u206F\uFEFF]',
        '',
        normalized
    )
    return cleaned.lower()


def delete_rows_by_keyword(sheet, column, keywords):
    """通用函数：删除指定列中包含任何关键词的行（包含匹配）"""
    if not keywords:
        return 0

    cleaned_keywords = [clean_text(kw) for kw in keywords]
    max_row = sheet.max_row
    rows_to_delete = []

    for row in range(1, max_row + 1):
        cell_value = sheet[f'{column}{row}'].value
        cell_clean = clean_text(cell_value)

        for kw in cleaned_keywords:
            if kw in cell_clean:
                rows_to_delete.append(row)
                break

    rows_to_delete.sort(reverse=True)
    deleted_count = 0

    for row in rows_to_delete:
        if row <= sheet.max_row:
            sheet.delete_rows(row)
            deleted_count += 1

    return deleted_count


def process_excel_columns(file_path):
    """完整处理流程：替换→填充→删除→多特殊公司处理→去重→按自定义顺序分类"""
    try:
        if not os.path.exists(file_path):
            print(f"错误: 文件 '{file_path}' 不存在")
            return

        print("开始加载Excel文件...")
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        sheet = workbook.active
        print("Excel文件加载完成")

        max_row = sheet.max_row
        max_col = sheet.max_column
        print(f"原始数据：共 {max_row} 行，{max_col} 列\n")

        # --------------------------
        # 步骤1：B列字段替换（全字段匹配）
        # --------------------------
        print("===== 步骤1：B列字段替换 =====")
        cleaned_rules = [(clean_text(k), clean_text(v)) for k, v in REPLACEMENT_RULES if clean_text(k)]
        replaced_count = 0

        for row in range(1, max_row + 1):
            cell = sheet[f'B{row}']
            original_value = cell.value

            if original_value is not None:
                cleaned_value = clean_text(str(original_value).strip())
                initial_value = cleaned_value

                for _ in range(MAX_REPLACE_ROUNDS):
                    changed = False
                    for key, value in cleaned_rules:
                        if cleaned_value == key:
                            cleaned_value = value
                            replaced_count += 1
                            changed = True
                            break
                    if not changed:
                        break

                if cleaned_value != initial_value:
                    cell.value = cleaned_value

            if row % 5000 == 0:
                print(f"已处理 {row}/{max_row} 行")

        print(f"替换完成：共替换 {replaced_count} 处\n")

        # --------------------------
        # 步骤2：A列填充
        # --------------------------
        print("===== 步骤2：A列填充 =====")
        current_a_value = None
        for row in range(1, max_row + 1):
            cell_value = sheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                current_a_value = cell_value
            elif current_a_value is not None:
                sheet[f'A{row}'].value = current_a_value
        print("A列填充完成\n")

        # --------------------------
        # 步骤3：删除A列包含指定关键词的行
        # --------------------------
        print(f"===== 步骤3：删除A列包含{DELETE_A_KEYWORDS}的行 =====")
        deleted_a = delete_rows_by_keyword(sheet, "A", DELETE_A_KEYWORDS)
        print(f"A列删除完成：共删除 {deleted_a} 行\n")
        max_row = sheet.max_row

        # --------------------------
        # 步骤4：删除B列包含指定关键词的行
        # --------------------------
        print(f"===== 步骤4：删除B列包含{DELETE_B_KEYWORDS[:5]}等关键词的行 =====")
        deleted_b = delete_rows_by_keyword(sheet, "B", DELETE_B_KEYWORDS)
        print(f"B列删除完成：共删除 {deleted_b} 行\n")
        max_row = sheet.max_row

        # --------------------------
        # 步骤5：处理多个特殊公司的行
        # 逻辑：A列匹配公司名称且B列不等于允许值 → 删除
        # --------------------------
        print("===== 步骤5：处理多个特殊公司的行 =====")
        if not SPECIAL_COMPANIES:
            print("无特殊公司配置，跳过此步骤\n")
        else:
            # 预处理特殊公司信息（清洗文本，提高匹配效率）
            processed_companies = [
                {
                    "company_clean": clean_text(company["company"]),
                    "allowed_b_clean": clean_text(company["allowed_b"]),
                    "company_original": company["company"]  # 保留原始名称用于显示
                }
                for company in SPECIAL_COMPANIES
            ]

            rows_to_delete = []
            # 遍历所有行检查特殊公司规则
            for row in range(1, max_row + 1):
                a_val = sheet[f'A{row}'].value
                b_val = sheet[f'B{row}'].value
                a_clean = clean_text(a_val)
                b_clean = clean_text(b_val)

                # 检查是否匹配任何特殊公司
                for comp in processed_companies:
                    if a_clean == comp["company_clean"]:
                        # 匹配到公司，检查B列是否为允许值
                        if b_clean != comp["allowed_b_clean"]:
                            rows_to_delete.append(row)
                            break  # 匹配到一个公司即停止检查其他公司

            # 执行删除（按行号降序）
            rows_to_delete.sort(reverse=True)
            deleted_special = 0
            for row in rows_to_delete:
                if row <= sheet.max_row:
                    sheet.delete_rows(row)
                    deleted_special += 1

            # 显示配置的特殊公司清单
            print("特殊公司配置清单：")
            for comp in SPECIAL_COMPANIES:
                print(f"- {comp['company']}：仅保留B列='{comp['allowed_b']}'的行")
            print(f"特殊公司行处理完成：共删除 {deleted_special} 行不符合条件的记录\n")
            max_row = sheet.max_row

        # --------------------------
        # 步骤6：A/B列组合去重
        # --------------------------
        print("===== 步骤6：A/B列组合去重 =====")
        seen_pairs = set()
        duplicate_rows = []

        for row in range(1, max_row + 1):
            a_clean = clean_text(sheet[f'A{row}'].value)
            b_clean = clean_text(sheet[f'B{row}'].value)
            pair = (a_clean, b_clean)

            if not a_clean and not b_clean:
                continue

            if pair in seen_pairs:
                duplicate_rows.append(row)
            else:
                seen_pairs.add(pair)

        # 批量删除重复行
        duplicate_rows.sort(reverse=True)
        batches = []
        if duplicate_rows:
            current_start = duplicate_rows[0]
            current_length = 1

            for row in duplicate_rows[1:]:
                if row == current_start - 1:
                    current_length += 1
                    current_start = row
                else:
                    batches.append((current_start, current_length))
                    current_start = row
                    current_length = 1
            batches.append((current_start, current_length))

        deleted_duplicate = 0
        for start_row, length in batches:
            if start_row <= sheet.max_row:
                sheet.delete_rows(start_row, length)
                deleted_duplicate += length

        print(f"去重完成：共删除 {deleted_duplicate} 行重复数据\n")
        max_row = sheet.max_row

        # --------------------------
        # 步骤7：按B列自定义顺序分类
        # --------------------------
        print("===== 步骤7：按B列自定义顺序分类 =====")
        # 创建排序优先级字典
        sort_priority = {clean_text(item): idx for idx, item in enumerate(B_COLUMN_SORT_ORDER)}
        default_priority = len(B_COLUMN_SORT_ORDER)  # 未指定项的优先级

        rows_data = []
        for row in range(1, max_row + 1):
            a_val = sheet[f'A{row}'].value
            b_val = sheet[f'B{row}'].value
            if (not a_val or str(a_val).strip() == '') and (not b_val or str(b_val).strip() == ''):
                continue

            row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
            b_clean = clean_text(b_val)
            priority = sort_priority.get(b_clean, default_priority)
            sort_key = (priority, b_clean)  # 排序键：(优先级, 清洗后的值)

            rows_data.append((sort_key, row_data))

        # 按自定义顺序排序
        rows_data.sort(key=lambda x: x[0])
        print(f"分类完成：共 {len(rows_data)} 行有效数据")
        print(f"排序顺序：{', '.join(B_COLUMN_SORT_ORDER)}，其他项排在最后")

        # 写入分类后的数据
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                sheet.cell(row=row, column=col).value = None

        for new_row, (_, row_data) in enumerate(rows_data, start=1):
            for col, value in enumerate(row_data, start=1):
                sheet.cell(row=new_row, column=col).value = value

            if new_row % 5000 == 0:
                print(f"已写入 {new_row} 行数据")

        # --------------------------
        # 保存文件
        # --------------------------
        print("\n正在保存文件...")
        workbook.save(file_path)
        print(f"所有处理完成，已覆盖原文件: {file_path}")
        print(f"最终统计：删除A列{deleted_a}行，删除B列{deleted_b}行，"
              f"删除特殊公司行{deleted_special if SPECIAL_COMPANIES else 0}行，"
              f"删除重复{deleted_duplicate}行，剩余{len(rows_data)}行")

    except Exception as e:
        print(f"\n处理错误: {str(e)}")


if __name__ == "__main__":
    EXCEL_PATH = r"E:\System\desktop\PY\SQE\关系梳理\1_采购入库单副本副本.xlsx"  # 可修改文件路径
    process_excel_columns(EXCEL_PATH)
