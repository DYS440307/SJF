import openpyxl
import os
import re

# ================ 可配置参数（放在最顶部方便修改） ================
# 1. 处理模式设置
process_all_months = False  # True=处理12个月份，False=处理单个月份
target_month_num = 1  # 单个月份模式时生效（1-12）
should_merge_cells = False  # True=合并第二列相同单元格，False=不合并

# 2. 文件路径设置
file1_path = r"E:\System\desktop\PY\SQE\关系梳理\2_惠州声乐品质履历_IQC检验记录汇总 - 副本.xlsx"
file2_path = r"E:\System\desktop\PY\SQE\声乐QCDS综合评分表_优化 - 副本.xlsx"


# ================================================================


def clean_text(text):
    """清洗文本：去除空格、特殊字符并统一为小写，增强匹配度"""
    if not text:
        return ""
    text_str = str(text).strip()
    text_str = re.sub(r'\s+', '', text_str)  # 去除所有空格
    text_str = re.sub(r'[^\w\u4e00-\u9fa5]', '', text_str)  # 保留字母、数字和中文
    return text_str.lower()  # 统一小写，忽略大小写差异


def merge_same_cells(worksheet, column):
    """合并指定列中连续相同的单元格"""
    if worksheet.max_row < 2:
        return  # 行数太少无需合并

    start_row = 6  # 从第6行开始处理
    current_value = worksheet.cell(row=start_row, column=column).value

    for row in range(start_row + 1, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=row, column=column).value

        # 如果当前值与上一个不同，检查是否需要合并
        if cell_value != current_value:
            # 只有当开始行和结束行不同时才合并
            if row - 1 > start_row:
                worksheet.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column
                )
            start_row = row
            current_value = cell_value

    # 处理最后一组连续相同的单元格
    if worksheet.max_row > start_row:
        worksheet.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=worksheet.max_row,
            end_column=column
        )


def process_month(ws1, ws2, month_num, month_name, should_merge):
    """处理单个月份的数据"""
    total_count = {}  # 存储总数量统计
    ng_count = {}  # 存储NG数量统计
    total_processed = 0
    valid_count = 0
    ng_total = 0

    print(f"\n开始处理{month_name}数据...")
    # 从第2行开始统计（跳过表头）
    for row in ws1.iter_rows(min_row=2, values_only=True):
        total_processed += 1
        # 提取前四列数据（日期、供应商、部品、状态）
        date, supplier, part, status = row[:4]

        # 跳过关键信息为空的行
        if not supplier or not part:
            continue

        # 判断是否为目标月份数据
        is_target_month = False
        if isinstance(date, str):
            if month_name in date:
                is_target_month = True
        elif hasattr(date, "month"):
            if date.month == month_num:
                is_target_month = True

        if not is_target_month:
            continue

        # 清洗文本后生成匹配键
        cleaned_supplier = clean_text(supplier)
        cleaned_part = clean_text(part)
        key = (cleaned_supplier, cleaned_part)

        # 统计总数量
        total_count[key] = total_count.get(key, 0) + 1
        valid_count += 1

        # 统计NG数量（判断状态是否为NG，不区分大小写）
        if status and str(status).strip().lower() == "ng":
            ng_count[key] = ng_count.get(key, 0) + 1
            ng_total += 1

    print(f"{month_name}处理完成：共{total_processed}行，有效数据{valid_count}行，其中NG数据{ng_total}行")
    print(f"统计到{len(total_count)}种供应商-部品组合")

    # 写入当前月份工作表
    print(f"开始写入{month_name}数据...")
    updated_rows = 0

    # 从第6行开始处理（文件2数据起始行）
    for row_num in range(6, ws2.max_row + 1):
        # 获取文件2中的部品和供应商（B列和C列）
        part = ws2.cell(row=row_num, column=2).value
        supplier = ws2.cell(row=row_num, column=3).value

        # 跳过空值行
        if not supplier or not part:
            updated_rows += 1
            continue

        # 清洗文本并生成匹配键
        cleaned_supplier = clean_text(supplier)
        cleaned_part = clean_text(part)
        key = (cleaned_supplier, cleaned_part)

        # 写入总数量（第4列）- 未匹配写入"/"
        total = total_count.get(key, "/")
        ws2.cell(row=row_num, column=4).value = total

        # 写入NG数量（第5列）- 未匹配写入"/"，0则显示0
        ng = ng_count.get(key)
        ws2.cell(row=row_num, column=5).value = ng if ng is not None else "/"

        updated_rows += 1

    # 根据开关决定是否合并单元格
    if should_merge:
        merge_same_cells(ws2, 2)
        print(f"{month_name}已合并第二列相同单元格")
    else:
        print(f"{month_name}未执行单元格合并（已关闭）")

    print(f"{month_name}处理完成：共更新{updated_rows}行数据")
    return True


# 所有月份的名称映射
all_months = [(i, f"{i}月") for i in range(1, 13)]
target_month_name = f"{target_month_num}月"

# ================== 加载文件 ==================
if not os.path.exists(file1_path):
    raise FileNotFoundError(f"文件1不存在：{file1_path}")
if not os.path.exists(file2_path):
    raise FileNotFoundError(f"文件2不存在：{file2_path}")

# 确保文件未被占用
try:
    # 只读模式加载文件1（提高性能）
    wb1 = openpyxl.load_workbook(file1_path, read_only=True, data_only=True)
    # 普通模式加载文件2（需要写入）
    wb2 = openpyxl.load_workbook(file2_path)
except PermissionError:
    raise PermissionError("文件可能被其他程序占用，请关闭后再试")

ws1 = wb1.active  # 默认激活的工作表

# ================== 处理数据 ==================
try:
    if process_all_months:
        # 处理所有12个月份
        for month_num, month_name in all_months:
            if month_name not in wb2.sheetnames:
                print(f"警告：文件2中未找到工作表{month_name}，已跳过")
                continue

            ws2 = wb2[month_name]
            process_month(ws1, ws2, month_num, month_name, should_merge_cells)
    else:
        # 处理单个月份
        if target_month_name not in wb2.sheetnames:
            raise ValueError(f"文件2中未找到工作表：{target_month_name}")

        ws2 = wb2[target_month_name]
        process_month(ws1, ws2, target_month_num, target_month_name, should_merge_cells)

    # ================== 保存文件 ==================
    wb1.close()  # 先关闭只读文件
    wb2.save(file2_path)  # 直接覆盖原文件
    wb2.close()
    print(f"\n结果已保存至原文件：{file2_path}")

except PermissionError:
    print(f"\n保存失败：文件可能被其他程序占用，请关闭后重试")
except Exception as e:
    print(f"\n处理时发生错误：{str(e)}")
