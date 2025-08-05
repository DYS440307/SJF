import openpyxl
import os


def process_suppliers(file_path):
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"❌ 错误：文件 {file_path} 不存在")
        return

    try:
        print(f"📂 开始处理文件：{file_path}")
        print(f"⏳ 正在加载工作簿...")

        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        # 获取第一个工作表
        ws = wb.active

        print(f"✅ 工作簿加载完成，正在准备处理数据")
        print(f"📊 检测到工作表包含 {ws.max_row} 行数据")

        # 记录原始行数，用于循环
        max_row = ws.max_row
        total_rows = max_row  # 保存总初始行数用于进度计算
        # 从第二行开始处理（假设第一行是标题）
        row = 2

        while row <= max_row:
            # 显示进度
            progress = (row / total_rows) * 100
            print(f"\r🔄 处理进度：{progress:.1f}%（正在处理第 {row} 行）", end="")

            # 获取子项物料编码
            item_code = ws.cell(row=row, column=1).value

            # 如果没有物料编码，跳过此行
            if not item_code:
                row += 1
                continue

            # 收集所有供应商名称（B列到G列，即列索引2到7）
            suppliers = []
            for col in range(2, 8):  # B到G列
                supplier = ws.cell(row=row, column=col).value
                if supplier and str(supplier).strip():  # 确保供应商名称不为空
                    suppliers.append(supplier)

            # 如果有多个供应商，需要插入新行
            if len(suppliers) > 1:
                print(f"\n🔍 发现物料 {item_code} 有 {len(suppliers)} 个供应商，正在拆分...")

                # 在当前行下方插入新行
                for i in range(len(suppliers) - 1):
                    ws.insert_rows(row + 1)
                    max_row += 1  # 增加总行数计数

                # 将供应商分配到各行，并设置父项
                for i, supplier in enumerate(suppliers):
                    current_row = row + i
                    # 清空当前行的所有供应商列
                    for col in range(2, 8):
                        ws.cell(row=current_row, column=col).value = None
                    # 设置当前供应商
                    ws.cell(row=current_row, column=2).value = supplier
                    # 设置父项为原始物料编码
                    ws.cell(row=current_row, column=8).value = item_code

                print(f"✅ 物料 {item_code} 拆分完成，生成了 {len(suppliers)} 行数据")
                # 移动到下一组数据
                row += len(suppliers)
            else:
                # 只有一个供应商，直接处理父项
                if suppliers:  # 如果有供应商
                    # 设置父项为原始物料编码
                    ws.cell(row=row, column=8).value = item_code
                row += 1

        # 保存修改后的文件，添加"_processed"后缀
        dir_name, file_name = os.path.split(file_path)
        name, ext = os.path.splitext(file_name)
        new_file_path = os.path.join(dir_name, f"{name}_processed{ext}")

        print(f"\n💾 正在保存处理后的文件...")
        wb.save(new_file_path)
        print(f"🎉 处理完成！文件已保存至：{new_file_path}")
        print(f"📊 处理总结：共处理 {total_rows} 行原始数据")

    except Exception as e:
        print(f"\n❌ 处理过程中发生错误：{str(e)}")


if __name__ == "__main__":
    # Excel文件路径
    excel_path = r"E:\System\download\1_merged.xlsx"
    process_suppliers(excel_path)
