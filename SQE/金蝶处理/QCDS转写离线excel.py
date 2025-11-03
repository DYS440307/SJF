import openpyxl

# Excel 文件路径
file_path = r"E:\System\download\2025年声乐QCDS综合评分表_优化.xlsx"

# 打开工作簿与第一个工作表
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 从第二行开始遍历（假设第1行为表头）
for row in range(2, ws.max_row + 1):
    # 读取组1与组2内容
    mat1 = ws.cell(row, 1).value  # 主供物料名称1
    sup1 = ws.cell(row, 2).value  # 供应商名称1
    score1 = ws.cell(row, 3).value  # 价格得分1

    mat2 = ws.cell(row, 5).value  # 主供物料名称2
    sup2 = ws.cell(row, 6).value  # 供应商名称2
    score2 = ws.cell(row, 7).value  # 价格得分2

    # 比较组1与组2的物料名称和供应商名称是否相同（非空情况下）
    if mat1 and sup1 and mat2 and sup2 and str(mat1).strip() == str(mat2).strip() and str(sup1).strip() == str(sup2).strip():
        # 映射组2的价格得分到组1
        ws.cell(row, 3).value = score2

# 保存修改
wb.save(file_path)
print("✅ 映射完成，文件已更新。")
