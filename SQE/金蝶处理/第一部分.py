import openpyxl

# 文件路径
file_path = r"E:\System\desktop\PY\图纸归档系统\BOm原档 - 副本.xlsx"

try:
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 获取当前活跃工作表
    max_row = ws.max_row  # 最大行数

    # ----------------------
    # 处理第二列（B列）：
    # 1. 用B1的值填充下方空单元格（遇到非空则更新复制源）
    # 2. 填充完成后清空B1
    # 3. 遍历所有行，若A列和B列同时有值，则清空B列的值
    # ----------------------
    first_value_col2 = ws.cell(row=1, column=2).value  # 保存B1初始值
    current_value = first_value_col2

    if current_value is not None:
        # 步骤1：填充下方空单元格
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=2)
            if cell.value is None:
                cell.value = current_value
            else:
                current_value = cell.value  # 更新复制源

        # 步骤2：清空B1
        ws.cell(row=1, column=2).value = None

    # 步骤3：检查所有行，若A列和B列同时有值，则清空B列
    for row in range(1, max_row + 1):
        col1_value = ws.cell(row=row, column=1).value  # 第一列（A列）值
        col2_value = ws.cell(row=row, column=2).value  # 第二列（B列）值
        # 若两列同时存在数值（非空），则清空第二列
        if col1_value is not None and col2_value is not None:
            ws.cell(row=row, column=2).value = None

    # ----------------------
    # 处理第5、6、7列（E、F、G列）：
    # 1. 整体向下移动一行
    # 2. 完成移动后，清空原第一行
    # ----------------------
    for col in [5, 6, 7]:
        # 向下移动一行（从最后一行向上遍历，避免覆盖）
        for row in range(max_row, 0, -1):
            ws.cell(row=row + 1, column=col).value = ws.cell(row=row, column=col).value
        # 清空原第一行
        ws.cell(row=1, column=col).value = None

    # ----------------------
    # 去重逻辑（保留首段，删除重复段）
    # ----------------------
    seen = set()
    rows_to_delete = []

    row = 1
    while row <= ws.max_row:
        a_val = ws.cell(row=row, column=1).value
        if a_val is not None:
            if a_val in seen:
                # 重复编号 -> 删除该段
                start_row = row
                end_row = row
                # 找段尾（下一个A有值之前）
                while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=1).value is None:
                    end_row += 1
                # 标记删除区间
                rows_to_delete.extend(range(start_row, end_row + 1))
                row = end_row + 1
                continue
            else:
                seen.add(a_val)
        row += 1

    # 倒序删除，防止索引错乱
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    # 保存修改
    wb.save(file_path)
    print(f"✅ 操作完成并去重，已保存：{file_path}")
    print(f"删除了 {len(rows_to_delete)} 行")

except Exception as e:
    print(f"处理过程中发生错误：{str(e)}")
