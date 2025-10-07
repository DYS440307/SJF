import openpyxl

# 向下复制单元格
def copy_cells_down(file_path, column='A'):
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    # 获取第一个工作表
    ws = wb.active

    # 记录当前需要复制的值
    current_value = None

    # 遍历所有行
    for row in range(1, ws.max_row + 1):
        cell = ws[f"{column}{row}"]
        cell_value = cell.value

        # 如果当前单元格有值，更新当前需要复制的值
        if cell_value is not None and str(cell_value).strip() != "":
            current_value = cell_value
        # 如果当前单元格为空且有可复制的值，则进行复制
        elif current_value is not None:
            cell.value = current_value

    # 保存修改后的文件，添加"已复制"后缀避免覆盖原文件
    new_file_path = file_path.replace('.xlsx', '_已复制.xlsx')
    wb.save(new_file_path)
    print(f"处理完成，文件已保存至: {new_file_path}")


if __name__ == "__main__":
    # 你的Excel文件路径
    excel_path = r"E:\System\desktop\PY\SQE\关系梳理\1_采购入库单.xlsx"
    # 处理A列
    copy_cells_down(excel_path, 'A')
