import openpyxl
import os
from tqdm import tqdm

# 文件路径 - 已更新为需要处理的文件
FILE_PATH = r"E:\System\download\物料清单（原档） - 副本.xlsx"


def process_excel_file(file_path, process_all_sheets=False):
    """
    1. 在A列有数值的单元格上方插入空白行，并将A、B、C列该单元格内容上移到新行
    2. 对A列每个有值单元格，将其到下方第一个有值A列单元格之间的空白行的F列，填充为该单元格的值

    参数:
    file_path: Excel文件路径
    process_all_sheets: 是否处理所有工作表(默认为False，只处理活动工作表)
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 确定需要处理的工作表
        sheets_to_process = wb.sheetnames if process_all_sheets else [wb.active.title]

        # 遍历处理每个工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            print(f"正在处理工作表: {sheet_name}")

            # 第一步：在A列有值的单元格上方插入空白行并上移A、B、C列内容
            print("执行第一步：插入空白行并上移A、B、C列内容...")
            max_row = ws.max_row
            with tqdm(total=max_row, desc="处理进度") as pbar:
                row = max_row
                while row >= 1:
                    a_cell = ws.cell(row=row, column=1)
                    a_value = a_cell.value

                    if a_value is not None and str(a_value).strip() != "":
                        # 在当前行上方插入空白行
                        ws.insert_rows(row)

                        # 将A、B、C列内容上移到新插入的行
                        # A列（第1列）
                        ws.cell(row=row, column=1).value = a_cell.value
                        a_cell.value = None

                        # B列（第2列）
                        b_cell = ws.cell(row=row + 1, column=2)
                        ws.cell(row=row, column=2).value = b_cell.value
                        b_cell.value = None

                        # C列（第3列）
                        c_cell = ws.cell(row=row + 1, column=3)
                        ws.cell(row=row, column=3).value = c_cell.value
                        c_cell.value = None

                    row -= 1
                    pbar.update(1)

            # 第二步：处理F列填充逻辑（对所有A列有值单元格）
            print("\n执行第二步：填充F列空白行...")
            # 收集A列所有有值单元格的行号和对应值
            target_rows = []
            new_max_row = ws.max_row  # 由于插入了行，最大行数已变化
            for row in range(1, new_max_row + 1):
                a_value = ws.cell(row=row, column=1).value
                if a_value is not None and str(a_value).strip() != "":
                    target_rows.append((row, a_value))  # 存储(行号, 值)元组

            # 对每个有值的A列单元格，处理其下方到下一个有值A列之间的F列
            with tqdm(total=len(target_rows), desc="F列填充进度") as pbar:
                for row_start, value in target_rows:
                    # 找到下方第一个有值的A列单元格行号
                    row_end = None
                    for row in range(row_start + 1, new_max_row + 1):
                        a_value = ws.cell(row=row, column=1).value
                        if a_value is not None and str(a_value).strip() != "":
                            row_end = row
                            break
                    # 如果没有找到下一个有值的行，就处理到最后一行
                    if row_end is None:
                        row_end = new_max_row + 1  # 确保range能包含最后一行

                    # 填充F列（第6列）为当前A列单元格的值
                    for row in range(row_start + 1, row_end):
                        ws.cell(row=row, column=6).value = value

                    pbar.update(1)

        # 保存修改后的文件，避免覆盖原文件
        new_file_path = file_path.replace('.xlsx', '_modified.xlsx')
        counter = 1
        while os.path.exists(new_file_path):
            new_file_path = file_path.replace('.xlsx', f'_modified_{counter}.xlsx')
            counter += 1

        wb.save(new_file_path)
        print(f"\n处理完成，文件已保存至: {new_file_path}")
        return new_file_path

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        return None


if __name__ == "__main__":
    # 调用函数
    process_excel_file(
        file_path=FILE_PATH,
        process_all_sheets=False  # 只处理活动工作表
    )
