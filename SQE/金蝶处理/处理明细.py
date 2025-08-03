import openpyxl
import os
from tqdm import tqdm


def copy_a_column_downwards(file_path, process_all_sheets=False):
    """
    对A列每个有值单元格，将其值向下复制，直到遇到下一个有值的A列单元格为止

    参数:
    file_path: Excel文件路径
    process_all_sheets: 是否处理所有工作表(默认为False，只处理活动工作表)
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 确定需要处理的工作表
        sheets_to_process = wb.sheetnames if process_all_sheets else [wb.active.title]

        # 遍历处理每个工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            print(f"正在处理工作表: {sheet_name}")

            # 收集A列所有有值单元格的行号和对应值
            target_rows = []
            max_row = ws.max_row
            for row in range(1, max_row + 1):
                a_value = ws.cell(row=row, column=1).value
                if a_value is not None and str(a_value).strip() != "":
                    target_rows.append((row, a_value))  # 存储(行号, 值)元组

            # 对每个有值的A列单元格，向下复制其值直到遇到下一个有值单元格
            with tqdm(total=len(target_rows), desc="处理进度") as pbar:
                for i, (row_start, value) in enumerate(target_rows):
                    # 确定复制的结束行（下一个有值单元格的上一行）
                    if i < len(target_rows) - 1:
                        # 下一个有值单元格的行号
                        next_row = target_rows[i + 1][0]
                        row_end = next_row - 1
                    else:
                        # 最后一个有值单元格，复制到工作表末尾
                        row_end = max_row

                    # 向下复制值（从当前行的下一行开始到结束行）
                    for row in range(row_start + 1, row_end + 1):
                        ws.cell(row=row, column=1).value = value

                    pbar.update(1)

        # 保存修改后的文件，避免覆盖原文件
        new_file_path = file_path.replace('.xlsx', '_modified.xlsx')
        counter = 1
        while os.path.exists(new_file_path):
            new_file_path = file_path.replace('.xlsx', f'_modified_{counter}.xlsx')
            counter += 1

        wb.save(new_file_path)
        print(f"\n处理完成，文件已保存至: {new_file_path}")
        return new_file_path

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        return None


if __name__ == "__main__":
    # 文件路径 - 已设置为你的Excel文件路径
    file_path = r"E:\System\download\采购入库单_2025080310381581_236281.xlsx"

    # 调用函数
    copy_a_column_downwards(
        file_path=file_path,
        process_all_sheets=False  # 只处理活动工作表
    )
