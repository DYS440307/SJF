import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

# ==============================
# 配置区域
# ==============================
excel_path = r"E:\System\download\声乐设备清单_2025年5月.xlsx"
output_folder = r"E:\System\download\图片文件夹"

# 确保输出文件夹存在
os.makedirs(output_folder, exist_ok=True)

# ==============================
# 打开Excel文件
# ==============================
wb = load_workbook(excel_path)
ws = wb.active

# 获取所有图片
for image in ws._images:
    # 图片锚点信息，确定所在单元格
    anchor = image.anchor._from
    row = anchor.row + 1  # openpyxl内部是从0开始的
    col = anchor.col + 1

    # 仅处理第2列（B列）
    if col != 2:
        continue

    # 获取序号和设备名称
    seq = ws.cell(row=row, column=1).value  # A列 序号
    name = ws.cell(row=row, column=3).value  # C列 设备名称

    # 构建文件名
    if not seq or not name:
        continue
    filename = f"{seq}_{name}.jpg"
    filepath = os.path.join(output_folder, filename)

    # 保存图片
    img_data = image._data()
    img = PILImage.open(BytesIO(img_data))
    img.convert("RGB").save(filepath, "JPEG")

    print(f"✅ 已提取图片: {filepath}")

print("\n🎯 所有图片提取完成！")
