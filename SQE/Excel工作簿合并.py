import openpyxl
from openpyxl import Workbook

# =============================
# 配置区
# =============================
src_path = r"E:\System\download\IQC样品明细.xlsx"
dst_path = r"E:\System\download\IQC样品明细_合集.xlsx"

# =============================
# 读取源文件
# =============================
wb_src = openpyxl.load_workbook(src_path, data_only=True)
sheet_names = wb_src.sheetnames

# 新建目标文件
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "合集"

# =============================
# 合并逻辑
# =============================
for name in sheet_names:
    ws = wb_src[name]
    print(f"正在处理工作表：{name}")

    # 跳过第一行（表头）
    rows = list(ws.iter_rows(values_only=True))[1:]

    for row in rows:
        if not row:
            continue

        # 去除每个单元格中的空格（前中后）
        cleaned = []
        for cell in row:
            if isinstance(cell, str):
                cleaned.append(cell.replace(" ", ""))  # 去掉所有空格
            else:
                cleaned.append(cell)

        # 跳过首列为空的行
        if cleaned[0] in (None, "", " "):
            continue

        ws_new.append(cleaned)

# =============================
# 保存结果
# =============================
wb_new.save(dst_path)
print(f"✅ 合并完成！结果已保存为：{dst_path}")
