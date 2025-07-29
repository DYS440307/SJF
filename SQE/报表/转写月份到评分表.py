import openpyxl
import os


def sync_supplier_data():
    # 定义文件路径
    source_path = r"E:\System\desktop\PY\SQE\2025年.xlsx"
    target_path = r"E:\System\desktop\PY\SQE\4月供方评价考核汇总表.xlsx"

    # 检查文件是否存在
    if not os.path.exists(source_path):
        print(f"错误：源文件不存在 - {source_path}")
        return
    if not os.path.exists(target_path):
        print(f"错误：目标文件不存在 - {target_path}")
        return

    try:
        # 打开源Excel文件
        source_wb = openpyxl.load_workbook(source_path, data_only=True)
        trend_sheet = source_wb["供应商质量表现趋势"]

        # 读取源数据
        supplier_name = trend_sheet["B3"].value  # 供应商名称
        if not supplier_name:
            print("错误：源表中B3单元格未找到供应商名称")
            return

        # 4月对应的来料总数（D是1月，E2月，F3月，G4月）
        april_total = trend_sheet["G3"].value  # D3~O3是每月来料总数，G对应4月
        # 4月对应的合格率（D6~O6是每月合格率，G对应4月）
        april_qualified_rate = trend_sheet["G6"].value

        if april_total is None:
            print("错误：未获取到4月来料总数数据")
            return

        # 处理来料总数的类型转换，避免比较错误
        try:
            april_total = float(april_total)
            if april_total.is_integer():
                april_total = int(april_total)
        except (ValueError, TypeError):
            print(f"错误：4月来料总数 '{april_total}' 不是有效的数字")
            return

        print(f"处理供应商：{supplier_name}，4月来料总数：{april_total}")

        # 打开目标Excel文件
        target_wb = openpyxl.load_workbook(target_path, data_only=False)

        # 根据来料总数判断写入位置
        if april_total <= 5:
            # 写入"月供货≤5批"表，从第7行开始
            if "月供货≤5批" not in target_wb.sheetnames:
                print("错误：目标文件中未找到'月供货≤5批'工作表")
                return

            low_volume_sheet = target_wb["月供货≤5批"]
            # 找到第7行及以后的第一个空行写入
            row = 7  # 从第7行开始
            while low_volume_sheet[f"A{row}"].value:
                row += 1
            low_volume_sheet[f"A{row}"].value = supplier_name
            print(f"供应商 {supplier_name} 已写入'月供货≤5批'表，行号：{row}")
        else:
            # 写入"绩效考核汇总表"，从第7行开始
            if "绩效考核汇总表" not in target_wb.sheetnames:
                print("错误：目标文件中未找到'绩效考核汇总表'工作表")
                return

            performance_sheet = target_wb["绩效考核汇总表"]
            # 确保从第7行开始写入
            row = 7
            while performance_sheet[f"C{row}"].value:
                row += 1
            performance_sheet[f"C{row}"].value = supplier_name
            print(f"供应商 {supplier_name} 已写入'绩效考核汇总表'的C{row}单元格")

        # 写入4月合格数据，从第7行开始
        if "绩效考核汇总表" in target_wb.sheetnames:
            performance_sheet = target_wb["绩效考核汇总表"]
            # 找到D列第7行及以后的第一个空行
            row = 7
            while performance_sheet[f"D{row}"].value:
                row += 1
            performance_sheet[f"D{row}"].value = april_qualified_rate
            print(f"4月合格数据 {april_qualified_rate} 已写入'绩效考核汇总表'的D{row}单元格")

        # 保存目标文件修改
        target_wb.save(target_path)
        print("数据同步完成")

    except Exception as e:
        print(f"处理过程中发生错误：{str(e)}")
    finally:
        # 确保工作簿关闭
        if 'source_wb' in locals():
            source_wb.close()
        if 'target_wb' in locals():
            target_wb.close()


if __name__ == "__main__":
    sync_supplier_data()
