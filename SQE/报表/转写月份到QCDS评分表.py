import openpyxl
from openpyxl.utils import get_column_letter
import logging
import re

# ==================== 月份选择配置（可自由修改）====================
# 在这里指定需要处理的月份，支持数字（如7）或中文（如"7月"、"七月"）
TARGET_MONTH = "1月"  # 可修改为其他月份，如"8"、"九月"等
# ==================================================================

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def normalize_text(text):
    """标准化文本：仅去除空白字符，保留原始字符和大小写"""
    if text is None:
        return ""
    # 转换为字符串并去除所有空白字符（空格、制表符等）
    return re.sub(r'\s+', '', str(text))


def process_suppliers(file1_path, file2_path, target_month):
    try:
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path)
        ws1 = wb1.active
        ws2 = wb2.active
        logger.info(f"成功加载文件：{file1_path} 和 {file2_path}")
        logger.info(f"当前处理的目标月份：{target_month}")

        # 定位目标月份所在列（第2行）
        target_col = None
        target_month_str = str(target_month).strip()
        # 统一月份判断标准（支持数字、"7月"、"七月"等格式）
        month_variants = {
            target_month_str,
            f"{target_month_str}月",
            str(int(target_month_str) if target_month_str.isdigit() else "")
        }

        for col in range(1, ws1.max_column + 1):
            val = ws1.cell(row=2, column=col).value
            if val is not None:
                val_str = str(val).strip()
                # 检查单元格值是否匹配目标月份的任何一种形式
                if val_str in month_variants or normalize_text(val_str) in month_variants:
                    target_col = col
                    logger.info(f"找到{target_month}所在列：{get_column_letter(target_col)}")
                    break

        if target_col is None:
            logger.error(f"未在文件1的第2行找到{target_month}所在列")
            return False

        # 预处理文件2：缓存供应商全称（标准化后和原始）与行号
        supplier_full_list = []
        for r2 in range(2, ws2.max_row + 1):
            full_name = ws2.cell(row=r2, column=3).value  # C列是全称
            if full_name:
                # 保存原始名称和标准化名称（去除空白）
                normalized_full = normalize_text(full_name)
                supplier_full_list.append((str(full_name), normalized_full, r2))
        logger.info(f"文件2共加载 {len(supplier_full_list)} 个供应商全称")

        # 遍历文件1的供应商
        row = 3  # 供应商简称从第3行（B3）开始
        processed_count = 0
        matched_count = 0

        while row <= ws1.max_row:
            # 获取供应商名称（B列）
            supplier_short = ws1.cell(row=row, column=2).value
            if not supplier_short:
                row += 4
                continue

            # 处理供应商简称
            supplier_short_str = str(supplier_short).strip()
            supplier_short_normalized = normalize_text(supplier_short_str)

            # 过滤标题行
            if supplier_short_str in ("", "供应商", "供应商简称"):
                row += 4
                continue

            processed_count += 1
            logger.debug(f"处理供应商：'{supplier_short_str}'（标准化：'{supplier_short_normalized}'，行号：{row}）")

            # 获取目标月份合格率（简称行+3行）
            pass_rate_row = row + 3
            if pass_rate_row > ws1.max_row:
                logger.warning(f"供应商 '{supplier_short_str}' 的合格率行超出范围，跳过")
                row += 4
                continue

            pass_rate_cell = ws1.cell(row=pass_rate_row, column=target_col)
            pass_rate = pass_rate_cell.value

            # 处理特殊情况
            if isinstance(pass_rate, str):
                if "本月未来料" in pass_rate:
                    logger.info(f"{supplier_short_str}：本月未来料，跳过")
                    row += 4
                    continue
                # 尝试转换带%的字符串
                try:
                    pass_rate = float(pass_rate.replace("%", "")) / 100
                except ValueError:
                    logger.warning(f"{supplier_short_str}：合格率非数字（{pass_rate}），跳过")
                    row += 4
                    continue
            elif not isinstance(pass_rate, (int, float)):
                logger.warning(f"{supplier_short_str}：合格率格式无效（{pass_rate}），跳过")
                row += 4
                continue

            value_to_write = round(pass_rate * 45, 2)

            # 匹配逻辑：先尝试标准化文本匹配，再尝试原始文本匹配
            matched = False
            # 1. 先使用去除空白后的文本进行匹配（解决空格导致的匹配失败）
            for full_name, normalized_full, r2 in supplier_full_list:
                if supplier_short_normalized in normalized_full:
                    logger.info(
                        f"匹配成功（标准化）：'{supplier_short_str}' 被包含在 '{full_name}' 中（行{r2}），写入值：{value_to_write}")
                    # 写入E列
                    target_cell = ws2.cell(row=r2, column=5)
                    # 处理合并单元格
                    merged_target = None
                    for merged_range in ws2.merged_cells.ranges:
                        if target_cell.coordinate in merged_range:
                            merged_target = ws2[merged_range.start_cell.coordinate]
                            break
                    if merged_target:
                        merged_target.value = value_to_write
                    else:
                        target_cell.value = value_to_write
                    matched = True
                    matched_count += 1
                    break  # 找到第一个匹配后退出

            # 2. 如果标准化匹配失败，尝试原始文本匹配
            if not matched:
                for full_name, _, r2 in supplier_full_list:
                    if supplier_short_str in full_name:
                        logger.info(
                            f"匹配成功（原始）：'{supplier_short_str}' 被包含在 '{full_name}' 中（行{r2}），写入值：{value_to_write}")
                        # 写入E列
                        target_cell = ws2.cell(row=r2, column=5)
                        # 处理合并单元格
                        merged_target = None
                        for merged_range in ws2.merged_cells.ranges:
                            if target_cell.coordinate in merged_range:
                                merged_target = ws2[merged_range.start_cell.coordinate]
                                break
                        if merged_target:
                            merged_target.value = value_to_write
                        else:
                            target_cell.value = value_to_write
                        matched = True
                        matched_count += 1
                        break

            if not matched:
                logger.warning(f"未匹配到供应商：'{supplier_short_str}'（标准化：'{supplier_short_normalized}'）")
                # 输出可能的相似项，帮助调试
                similar_items = []
                for full_name, normalized_full, r2 in supplier_full_list:
                    if supplier_short_normalized in normalized_full or supplier_short_str in full_name:
                        similar_items.append(f"[{r2}] {full_name}")
                if similar_items:
                    logger.warning(f"  可能的相似项：{', '.join(similar_items)}")

            row += 4  # 处理下一个供应商

        # 保存文件2
        wb2.save(file2_path)
        logger.info(f"\n处理总结：共处理 {processed_count} 个供应商，成功匹配 {matched_count} 个")
        return True

    except FileNotFoundError as e:
        logger.error(f"文件不存在：{e.filename}")
    except PermissionError:
        logger.error(f"没有权限操作文件，请关闭Excel后重试")
    except Exception as e:
        logger.error(f"处理过程中出错：{str(e)}", exc_info=True)
    finally:
        try:
            wb1.close()
            wb2.close()
        except:
            pass
    return False


if __name__ == "__main__":
    file1 = r"E:\System\desktop\PY\SQE\2025年.xlsx"
    file2 = r"E:\System\desktop\PY\SQE\声乐QCDS综合评分表 - 副本.xlsx"
    # 调用时传入目标月份
    process_suppliers(file1, file2, TARGET_MONTH)
