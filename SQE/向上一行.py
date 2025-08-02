import openpyxl
import os
from tqdm import tqdm


def insert_blank_rows_and_shift_a_only(file_path, process_all_sheets=False):
    """
    在A列有数值的单元格上方插入空白行，并仅将A列该单元格内容上移到新行
    其他列内容保持不变

    参数:
    file_path: Excel文件路径
    process_all_sheets: 是否处理所有工作表(默认为False，只处理活动工作表)
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path)

        # 确定需要处理的工作表
        sheets_to_process = wb.sheetnames if process_all_sheets else [wb.active.title]

        # 遍历处理每个工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            print(f"正在处理工作表: {sheet_name}")

            # 获取最大行数并创建进度条
            max_row = ws.max_row
            with tqdm(total=max_row, desc="处理进度") as pbar:
                # 从下往上遍历，避免插入行后影响索引
                row = max_row
                while row >= 1:
                    # 只检查A列(第1列)
                    a_cell = ws.cell(row=row, column=1)
                    a_value = a_cell.value

                    if a_value is not None and str(a_value).strip() != "":
                        # 在当前行上方插入空白行
                        ws.insert_rows(row)

                        # 仅将A列内容上移到新插入的行
                        ws.cell(row=row, column=1).value = a_value
                        # 清空原A列单元格内容
                        a_cell.value = None

                    row -= 1
                    pbar.update(1)

        # 保存修改后的文件，避免覆盖原文件
        new_file_path = file_path.replace('.xlsx', '_modified.xlsx')
        counter = 1
        while os.path.exists(new_file_path):
            new_file_path = file_path.replace('.xlsx', f'_modified_{counter}.xlsx')
            counter += 1

        wb.save(new_file_path)
        print(f"\n处理完成，文件已保存至: {new_file_path}")
        return new_file_path

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        return None


if __name__ == "__main__":
    # 文件路径 - 请根据实际情况修改
    file_path = r"E:\System\download\物料清单_2025080208224874_236281.xlsx"

    # 调用函数，修正了函数名
    insert_blank_rows_and_shift_a_only(  # 这里修正了函数名称
        file_path=file_path,
        process_all_sheets=False  # 只处理活动工作表
    )
