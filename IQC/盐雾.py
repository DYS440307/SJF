import warnings
from openpyxl import load_workbook

# 抑制 openpyxl 对图表读取时的警告
warnings.filterwarnings("ignore", ".*Unable to read chart.*", UserWarning)

# 文件路径
file1 = r"E:\System\desktop\2025年.xlsx"
file2 = r"E:\System\desktop\盐雾实验记录.xlsx"

# 要操作的工作表名称（例如"1月")
sheet_name = "1月"

# 关键词列表
keywords = ['T铁', 'U铁', '盆架', '钕铁硼', '华司']

# 加载工作簿与指定工作表
wb1 = load_workbook(filename=file1, data_only=True)
sheet1 = wb1[sheet_name]
wb2 = load_workbook(filename=file2)
sheet2 = wb2[sheet_name]

# 从目标表的第4行开始写入数据
start_row = 4
current_row = start_row

# 迁移源表前四列数据
for row in sheet1.iter_rows(min_row=2, values_only=False):
    val = row[2].value
    if val and any(kw in str(val) for kw in keywords):
        for col_idx in range(4):
            sheet2.cell(row=current_row, column=col_idx+1).value = row[col_idx].value
        current_row += 1

# 收集相同日期的合并区间
merge_groups = []
if current_row > start_row:
    grp_start = start_row
    prev_date = sheet2.cell(row=grp_start, column=1).value
    for r in range(start_row+1, current_row):
        cur_date = sheet2.cell(row=r, column=1).value
        if cur_date != prev_date:
            if grp_start < r - 1:
                merge_groups.append((grp_start, r - 1))
            grp_start = r
            prev_date = cur_date
    if grp_start < current_row - 1:
        merge_groups.append((grp_start, current_row - 1))

# 合并第1列（日期）
for start, end in merge_groups:
    sheet2.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)

# 填充所有已写入行的第5-8列数据
for r in range(start_row, current_row):
    # 仅对迁移过的行进行填充
    sheet2.cell(row=r, column=5).value = '5PCS'
    sheet2.cell(row=r, column=6).value = '无'
    sheet2.cell(row=r, column=7).value = '合格'
    sheet2.cell(row=r, column=8).value = '邓洋枢'

# 保存目标工作簿
wb2.save(filename=file2)
print(f"数据迁移完成：共迁移 {current_row - start_row} 行，合并第1列并填充第5-8列。写入到 '{sheet_name}' 表，从第{start_row}行开始。")
