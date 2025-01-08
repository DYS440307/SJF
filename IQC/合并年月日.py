from openpyxl import load_workbook
from openpyxl import Workbook

# 加载原始 Excel 文件
file_path = r'F:\system\Pictures\转中\盐雾实验记录.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# 创建一个新的工作簿用于保存去重结果
new_wb = Workbook()
new_ws = new_wb.active

# 用于存储已见的组合
seen = set()

# 遍历原始数据，从第二行开始（假设第一行是表头）
for row in ws.iter_rows(min_row=2, values_only=True):
    # 组合为一个元组（送测日期，供应商，料号）
    group = (row[0], row[1], row[2])

    # 如果这个组合没有出现过，则写入新工作簿
    if group not in seen:
        seen.add(group)
        new_ws.append(row)

# 保存去重后的数据
output_path = r'F:\system\Pictures\转中\盐雾实验记录_去重.xlsx'
new_wb.save(output_path)

print(f'去重操作完成，结果已保存到 {output_path}')
