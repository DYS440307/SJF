import openpyxl
from datetime import datetime, timedelta
# Define file paths
file1_path = 'F:/system/Desktop/2024年.xlsx'
file2_path = 'F:/system/Desktop/锦丝线实验记录.xlsx'
# 使用前注意删除不良汇总避免干扰
# Define the keywords to look for in column C
keywords = ["锦丝线"]

# Load the workbook
wb1 = openpyxl.load_workbook(file1_path)
wb2 = openpyxl.load_workbook(file2_path)

# Select the first sheet of file2
sheet2 = wb2.active

# Find the first empty row in sheet2
start_row = sheet2.max_row + 1

# Iterate through all sheets in file1
for sheet_name in wb1.sheetnames:
    sheet1 = wb1[sheet_name]

    # Iterate through rows in sheet1 to find matching keywords and non-empty D column
    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row):
        cell_c = row[2]  # Column C
        cell_d = row[3]  # Column D

        if cell_c.value in keywords and cell_d.value is not None:
            # Get the values from columns A, B, C, D
            values = [row[0].value, row[1].value, row[2].value, row[3].value]

            # Write the values into the next available row in sheet2
            for col, value in enumerate(values, start=1):
                sheet2.cell(row=start_row, column=col, value=value)

            # Write the additional values into columns E, F, G, H
            additional_values = ['5PCS', '无', '合格', '邓洋枢']
            for col, value in enumerate(additional_values, start=5):
                sheet2.cell(row=start_row, column=col, value=value)

            # Move to the next row in sheet2
            start_row += 1

# Save the updated file2 with a new name to preserve the original
updated_file2_path = 'F:/system/Desktop/锦丝线实验记录.xlsx'
wb2.save(updated_file2_path)

# Confirm completion
updated_file2_path

# 加载 Excel 工作簿
wb = openpyxl.load_workbook('F:/system/Desktop/锦丝线实验记录.xlsx')
ws = wb.active  # 默认获取活动工作表

# 创建一个字典来存储每周的行号
weeks_dict = {}

# 遍历 A 列的所有单元格
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    if isinstance(cell.value, datetime):
        # 获取该日期是一年中的第几周
        week_num = cell.value.isocalendar()[1]
        # 将行号添加到对应周的列表中
        weeks_dict.setdefault(week_num, []).append(row)

# 合并同一周的单元格，并写入下一周的星期一的日期
for week_num, rows in weeks_dict.items():
    if len(rows) > 1:
        # 合并单元格
        ws.merge_cells(start_row=rows[0], start_column=1, end_row=rows[-1], end_column=1)
        # 计算下一周的星期一的日期
        next_monday_cell_value = ws.cell(row=rows[0], column=1).value + timedelta(days=(7 - ws.cell(row=rows[0], column=1).value.weekday())) if ws.cell(row=rows[0], column=1).value else None
        if next_monday_cell_value:
            next_monday = next_monday_cell_value.replace(hour=0, minute=0, second=0, microsecond=0)
            # 直接对左上角的单元格设置值为下一周的星期一的日期
            ws.cell(row=rows[0], column=1).value = next_monday

# 保存工作簿
wb.save('F:/system/Desktop/锦丝线实验记录.xlsx')
