import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# 读取Excel文件
file_path = r'F:\system\Pictures\转中\盐雾实验记录.xlsx'
df = pd.read_excel(file_path)

# 映射列名
df.columns = ['送测日期', '第二列', '第三列', '料号'] + list(df.columns[4:])

# 去重只去掉完全相同的行，保留所有的料号行
df_unique = df.drop_duplicates()

# 保存去重后的结果到新的Excel文件
output_file_path = r'F:\system\Pictures\转中\盐雾实验记录_去重.xlsx'
df_unique.to_excel(output_file_path, index=False)

# 加载去重后的Excel文件
wb = load_workbook(output_file_path)
ws = wb.active

# 合并相同送测日期的单元格
current_date = None
start_row = 2  # 从第二行开始，因为第一行是标题
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    if cell.value != current_date:
        if current_date is not None:
            # 合并同一日期的单元格
            ws.merge_cells(start_row=start_row, end_row=row-1, start_column=1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(vertical='center')
        current_date = cell.value
        start_row = row

# 合并最后一组相同送测日期的单元格
ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)
ws.cell(row=start_row, column=1).alignment = Alignment(vertical='center')

# 保存最终结果
wb.save(output_file_path)

print("去重并合并完成，结果已保存到文件：", output_file_path)
