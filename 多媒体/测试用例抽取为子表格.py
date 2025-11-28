import openpyxl
import re
from openpyxl.styles import Font


def transfer_test_cases_in_same_file(file_path, detail_sheet_name, template_sheet_name):
    # 加载工作簿（可写模式）
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)

    # 检查基础工作表是否存在
    if detail_sheet_name not in wb.sheetnames:
        raise ValueError(f"错误：未找到「{detail_sheet_name}」工作表")
    if template_sheet_name not in wb.sheetnames:
        raise ValueError(f"错误：未找到「{template_sheet_name}」工作表")

    detail_ws = wb[detail_sheet_name]
    template_ws = wb[template_sheet_name]

    # 超链接样式（蓝色+下划线）
    hyperlink_font = Font(color="0000FF", underline="single")
    created_sheets = []  # 记录创建的工作表名称（用于校验）

    for row in range(2, detail_ws.max_row + 1):
        # 获取数据（按列索引：C=3测试项目，D=4测试条件，E=5测试步骤，F=6判定标准）
        test_item = detail_ws.cell(row=row, column=3).value  # 测试项目（C列）
        test_conditions = detail_ws.cell(row=row, column=4).value  # 测试条件（D列）
        test_steps = detail_ws.cell(row=row, column=5).value  # 测试步骤（E列）
        criteria = detail_ws.cell(row=row, column=6).value  # 判定标准（F列）

        # 跳过不完整数据
        if not all([test_item, test_conditions, test_steps, criteria]):
            print(f"跳过行{row}：数据不完整（测试项目/条件/步骤/标准缺一）")
            continue

        # 1. 生成标准化工作表名称（避免特殊字符和重复）
        raw_name = str(test_item).strip()
        valid_name = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9_ ]', '', raw_name)  # 保留安全字符
        valid_name = valid_name[:31]  # 限制长度
        # 处理重复名称
        if valid_name in created_sheets:
            count = 2
            while f"{valid_name}_{count}" in created_sheets:
                count += 1
            valid_name = f"{valid_name}_{count}"

        # 2. 复制模板并创建新工作表
        try:
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = valid_name
            if valid_name not in wb.sheetnames:
                raise Exception("创建后未在工作表列表中找到")
            created_sheets.append(valid_name)
            print(f"行{row}：创建工作表「{valid_name}」成功")
        except Exception as e:
            print(f"行{row}：创建工作表失败，错误：{e}")
            continue

        # 3. 填充新工作表内容（按新要求写入单元格）
        new_ws["B3"] = test_item  # 测试项目写入B3
        new_ws["B6"] = test_conditions  # 测试条件写入B6（新调整）
        new_ws["B7"] = test_steps  # 测试步骤写入B8（新调整）

        # 拆分判定标准到C12:C14
        criteria_list = re.split(r'[\n;]?\s*(?=\d+\.)', str(criteria))
        criteria_list = [c.strip() for c in criteria_list if c.strip()]
        for i in range(3):
            new_ws[f"C{12 + i}"] = criteria_list[i] if i < len(criteria_list) else ""

        # 4. 设置超链接（指向新工作表A1）
        try:
            hyperlink = f"#'{valid_name}'!A1"  # 强制用单引号包裹，确保匹配
            if valid_name not in wb.sheetnames:
                raise Exception("目标工作表不存在")

            cell = detail_ws.cell(row=row, column=3)
            cell.value = test_item
            cell.hyperlink = hyperlink
            cell.font = hyperlink_font
            print(f"行{row}：超链接设置成功")
        except Exception as e:
            print(f"行{row}：超链接设置失败，错误：{e}")

    # 保存文件
    try:
        wb.save(file_path)
        print("\n==== 操作完成 ====")
        print(f"文件保存至：{file_path}")
        print(f"创建工作表数量：{len(created_sheets)}")
        print("提示：请关闭所有Excel窗口后重新打开文件，点击C列蓝色链接测试跳转")
    except PermissionError:
        print("\n保存失败！请关闭所有Excel窗口后重试（文件被占用）")
    except Exception as e:
        print(f"\n保存错误：{e}")


if __name__ == "__main__":
    # 新文件路径
    EXCEL_FILE_PATH = r"E:\System\desktop\睿音_UC000_测试用例完全版(WPS可能不兼容) - 副本.xlsx"
    DETAIL_SHEET_NAME = "详细测试用例"  # 确保与工作表名称完全一致（区分大小写）
    TEMPLATE_SHEET_NAME = "报告模板"  # 同上

    transfer_test_cases_in_same_file(EXCEL_FILE_PATH, DETAIL_SHEET_NAME, TEMPLATE_SHEET_NAME)