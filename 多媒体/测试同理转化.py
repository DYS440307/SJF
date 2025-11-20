import openpyxl
import re
from openpyxl.styles import Font


def transfer_test_cases_in_same_file(file_path, detail_sheet_name, template_sheet_name):
    # 加载工作簿（确保可写模式）
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)

    # 检查基础工作表是否存在
    if detail_sheet_name not in wb.sheetnames:
        raise ValueError(f"错误：未找到「{detail_sheet_name}」工作表")
    if template_sheet_name not in wb.sheetnames:
        raise ValueError(f"错误：未找到「{template_sheet_name}」工作表")

    detail_ws = wb[detail_sheet_name]
    template_ws = wb[template_sheet_name]

    # 超链接样式
    hyperlink_font = Font(color="0000FF", underline="single")
    created_sheets = []  # 记录所有成功创建的工作表名称（精确匹配）

    for row in range(2, detail_ws.max_row + 1):
        # 获取基础数据
        test_item = detail_ws.cell(row=row, column=3).value
        test_steps = detail_ws.cell(row=row, column=5).value
        criteria = detail_ws.cell(row=row, column=6).value

        if not all([test_item, test_steps, criteria]):
            print(f"跳过行{row}：测试项目信息不完整")
            continue

        # 1. 生成严格标准化的工作表名称（确保无任何歧义）
        raw_name = str(test_item).strip()
        # 只保留字母、数字、中文、下划线和空格（移除所有其他符号）
        valid_name = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9_ ]', '', raw_name)
        # 限制长度31字符
        valid_name = valid_name[:31]
        # 处理重复名称（添加序号）
        if valid_name in created_sheets:
            count = 2
            while f"{valid_name}_{count}" in created_sheets:
                count += 1
            valid_name = f"{valid_name}_{count}"

        # 2. 复制模板并创建工作表（强制校验名称）
        try:
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = valid_name
            # 立即校验工作表是否创建成功
            if valid_name not in wb.sheetnames:
                raise Exception(f"创建后未在工作表列表中找到「{valid_name}」")
            created_sheets.append(valid_name)
            print(f"行{row}：成功创建工作表「{valid_name}」")
        except Exception as e:
            print(f"行{row}：创建工作表失败「{valid_name}」，错误：{e}")
            continue

        # 3. 填充新工作表内容
        new_ws["B3"] = test_item
        new_ws["B6"] = test_steps
        # 拆分判定标准
        criteria_list = re.split(r'[\n;]?\s*(?=\d+\.)', str(criteria))
        criteria_list = [c.strip() for c in criteria_list if c.strip()]
        for i in range(3):
            new_ws[f"C{12 + i}"] = criteria_list[i] if i < len(criteria_list) else ""

        # 4. 生成超链接（严格匹配工作表名称）
        try:
            # 超链接格式：强制用单引号包裹，确保与创建的工作表名完全一致
            hyperlink = f"#'{valid_name}'!A1"
            # 验证超链接指向的工作表是否存在
            if valid_name not in wb.sheetnames:
                raise Exception(f"超链接指向的工作表「{valid_name}」不存在")

            # 设置单元格超链接
            cell = detail_ws.cell(row=row, column=3)
            cell.value = test_item  # 显示原始名称
            cell.hyperlink = hyperlink
            cell.font = hyperlink_font
            print(f"行{row}：超链接设置成功「{hyperlink}」")
        except Exception as e:
            print(f"行{row}：超链接设置失败，错误：{e}")

    # 保存文件（关键步骤）
    try:
        wb.save(file_path)
        print("\n==== 操作完成 ====")
        print(f"文件已保存至：{file_path}")
        print(f"共创建{len(created_sheets)}个工作表：{created_sheets}")
        print("请按以下步骤测试：")
        print("1. 关闭所有Excel窗口")
        print("2. 重新双击打开该Excel文件")
        print("3. 进入「详细测试用例」工作表，点击C列蓝色链接")
    except PermissionError:
        print("\n保存失败！请关闭所有Excel窗口后重试（文件被占用）")
    except Exception as e:
        print(f"\n保存失败：{e}")


if __name__ == "__main__":
    # 请务必确认以下参数与实际Excel一致
    EXCEL_FILE_PATH = r"E:\System\desktop\睿音_UC000_测试用例完全版(WPS可能不兼容) - 副本.xlsx"
    DETAIL_SHEET_NAME = "详细测试用例"  # 工作表名称必须完全匹配（包括空格和大小写）
    TEMPLATE_SHEET_NAME = "报告模板"  # 同上

    transfer_test_cases_in_same_file(EXCEL_FILE_PATH, DETAIL_SHEET_NAME, TEMPLATE_SHEET_NAME)