from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# 提取FO
# 找到 C42:C54 中的位置
Temp_FO_1 = 30
Temp_FO_2 = 49
# 提取灵敏度
# 7: 提取Temp_Fund中第二个工作表的数据
Temp_Fund_1 = 70
Temp_Fund_2 = 73
Temp_Fund_3 = 77
Temp_Fund_4 = 82
# 提取失真
# 8: 提取Temp_THD中第二个工作表的数据
Temp_THD_1 = 37
Temp_THD_2 = 110
# 9: 提取TempFile1目录下包名字“额定功率”文件
Temp_Name = "12302-500111"
Temp_Name2 = "性能测试"
# 上下限的设置
Temp_B11_LowLim = 300 * 0.8
Temp_B11_UpLim = 300 * 1.2
Temp_C11_LowLim = 6 * 0.85
Temp_C11_UpLim = 6 * 1.15
Temp_D11_LowLim = 74
Temp_D11_UpLim = 78
Temp_E11_LowLim = 0
Temp_E11_UpLim = 10

# 1: 提取目录下所有文件
directory = "F:/system/Desktop/PY/OQC/" + Temp_Name + "/源文件"
all_files = os.listdir(directory)
TempInput = None

# 2-4: 筛选文件
Temp_Imp = [file for file in all_files if "Imp" in file]
Temp_THD = [file for file in all_files if "THD" in file]
Temp_Fund = [file for file in all_files if "Fund" in file]

# 5-6: 提取Temp_Imp中第二个工作表的数据w
for file in Temp_Imp:
    wb = load_workbook(os.path.join(directory, file))
    sheet = wb.worksheets[1]

    # 提取阻抗
    Temp_Ohms1 = sheet['C61'].value
    Temp_Ohms2 = sheet['G61'].value
    Temp_Ohms3 = sheet['K61'].value
    Temp_Ohms4 = sheet['O61'].value
    Temp_Ohms5 = sheet['S61'].value

# 提取FH
max_row_C = max(range(Temp_FO_1, Temp_FO_2), key=lambda row: sheet.cell(row=row, column=3).value)
# 让 Temp_Fh1 等于位置左边一列的数值
Temp_Fh1 = sheet.cell(row=max_row_C, column=2).value
# 找到 G42:G54 中最
max_row_G = max(range(Temp_FO_1, Temp_FO_2), key=lambda row: sheet.cell(row=row, column=7).value)
# 让 Temp_Fh2 等于位置左边一列的数值
Temp_Fh2 = sheet.cell(row=max_row_G, column=6).value
# 找到 K42:K54 中最大值的位置
max_row_K = max(range(Temp_FO_1, Temp_FO_2), key=lambda row: sheet.cell(row=row, column=11).value)
# 让 Temp_Fh3 等于位置左边一列的数值
Temp_Fh3 = sheet.cell(row=max_row_K, column=10).value
# 找到 O42:O54 中最大值的位置
max_row_O = max(range(Temp_FO_1, Temp_FO_2), key=lambda row: sheet.cell(row=row, column=15).value)
# 让 Temp_Fh4 等于位置左边一列的数值
Temp_Fh4 = sheet.cell(row=max_row_O, column=14).value
# 找到 S42:S54 的位置
max_row_S = max(range(Temp_FO_1, Temp_FO_2), key=lambda row: sheet.cell(row=row, column=19).value)
# 让 Temp_Fh5 等于位置左边一列的数值
Temp_Fh5 = sheet.cell(row=max_row_S, column=18).value
wb.close()

# 提取灵敏度
# 7: 提取Temp_Fund中第二个工作表的数据
for file in Temp_Fund:
    wb = load_workbook(os.path.join(directory, file))
    sheet = wb.worksheets[1]
    Temp_Fund1 = sum(
        sheet.cell(row=row, column=3).value for row in [Temp_Fund_1, Temp_Fund_2, Temp_Fund_3, Temp_Fund_4]) / 4
    Temp_Fund2 = sum(
        sheet.cell(row=row, column=7).value for row in [Temp_Fund_1, Temp_Fund_2, Temp_Fund_3, Temp_Fund_4]) / 4
    Temp_Fund3 = sum(
        sheet.cell(row=row, column=11).value for row in [Temp_Fund_1, Temp_Fund_2, Temp_Fund_3, Temp_Fund_4]) / 4
    Temp_Fund4 = sum(
        sheet.cell(row=row, column=15).value for row in [Temp_Fund_1, Temp_Fund_2, Temp_Fund_3, Temp_Fund_4]) / 4
    Temp_Fund5 = sum(
        sheet.cell(row=row, column=19).value for row in [Temp_Fund_1, Temp_Fund_2, Temp_Fund_3, Temp_Fund_4]) / 4
wb.close()

# 提取失真
# 8: 提取Temp_THD中第二个工作表的数据
for file in Temp_THD:
    wb = load_workbook(os.path.join(directory, file))
    sheet = wb.worksheets[1]
    Temp_THD1 = max(sheet.cell(row=row, column=3).value for row in range(Temp_THD_1, Temp_THD_2))
    Temp_THD2 = max(sheet.cell(row=row, column=7).value for row in range(Temp_THD_1, Temp_THD_2))
    Temp_THD3 = max(sheet.cell(row=row, column=11).value for row in range(Temp_THD_1, Temp_THD_2))
    Temp_THD4 = max(sheet.cell(row=row, column=15).value for row in range(Temp_THD_1, Temp_THD_2))
    Temp_THD5 = max(sheet.cell(row=row, column=19).value for row in range(Temp_THD_1, Temp_THD_2))
    wb.close()

# 9: 提取TempFile1目录下包含“Temp_Name”文件
TempFile1_directory = "F:/system/Desktop/PY/OQC/" + Temp_Name + "/TEMP"
TempFile1 = [file for file in os.listdir(TempFile1_directory) if Temp_Name2 in file]

# 10: 判断TempFile1的A1和B4单元格中的数值是否包含Temp_Name
for file in TempFile1:
    wb = load_workbook(os.path.join(TempFile1_directory, file))
    sheet = wb.active

    # 写入Fh
    # 11-14: 将数据写入TempFile1中
    sheet['B12'] = Temp_Fh1
    sheet['B13'] = Temp_Fh2
    sheet['B14'] = Temp_Fh3
    sheet['B15'] = Temp_Fh4
    sheet['B16'] = Temp_Fh5
    # 写入阻抗
    sheet['C12'] = Temp_Ohms1
    sheet['C13'] = Temp_Ohms2
    sheet['C14'] = Temp_Ohms3
    sheet['C15'] = Temp_Ohms4
    sheet['C16'] = Temp_Ohms5
    # 写入灵敏度
    sheet['D12'] = Temp_Fund1
    sheet['D13'] = Temp_Fund2
    sheet['D14'] = Temp_Fund3
    sheet['D15'] = Temp_Fund4
    sheet['D16'] = Temp_Fund5
    # 写入失真
    sheet['E12'] = Temp_THD1
    sheet['E13'] = Temp_THD2
    sheet['E14'] = Temp_THD3
    sheet['E15'] = Temp_THD4
    sheet['E16'] = Temp_THD5
    wb.save(os.path.join(TempFile1_directory, file))
    wb.close()
print('实验前数据写入完成')

# 15: NG&OK判断
for file in TempFile1:
    wb = load_workbook(os.path.join(TempFile1_directory, file))
    sheet = wb.active
    # Extract data from B20 to B24
    TempFile1_Fh_values = [sheet.cell(row=row, column=2).value for row in range(12, 16)]
    # Extract data from D20 to D24
    TempFile1_Ohms_values = [sheet.cell(row=row, column=3).value for row in range(12, 16)]
    # Extract data from F20 to F24
    TempFile1_Fund_values = [sheet.cell(row=row, column=4).value for row in range(12, 16)]
    # Extract data from H20 to H24
    TempFile1_THD_values = [sheet.cell(row=row, column=5).value for row in range(12, 16)]


    # Check conditions and write OK/NG in B11:D11:F11:H11
    def check_condition(value, low_limit, up_limit):
        return "OK" if low_limit <= value <= up_limit else "NG"


    # 上下限的设置------注意修改
    sheet['B17'] = check_condition(TempFile1_Fh_values[0], Temp_B11_LowLim, Temp_B11_UpLim)
    sheet['C17'] = check_condition(TempFile1_Ohms_values[0], Temp_C11_LowLim, Temp_C11_UpLim)
    sheet['D17'] = check_condition(TempFile1_Fund_values[0], Temp_D11_LowLim, Temp_D11_UpLim)
    sheet['E17'] = check_condition(TempFile1_THD_values[0], Temp_E11_LowLim, Temp_E11_UpLim)
    print('数据结果判定完成')

    # 定义条件，假设条件是B17到E17单元格内容包含"NG"
    target_range = sheet['B17:E17']

    # 设置红色、加粗样式
    red_bold_font = Font(color='FF0000', bold=True)

    # 遍历目标范围
    for row in target_range:
        for cell in row:
            # 检查单元格内容是否包含"NG"
            if "NG" in str(cell.value):
                # 设置红色、加粗样式
                cell.font = red_bold_font

    wb.save(os.path.join(TempFile1_directory, file))
    wb.close()
