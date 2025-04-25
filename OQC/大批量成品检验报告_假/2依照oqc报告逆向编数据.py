import os
import openpyxl

# 源文件夹路径
source_folder = r"E:\System\desktop\PY\OQC\Input"
# 目标文件夹路径
target_folder = r"E:\System\desktop\PY\OQC\Output"
# 模板文件路径
template_file = r"E:\System\desktop\PY\OQC\转移模板.xlsx"


def read_and_write_data(source_path, target_folder):
    # 打开源文件
    source_workbook = openpyxl.load_workbook(source_path)
    source_sheet = source_workbook.active

    # 打开模板文件
    template_workbook = openpyxl.load_workbook(template_file)
    template_sheet = template_workbook.active

    # 读取并写入数据
    template_sheet['B4'] = source_sheet['M4'].value
    template_sheet['H4'] = source_sheet['C5'].value
    template_sheet['E3'] = source_sheet['R3'].value
    template_sheet['E4'] = source_sheet['H4'].value  # 将源文件的C5单元格数据写入到模板文件的E4单元格
    template_sheet['B11'] = source_sheet['F19'].value
    template_sheet['C11'] = source_sheet['F21'].value
    template_sheet['D11'] = source_sheet['F22'].value
    template_sheet['H3'] = source_sheet['C3'].value
    template_sheet['E11'] = source_sheet['F23'].value
    for i in range(5):
        template_sheet.cell(row=12 + i, column=2).value = source_sheet.cell(row=19, column=10 + i).value
        template_sheet.cell(row=12 + i, column=3).value = source_sheet.cell(row=21, column=10 + i).value
        template_sheet.cell(row=12 + i, column=4).value = source_sheet.cell(row=22, column=10 + i).value

    # 保存并重命名文件
    target_filename = f"{template_sheet['H3'].value}_{template_sheet['B4'].value}.xlsx"
    target_path = os.path.join(target_folder, target_filename)

    # 确保目标文件夹及其所有父级文件夹都存在
    os.makedirs(os.path.dirname(target_path), exist_ok=True)

    template_workbook.save(target_path)

    # 关闭工作簿
    source_workbook.close()
    template_workbook.close()


# 确保目标文件夹存在
if not os.path.exists(target_folder):
    os.makedirs(target_folder)

# 处理源文件夹下的每个文件
for filename in os.listdir(source_folder):
    source_path = os.path.join(source_folder, filename)

    # 只处理xlsx文件
    if filename.endswith(".xlsx"):
        read_and_write_data(source_path, target_folder)

print("处理完成！")
